import sys
import os
import json
from pathlib import Path
import pandas as pd
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configure Tesseract path for Windows
try:
    # Try common Windows installation paths
    possible_paths = [
        r'C:\Program Files\Tesseract-OCR\tesseract.exe',
        r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
        r'C:\Users\{}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'.format(os.getenv('USERNAME', '')),
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            pytesseract.pytesseract.tesseract_cmd = path
            break
    else:
        # If not found in common paths, try to find it in PATH
        import shutil
        tesseract_path = shutil.which('tesseract')
        if tesseract_path:
            pytesseract.pytesseract.tesseract_cmd = tesseract_path
except Exception as e:
    print(f"Warning: Could not configure Tesseract path: {e}")
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
import barcode
from barcode.writer import ImageWriter
import hashlib
import requests
from datetime import datetime, timedelta
import serial
import time
import subprocess
import tempfile
import win32print
import win32api

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QGridLayout, QLabel, QPushButton, QTextEdit, QLineEdit, 
    QFileDialog, QMessageBox, QProgressBar, QStatusBar, QFrame,
    QScrollArea, QGroupBox, QSplitter, QComboBox, QDialog, 
    QDialogButtonBox, QListWidget, QTableWidget, QTableWidgetItem,
    QHeaderView, QPlainTextEdit, QCheckBox, QTabWidget, QDateEdit,
    QStyledItemDelegate, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem,
    QGraphicsRectItem
)
from PySide6.QtCore import Qt, QThread, Signal, QTimer, QSize, QDate, QRectF, QPointF
from PySide6.QtGui import QFont, QPalette, QColor, QIcon, QPixmap, QPen, QBrush, QPainter

# Import Supabase configuration
try:
    from supabase_config import save_generated_barcodes, upload_store_orders_from_excel, get_supabase_client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False
    print("Warning: Supabase configuration not available. Some features may be disabled.")


class ProcessingThread(QThread):
    """Background thread for PDF processing operations"""
    progress_signal = Signal(str)
    finished_signal = Signal(bool, dict)
    
    def __init__(self, app_instance):
        super().__init__()
        self.app = app_instance
    
    def run(self):
        try:
            result = self.app.process_picking_dockets_internal()
            self.finished_signal.emit(True, result)
        except Exception as e:
            self.progress_signal.emit(f"Error: {str(e)}")
            self.finished_signal.emit(False, {"error": str(e)})





class ProcessingResultsDialog(QDialog):
    """Professional dialog for displaying processing results"""
    
    def __init__(self, results, parent=None):
        super().__init__(parent)
        self.results = results
        self.setWindowTitle("Processing Results")
        self.setModal(True)
        self.resize(900, 700)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header_layout = QHBoxLayout()
        
        # Status icon
        status_icon = "✓"
        status_color = "#10b981"
        
        if results.get('driver_files_created', 0) == 0:
            status_icon = "⚠"
            status_color = "#f59e0b"
        elif results.get('failed_files'):
            status_icon = "⚠"
            status_color = "#f59e0b"
        
        status_label = QLabel(status_icon)
        status_label.setStyleSheet(f"font-size: 32px; color: {status_color}; font-weight: bold;")
        header_layout.addWidget(status_label)
        
        # Title
        title_text = "Picking Dockets Processing Completed Successfully"
        if results.get('driver_files_created', 0) == 0:
            title_text = "Picking Dockets Processing Completed - No Matching Orders Found"
        elif results.get('failed_files'):
            title_text = "Picking Dockets Processing Completed with Some Issues"
        
        title_label = QLabel(title_text)
        title_label.setObjectName("resultTitle")
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        layout.addLayout(header_layout)
        
        # Summary statistics
        stats_frame = QFrame()
        stats_frame.setObjectName("statsFrame")
        stats_layout = QGridLayout(stats_frame)
        
        # Stats
        stats_data = [
            ("Picking PDF Files Processed", str(results.get('processed_files', 0))),
            ("Pages Scanned", str(results.get('total_pages', 0))),
            ("Driver Picking PDFs Created", str(results.get('driver_files_created', 0))),
            ("Barcodes Generated", str(results.get('barcodes_generated', 0))),
            ("Failed Files", str(len(results.get('failed_files', []))))
        ]
        
        for i, (label, value) in enumerate(stats_data):
            label_widget = QLabel(label + ":")
            label_widget.setObjectName("statsLabel")
            value_widget = QLabel(value)
            value_widget.setObjectName("statsValue")
            
            stats_layout.addWidget(label_widget, i, 0)
            stats_layout.addWidget(value_widget, i, 1)
        
        layout.addWidget(stats_frame)
        
        # Create tabbed interface
        tab_widget = QTabWidget()
        
        # Tab 1: Created Files
        if results.get('created_files'):
            files_tab = self.create_files_tab(results.get('created_files', []))
            tab_widget.addTab(files_tab, "Created Files")
        

        # Next: Driver Details
        if results.get('driver_details'):
            driver_tab = self.create_driver_tab(results.get('driver_details', {}))
            tab_widget.addTab(driver_tab, "Driver Details")
        
        # Tab 3: Barcode Status (if there are any issues)
        if results.get('barcode_generation_errors') or results.get('order_numbers_not_found'):
            barcode_status_tab = self.create_barcode_status_tab(results)
            tab_widget.addTab(barcode_status_tab, "Barcode Status")
        
        # Tab 4: Failed Files (if any)
        if results.get('failed_files'):
            failed_tab = self.create_failed_tab(results.get('failed_files', []))
            tab_widget.addTab(failed_tab, "Failed Files")
        
        # If no tabs were created, show diagnostic information
        if tab_widget.count() == 0:
            if results.get('error') == "No matching orders found in picking docket PDF files":
                diagnostic_tab = self.create_diagnostic_tab(results)
                tab_widget.addTab(diagnostic_tab, "Diagnostic Info")
            else:
                empty_tab = self.create_empty_tab()
                tab_widget.addTab(empty_tab, "Results")
        
        layout.addWidget(tab_widget)
        
        # Output directory info
        output_frame = QFrame()
        output_frame.setObjectName("outputFrame")
        output_layout = QVBoxLayout(output_frame)
        
        output_label = QLabel("Output Directory:")
        output_label.setObjectName("outputLabel")
        output_layout.addWidget(output_label)
        
        output_path = QLabel(results.get('output_dir', ''))
        output_path.setObjectName("outputPath")
        output_path.setWordWrap(True)
        output_layout.addWidget(output_path)
        
        layout.addWidget(output_frame)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        open_folder_btn = QPushButton("Open Output Folder")
        open_folder_btn.setObjectName("primaryButton")
        open_folder_btn.clicked.connect(self.open_output_folder)
        
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        
        button_layout.addWidget(open_folder_btn)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        
        layout.addLayout(button_layout)
        
        # Apply styling
        self.apply_results_styling()
    
    def create_files_tab(self, files):
        """Create tab showing created files"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["File Name", "Status"])
        table.setRowCount(len(files))
        
        for i, filename in enumerate(files):
            # File name
            name_item = QTableWidgetItem(filename)
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(i, 0, name_item)
            
            # Status
            status_item = QTableWidgetItem("✓ Created")
            status_item.setFlags(status_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(i, 1, status_item)
        
        table.resizeColumnsToContents()
        table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(table)
        
        return widget

    
    def create_driver_tab(self, driver_details):
        """Create tab showing driver details"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["Driver", "Total Pages", "Orders"])
        table.setRowCount(len(driver_details))
        
        for i, (driver, details) in enumerate(driver_details.items()):
            # Driver
            driver_item = QTableWidgetItem(f"Driver {driver}")
            driver_item.setFlags(driver_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(i, 0, driver_item)
            
            # Total Pages
            total_pages = details.get('page_count', 0)
            pages_item = QTableWidgetItem(str(total_pages))
            pages_item.setFlags(pages_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(i, 1, pages_item)
            
            # Orders
            orders = details.get('orders', [])
            orders_text = ", ".join(orders[:3])
            if len(orders) > 3:
                orders_text += f" ... (+{len(orders)-3} more)"
            orders_item = QTableWidgetItem(orders_text)
            orders_item.setFlags(orders_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(i, 2, orders_item)
        
        table.resizeColumnsToContents()
        table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(table)
        
        return widget
    
    def create_failed_tab(self, failed_files):
        """Create tab showing failed files"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["File Name", "Status"])
        table.setRowCount(len(failed_files))
        
        for i, filename in enumerate(failed_files):
            # File name
            name_item = QTableWidgetItem(filename)
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(i, 0, name_item)
            
            # Status
            status_item = QTableWidgetItem("✗ Failed")
            status_item.setFlags(status_item.flags() & ~Qt.ItemIsEditable)
            table.setItem(i, 1, status_item)
        
        table.resizeColumnsToContents()
        table.horizontalHeader().setStretchLastSection(True)
        layout.addWidget(table)
        
        return widget
    
    def create_empty_tab(self):
        """Create empty tab for when no results are available"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Center the message
        layout.addStretch()
        
        message_label = QLabel("No results to display")
        message_label.setAlignment(Qt.AlignCenter)
        message_label.setStyleSheet("color: #6b7280; font-size: 14px; font-style: italic;")
        layout.addWidget(message_label)
        
        layout.addStretch()
        
        return widget
    
    def create_barcode_status_tab(self, results):
        """Create tab showing barcode generation and order number matching status"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Summary section
        summary_label = QLabel("Barcode Generation and Order Number Status")
        summary_label.setObjectName("sectionTitle")
        layout.addWidget(summary_label)
        
        # Statistics
        stats_text = QLabel(f"""
        <b>Summary:</b>
        <br>• Barcodes Generated: {len(results.get('order_numbers_found_in_pdfs', []))}
        <br>• Barcode Generation Failures: {len(results.get('barcode_generation_errors', {}))}
        <br>• Order Numbers Found in PDFs: {len(results.get('order_numbers_found_in_pdfs', []))}
        <br>• Order Numbers Not Found in PDFs: {len(results.get('order_numbers_not_found', []))}
        """)
        stats_text.setObjectName("infoText")
        stats_text.setWordWrap(True)
        layout.addWidget(stats_text)
        
        # Barcode generation errors
        if results.get('barcode_generation_errors'):
            errors_label = QLabel("❌ Barcode Generation Failures:")
            errors_label.setObjectName("warningText")
            layout.addWidget(errors_label)
            
            errors_table = QTableWidget()
            errors_table.setColumnCount(2)
            errors_table.setHorizontalHeaderLabels(["Order Number", "Error Reason"])
            errors_table.setRowCount(len(results.get('barcode_generation_errors', {})))
            
            for i, (order_id, error) in enumerate(results.get('barcode_generation_errors', {}).items()):
                # Order number
                order_item = QTableWidgetItem(str(order_id))
                order_item.setFlags(order_item.flags() & ~Qt.ItemIsEditable)
                errors_table.setItem(i, 0, order_item)
                
                # Error reason
                error_item = QTableWidgetItem(error)
                error_item.setFlags(error_item.flags() & ~Qt.ItemIsEditable)
                errors_table.setItem(i, 1, error_item)
            
            errors_table.resizeColumnsToContents()
            errors_table.horizontalHeader().setStretchLastSection(True)
            layout.addWidget(errors_table)
        
        # Order numbers not found in PDFs
        if results.get('order_numbers_not_found'):
            not_found_label = QLabel("❌ Order Numbers Not Found in PDF Files:")
            not_found_label.setObjectName("warningText")
            layout.addWidget(not_found_label)
            
            not_found_table = QTableWidget()
            not_found_table.setColumnCount(2)
            not_found_table.setHorizontalHeaderLabels(["Order Number", "Status"])
            not_found_table.setRowCount(len(results.get('order_numbers_not_found', [])))
            
            for i, order_id in enumerate(results.get('order_numbers_not_found', [])):
                # Order number
                order_item = QTableWidgetItem(str(order_id))
                order_item.setFlags(order_item.flags() & ~Qt.ItemIsEditable)
                not_found_table.setItem(i, 0, order_item)
                
                # Status
                status_item = QTableWidgetItem("No matching pages found in any PDF file")
                status_item.setFlags(status_item.flags() & ~Qt.ItemIsEditable)
                not_found_table.setItem(i, 1, status_item)
            
            not_found_table.resizeColumnsToContents()
            not_found_table.horizontalHeader().setStretchLastSection(True)
            layout.addWidget(not_found_table)
        
        # Successfully processed order numbers
        if results.get('order_numbers_found_in_pdfs'):
            success_label = QLabel("✅ Successfully Processed Order Numbers:")
            success_label.setObjectName("successText")
            layout.addWidget(success_label)
            
            success_table = QTableWidget()
            success_table.setColumnCount(2)
            success_table.setHorizontalHeaderLabels(["Order Number", "Pages Found"])
            success_table.setRowCount(len(results.get('order_numbers_found_in_pdfs', [])))
            
            for i, order_id in enumerate(results.get('order_numbers_found_in_pdfs', [])):
                # Order number
                order_item = QTableWidgetItem(str(order_id))
                order_item.setFlags(order_item.flags() & ~Qt.ItemIsEditable)
                success_table.setItem(i, 0, order_item)
                
                # Pages found
                page_count = len(results.get('driver_details', {}).get(order_id, {}).get('orders', []))
                pages_item = QTableWidgetItem(str(page_count))
                pages_item.setFlags(pages_item.flags() & ~Qt.ItemIsEditable)
                success_table.setItem(i, 1, pages_item)
            
            success_table.resizeColumnsToContents()
            success_table.horizontalHeader().setStretchLastSection(True)
            layout.addWidget(success_table)
        
        layout.addStretch()
        return widget
    
    def create_diagnostic_tab(self, results):
        """Create diagnostic tab when no matching orders are found"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # Diagnostic message
        diagnostic_text = QPlainTextEdit()
        diagnostic_text.setReadOnly(True)
        diagnostic_text.setStyleSheet("font-family: 'Consolas', monospace; font-size: 12px; background-color: #f8fafc;")
        
        message = "No matching orders were found in the picking docket PDF files.\n\n"
        message += "Troubleshooting Steps:\n"
        message += "=" * 50 + "\n\n"
        message += "1. Check that your picking docket PDF files contain order IDs\n"
        message += "2. Ensure you have loaded delivery sequence data first\n"
        message += "3. Order ID matching is case-insensitive (AA061B4Y = aa061b4y)\n\n"
        message += f"Processing Summary:\n"
        message += f"- Picking PDF files processed: {results.get('processed_files', 0)}\n"
        message += f"- Total pages scanned: {results.get('total_pages', 0)}\n"
        message += f"- No matching order IDs found\n\n"
        message += "Common Issues:\n"
        message += "- Order IDs in picking PDFs don't match those in delivery data\n"
        message += "- PDF contains images that need OCR processing\n"
        message += "- No delivery sequence data loaded\n"
        message += "- Text extraction failed from PDF pages\n\n"
        message += "Note: You need delivery sequence data loaded before processing picking dockets."
        
        diagnostic_text.setPlainText(message)
        layout.addWidget(diagnostic_text)
        
        return widget
    
    def open_output_folder(self):
        """Open output folder"""
        if hasattr(self.parent(), 'open_output_directory'):
            self.parent().open_output_directory(self.results.get('output_dir', ''))
    
    def apply_results_styling(self):
        """Apply styling to results dialog"""
        self.setStyleSheet("""
            QDialog {
                background-color: #f8fafc;
            }
            
            QLabel#resultTitle {
                font-size: 18px;
                font-weight: bold;
                color: #1e293b;
                margin-bottom: 10px;
            }
            
            QFrame#statsFrame {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 15px;
            }
            
            QLabel#statsLabel {
                font-weight: 600;
                color: #374151;
                font-size: 14px;
            }
            
            QLabel#statsValue {
                font-size: 14px;
                color: #10b981;
                font-weight: bold;
            }
            
            QFrame#outputFrame {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 15px;
            }
            
            QLabel#outputLabel {
                font-weight: 600;
                color: #374151;
                font-size: 14px;
                margin-bottom: 5px;
            }
            
            QLabel#outputPath {
                color: #6b7280;
                font-size: 12px;
                font-family: 'Consolas', monospace;
            }
            
            QTabWidget::pane {
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                background-color: white;
            }
            
            QTabBar::tab {
                background-color: #f1f5f9;
                border: 1px solid #e2e8f0;
                border-bottom: none;
                border-radius: 6px 6px 0 0;
                padding: 8px 16px;
                margin-right: 2px;
                color: #64748b;
                font-weight: 500;
            }
            
            QTabBar::tab:selected {
                background-color: white;
                color: #1e293b;
                border-bottom: 2px solid #2563eb;
            }
            
            QTabBar::tab:hover {
                background-color: #e2e8f0;
                color: #374151;
            }
            
            QTableWidget {
                gridline-color: #f1f5f9;
                border: none;
                background-color: white;
            }
            
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #f1f5f9;
            }
            
            QTableWidget::item:selected {
                background-color: #eff6ff;
                color: #1e40af;
            }
            
            QTableWidget QHeaderView::section {
                background-color: #f8fafc;
                border: none;
                border-bottom: 2px solid #e2e8f0;
                padding: 10px;
                font-weight: 600;
                color: #374151;
            }
        """)


class OrderEntryDialog(QDialog):
    """Dialog for entering order data in a table format"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add New Orders")
        self.setModal(True)
        self.resize(1200, 600)
        
        # Initialize data
        self.order_data = []
        
        # Setup UI
        self.init_ui()
        self.apply_styling()
    
    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header_label = QLabel("Enter Order Details")
        header_label.setObjectName("headerTitle")
        layout.addWidget(header_label)
        
        # Instructions
        instructions = QLabel(
            "Fill in the table below with your order information. "
            "You can add multiple rows and edit any cell by double-clicking or using arrow keys to navigate."
        )
        instructions.setObjectName("infoText")
        instructions.setWordWrap(True)
        layout.addWidget(instructions)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Order Number", "Item Code", "Product Description", "Barcode", 
            "Customer Type", "Quantity", "Site Name", "Account Code", 
            "Dispatch Code", "Route"
        ])
        
        # Set initial rows
        self.table.setRowCount(5)
        
        # Initialize all cells with empty items for direct editing
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = QTableWidgetItem("")
                self.table.setItem(row, col, item)
        
        # Make table editable and enable keyboard navigation
        self.table.setEditTriggers(QTableWidget.DoubleClicked | QTableWidget.EditKeyPressed | QTableWidget.AnyKeyPressed)
        
        # Enable inline editing (no popup text box)
        self.table.setItemDelegate(QStyledItemDelegate())
        
        # Enable keyboard navigation
        self.table.setTabKeyNavigation(True)
        self.table.setFocusPolicy(Qt.StrongFocus)
        
        # Set column widths for better spacing
        column_widths = [120, 100, 200, 120, 120, 80, 120, 120, 120, 100]
        for i, width in enumerate(column_widths):
            self.table.setColumnWidth(i, width)
        
        # Set row height for better readability
        self.table.verticalHeader().setDefaultSectionSize(35)
        
        # Resize columns to content but respect minimum widths
        self.table.resizeColumnsToContents()
        for i, width in enumerate(column_widths):
            if self.table.columnWidth(i) < width:
                self.table.setColumnWidth(i, width)
        
        # Enable horizontal scrolling for smaller screens
        self.table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        
        layout.addWidget(self.table)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        add_row_btn = QPushButton("➕ Add Row")
        add_row_btn.setObjectName("secondaryButton")
        add_row_btn.clicked.connect(self.add_row)
        
        clear_btn = QPushButton("Clear All")
        clear_btn.setObjectName("secondaryButton")
        clear_btn.clicked.connect(self.clear_table)
        
        button_layout.addWidget(add_row_btn)
        button_layout.addWidget(clear_btn)
        button_layout.addStretch()
        
        # Dialog buttons
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self.button_box.accepted.connect(self.accept)
        self.button_box.rejected.connect(self.reject)
        
        button_layout.addWidget(self.button_box)
        layout.addLayout(button_layout)
        
        # Connect keyboard events for better navigation
        self.table.keyPressEvent = self.table_key_press_event
    
    def table_key_press_event(self, event):
        """Handle keyboard navigation and editing"""
        if event.key() in [Qt.Key_Return, Qt.Key_Enter]:
            # Move to next row when Enter is pressed
            current_row = self.table.currentRow()
            current_col = self.table.currentColumn()
            
            if current_row < self.table.rowCount() - 1:
                self.table.setCurrentCell(current_row + 1, current_col)
            else:
                # If at last row, add a new row and move there
                self.add_row()
                self.table.setCurrentCell(self.table.rowCount() - 1, current_col)
            
            # Start editing the new cell
            self.table.editItem(self.table.item(self.table.currentRow(), self.table.currentColumn()))
        elif event.key() == Qt.Key_Tab:
            # Handle Tab navigation
            current_row = self.table.currentRow()
            current_col = self.table.currentColumn()
            
            if current_col < self.table.columnCount() - 1:
                # Move to next column
                self.table.setCurrentCell(current_row, current_col + 1)
            else:
                # Move to first column of next row
                if current_row < self.table.rowCount() - 1:
                    self.table.setCurrentCell(current_row + 1, 0)
                else:
                    # If at last row, add a new row
                    self.add_row()
                    self.table.setCurrentCell(self.table.rowCount() - 1, 0)
            
            # Start editing the new cell
            self.table.editItem(self.table.item(self.table.currentRow(), self.table.currentColumn()))
        elif event.key() == Qt.Key_Backtab:
            # Handle Shift+Tab navigation
            current_row = self.table.currentRow()
            current_col = self.table.currentColumn()
            
            if current_col > 0:
                # Move to previous column
                self.table.setCurrentCell(current_row, current_col - 1)
            else:
                # Move to last column of previous row
                if current_row > 0:
                    self.table.setCurrentCell(current_row - 1, self.table.columnCount() - 1)
            
            # Start editing the new cell
            self.table.editItem(self.table.item(self.table.currentRow(), self.table.currentColumn()))
        else:
            # For any other key, start editing if not already editing
            if not self.table.state() == QTableWidget.EditingState:
                current_item = self.table.item(self.table.currentRow(), self.table.currentColumn())
                if not current_item:
                    current_item = QTableWidgetItem("")
                    self.table.setItem(self.table.currentRow(), self.table.currentColumn(), current_item)
                self.table.editItem(current_item)
            
            # Call the original keyPressEvent
            QTableWidget.keyPressEvent(self.table, event)
    
    def add_row(self):
        """Add a new row to the table"""
        current_rows = self.table.rowCount()
        self.table.setRowCount(current_rows + 1)
        
        # Initialize the new row with empty items for direct editing
        for col in range(self.table.columnCount()):
            item = QTableWidgetItem("")
            self.table.setItem(current_rows, col, item)
    
    def clear_table(self):
        """Clear all data from the table"""
        self.table.setRowCount(0)
        self.table.setRowCount(5)
        
        # Initialize all cells with empty items for direct editing
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = QTableWidgetItem("")
                self.table.setItem(row, col, item)
    
    def get_order_data(self):
        """Extract order data from the table"""
        data = []
        
        for row in range(self.table.rowCount()):
            row_data = {}
            has_data = False
            
            # Extract data from each column
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                value = item.text() if item else ""
                
                # Map column headers to expected field names
                if col == 0:  # Order Number
                    row_data['Order Number'] = value
                elif col == 1:  # Item Code
                    row_data['Item Code'] = value
                elif col == 2:  # Product Description
                    row_data['Product Description'] = value
                elif col == 3:  # Barcode
                    row_data['Barcode'] = value
                elif col == 4:  # Customer Type
                    row_data['Customer Type'] = value
                elif col == 5:  # Quantity
                    row_data['Quantity'] = value
                elif col == 6:  # Site Name
                    row_data['Site Name'] = value
                elif col == 7:  # Account Code
                    row_data['Account Code'] = value
                elif col == 8:  # Dispatch Code
                    row_data['Dispatch Code'] = value
                elif col == 9:  # Route
                    row_data['Route'] = value
                
                if value.strip():
                    has_data = True
            
            # Only add rows that have at least some data
            if has_data:
                data.append(row_data)
        
        return data
    
    def apply_styling(self):
        """Apply styling to the dialog"""
        self.setStyleSheet("""
            QDialog {
                background-color: #f8fafc;
            }
            
            QLabel#headerTitle {
                font-size: 18px;
                font-weight: bold;
                color: #1e293b;
                margin-bottom: 10px;
            }
            
            QLabel#infoText {
                color: #6b7280;
                font-size: 14px;
                margin-bottom: 10px;
            }
            
            QTableWidget {
                gridline-color: #e2e8f0;
                border: 2px solid #e2e8f0;
                border-radius: 8px;
                background-color: white;
                selection-background-color: #eff6ff;
                selection-color: #1e40af;
                font-size: 13px;
            }
            
            QTableWidget::item {
                padding: 8px 12px;
                border-bottom: 1px solid #f1f5f9;
                border-right: 1px solid #f1f5f9;
            }
            
            QTableWidget::item:selected {
                background-color: #eff6ff;
                color: #1e40af;
                font-weight: 500;
            }
            
            QTableWidget::item:focus {
                background-color: #fef3c7;
                border: 2px solid #f59e0b;
            }
            
            QTableWidget QHeaderView::section {
                background-color: #f8fafc;
                border: none;
                border-bottom: 2px solid #e2e8f0;
                border-right: 1px solid #e2e8f0;
                padding: 12px 8px;
                font-weight: 600;
                color: #374151;
                font-size: 13px;
            }
            
            QTableWidget QHeaderView::section:first {
                border-left: none;
            }
            
            QTableWidget QHeaderView::section:last {
                border-right: none;
            }
            
            QPushButton#secondaryButton {
                background-color: #6b7280;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-weight: 500;
                font-size: 14px;
                min-width: 100px;
            }
            
            QPushButton#secondaryButton:hover {
                background-color: #4b5563;
            }
            
            QPushButton#secondaryButton:pressed {
                background-color: #374151;
            }
            
            QDialogButtonBox QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-weight: 500;
                font-size: 14px;
                min-width: 80px;
            }
            
            QDialogButtonBox QPushButton:hover {
                background-color: #1d4ed8;
            }
            
            QDialogButtonBox QPushButton:pressed {
                background-color: #1e40af;
            }
        """)


class PDFGraphicsView(QGraphicsView):
    """Custom graphics view for PDF coordinate selection"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dialog = parent
        self.drawing = False
        self.start_point = None
        self.end_point = None
    
    def mousePressEvent(self, event):
        """Handle mouse press for starting rectangle selection"""
        if event.button() == Qt.LeftButton:
            self.drawing = True
            self.start_point = self.mapToScene(event.pos())
            self.end_point = self.start_point
            if self.parent_dialog:
                self.parent_dialog.status_label.setText("Drag to create selection rectangle")
        super().mousePressEvent(event)
    
    def mouseMoveEvent(self, event):
        """Handle mouse move for drawing rectangle"""
        if self.drawing:
            self.end_point = self.mapToScene(event.pos())
            if self.parent_dialog:
                self.parent_dialog.update_selection_rect()
        super().mouseMoveEvent(event)
    
    def mouseReleaseEvent(self, event):
        """Handle mouse release for finishing rectangle selection"""
        if event.button() == Qt.LeftButton and self.drawing:
            self.drawing = False
            self.end_point = self.mapToScene(event.pos())
            if self.parent_dialog:
                self.parent_dialog.finish_selection()
        super().mouseReleaseEvent(event)


class MultiRegionCoordinateSelectorDialog(QDialog):
    """Dialog for selecting multiple OCR regions visually on PDF"""
    
    def __init__(self, pdf_path, parent=None):
        super().__init__(parent)
        self.pdf_path = pdf_path
        self.current_region = 'region_1'  # Start with region 1
        
        # Get regions from parent application if available
        if parent and hasattr(parent, 'ocr_regions'):
            self.regions = parent.ocr_regions.copy()
        else:
            # Fallback to default regions
            self.regions = {
                'region_1': {'coordinates': None, 'color': 'red', 'name': 'Region 1'},
                'region_2': {'coordinates': None, 'color': 'blue', 'name': 'Region 2'},
                'region_3': {'coordinates': None, 'color': 'green', 'name': 'Region 3'},
                'region_4': {'coordinates': None, 'color': 'orange', 'name': 'Region 4'},
                'region_5': {'coordinates': None, 'color': 'purple', 'name': 'Region 5'}
            }
        self.drawing = False
        self.start_point = None
        self.end_point = None
        
        self.setWindowTitle("Configure Multiple OCR Regions")
        self.setModal(True)
        self.resize(1200, 900)
        
        self.init_ui()
        self.load_pdf()
    
    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout(self)
        
        # Instructions
        instructions = QLabel(
            "Configure Multiple OCR Regions\n"
            "Select 5 different regions where OCR should extract text:\n"
            "• Region 1 (Red): e.g., 'Route Cork 1'\n"
            "• Region 2 (Blue): e.g., Order number or date\n"
            "• Region 3 (Green): e.g., Customer reference\n"
            "• Region 4 (Orange): Additional data field\n"
            "• Region 5 (Purple): Additional text that will be appended to Region 3\n\n"
            "Use the region selector below to switch between regions, then click and drag to draw rectangles."
        )
        instructions.setObjectName("sectionTitle")
        instructions.setWordWrap(True)
        layout.addWidget(instructions)
        
        # Region selector
        region_layout = QHBoxLayout()
        region_layout.addWidget(QLabel("Select Region:"))
        
        self.region_combo = QComboBox()
        # Dynamically add all regions from the configuration
        for region_id, region in self.regions.items():
            self.region_combo.addItem(f"{region['name']} ({region['color'].title()})", region_id)
        self.region_combo.currentTextChanged.connect(self.on_region_changed)
        region_layout.addWidget(self.region_combo)
        
        region_layout.addStretch()
        layout.addLayout(region_layout)
        
        # Region 5 save location configuration
        self.region_5_config_group = QGroupBox("Region 5 Save Location Configuration")
        self.region_5_config_group.setVisible(False)  # Initially hidden
        region_5_layout = QHBoxLayout(self.region_5_config_group)
        
        region_5_layout.addWidget(QLabel("Save Region 5 data to:"))
        self.region_5_save_location = QLineEdit()
        self.region_5_save_location.setPlaceholderText("e.g., Column L, Custom Field, etc.")
        self.region_5_save_location.setText("Column K (Region 5 Data)")  # Default value
        region_5_layout.addWidget(self.region_5_save_location)
        
        self.browse_save_location_btn = QPushButton("Browse...")
        self.browse_save_location_btn.clicked.connect(self.browse_save_location)
        region_5_layout.addWidget(self.browse_save_location_btn)
        
        layout.addWidget(self.region_5_config_group)
        
        # Graphics view for PDF display
        self.graphics_view = PDFGraphicsView(self)
        self.graphics_view.setMouseTracking(True)
        layout.addWidget(self.graphics_view)
        
        # Scene for graphics
        self.scene = QGraphicsScene()
        self.graphics_view.setScene(self.scene)
        
        # Status label
        self.status_label = QLabel("Select Region 1 (Red) and click and drag to select the OCR region")
        self.status_label.setObjectName("infoText")
        layout.addWidget(self.status_label)
        
        # Coordinate display
        self.coord_label = QLabel("Region 1 (Red): Not selected")
        self.coord_label.setObjectName("infoText")
        layout.addWidget(self.coord_label)
        
        # Region status display
        self.region_status_label = QLabel("Region 2 (Blue): Not selected | Region 3 (Green): Not selected")
        self.region_status_label.setObjectName("infoText")
        layout.addWidget(self.region_status_label)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.clear_btn = QPushButton("Clear Current Region")
        self.clear_btn.clicked.connect(self.clear_current_region)
        self.clear_btn.setEnabled(False)
        
        self.test_btn = QPushButton("Test Current Region")
        self.test_btn.clicked.connect(self.test_current_region)
        self.test_btn.setEnabled(False)
        self.test_btn.setObjectName("secondaryButton")
        
        self.show_all_btn = QPushButton("Show All Regions")
        self.show_all_btn.clicked.connect(self.show_all_regions)
        self.show_all_btn.setEnabled(False)
        self.show_all_btn.setObjectName("secondaryButton")
        
        self.save_btn = QPushButton("Save All Regions")
        self.save_btn.clicked.connect(self.save_all_regions)
        self.save_btn.setEnabled(False)
        self.save_btn.setObjectName("primaryButton")
        
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.clear_btn)
        button_layout.addWidget(self.test_btn)
        button_layout.addWidget(self.show_all_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.cancel_btn)
        button_layout.addWidget(self.save_btn)
        
        layout.addLayout(button_layout)
    
    def load_pdf(self):
        """Load PDF and display first page"""
        try:
            # Open PDF with PyMuPDF
            pdf_document = fitz.open(self.pdf_path)
            if len(pdf_document) == 0:
                QMessageBox.warning(self, "Error", "PDF file is empty or corrupted")
                return
            
            # Get first page
            page = pdf_document[0]
            
            # Convert to image
            mat = fitz.Matrix(1.0, 1.0)  # Use 1.0x zoom to match OCR processing
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            
            # Convert to QPixmap
            pixmap = QPixmap()
            pixmap.loadFromData(img_data)
            
            # Add to scene
            self.pixmap_item = QGraphicsPixmapItem(pixmap)
            self.scene.addItem(self.pixmap_item)
            
            # Store original image size for coordinate conversion
            self.original_size = (pix.width, pix.height)
            self.scale_factor = 1.0  # Same as matrix zoom
            
            pdf_document.close()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load PDF: {str(e)}")
    
    def on_region_changed(self):
        """Handle region selection change"""
        self.current_region = self.region_combo.currentData()
        region = self.regions[self.current_region]
        
        # Clear existing rectangles to show only the current region
        for item in self.scene.items():
            if isinstance(item, QGraphicsRectItem) and item != self.pixmap_item:
                self.scene.removeItem(item)
        
        # Draw only the current region if it has coordinates
        if region['coordinates']:
            x1, y1, x2, y2 = region['coordinates']
            rect = QRectF(x1, y1, x2 - x1, y2 - y1)
            
            # Set color based on current region
            if region['color'] == 'red':
                pen_color = QColor(255, 0, 0, 200)
            elif region['color'] == 'blue':
                pen_color = QColor(0, 0, 255, 200)
            elif region['color'] == 'green':
                pen_color = QColor(0, 255, 0, 200)
            elif region['color'] == 'orange':
                pen_color = QColor(255, 165, 0, 200)  # Orange color
            elif region['color'] == 'purple':
                pen_color = QColor(128, 0, 128, 200)  # Purple color
            else:
                pen_color = QColor(128, 128, 128, 200)
            
            rect_item = self.scene.addRect(rect, QPen(pen_color, 2))
            rect_item.setZValue(1)
            
            # Add label
            text_item = self.scene.addText(region['name'], QFont("Arial", 10, QFont.Bold))
            text_item.setDefaultTextColor(pen_color)
            text_item.setPos(rect.x(), rect.y() - 15)
            text_item.setZValue(2)
        
        # Update status labels
        self.status_label.setText(f"Select {region['name']} ({region['color'].title()}) and click and drag to select the OCR region")
        
        if region['coordinates']:
            self.coord_label.setText(f"{region['name']} ({region['color'].title()}): {region['coordinates']}")
            self.clear_btn.setEnabled(True)
            self.test_btn.setEnabled(True)
        else:
            self.coord_label.setText(f"{region['name']} ({region['color'].title()}): Not selected")
            self.clear_btn.setEnabled(False)
            self.test_btn.setEnabled(False)
        
        # Update region status
        self.update_region_status()
        
        # Show/hide region 5 configuration
        self.region_5_config_group.setVisible(self.current_region == 'region_5')
    
    def update_region_status(self):
        """Update the region status display"""
        status_parts = []
        for region_id, region in self.regions.items():
            if region_id != self.current_region:
                status = "Selected" if region['coordinates'] else "Not selected"
                status_parts.append(f"{region['name']} ({region['color'].title()}): {status}")
        
        self.region_status_label.setText(" | ".join(status_parts))
        
        # Enable save button if at least one region is configured
        configured_count = sum(1 for region in self.regions.values() if region['coordinates'])
        self.save_btn.setEnabled(configured_count > 0)
        self.show_all_btn.setEnabled(configured_count > 0)
    
    def draw_all_regions(self):
        """Draw all configured regions on the PDF"""
        # Clear existing rectangles
        for item in self.scene.items():
            if isinstance(item, QGraphicsRectItem) and item != self.pixmap_item:
                self.scene.removeItem(item)
        
        # Draw all regions
        for region_id, region in self.regions.items():
            if region['coordinates']:
                x1, y1, x2, y2 = region['coordinates']
                rect = QRectF(x1, y1, x2 - x1, y2 - y1)
                
                # Set color based on region
                if region['color'] == 'red':
                    pen_color = QColor(255, 0, 0, 200)
                elif region['color'] == 'blue':
                    pen_color = QColor(0, 0, 255, 200)
                elif region['color'] == 'green':
                    pen_color = QColor(0, 255, 0, 200)
                elif region['color'] == 'orange':
                    pen_color = QColor(255, 165, 0, 200)  # Orange color
                elif region['color'] == 'purple':
                    pen_color = QColor(128, 0, 128, 200)  # Purple color
                else:
                    pen_color = QColor(128, 128, 128, 200)
                
                rect_item = self.scene.addRect(rect, QPen(pen_color, 2))
                rect_item.setZValue(1)
                
                # Add label
                text_item = self.scene.addText(region['name'], QFont("Arial", 10, QFont.Bold))
                text_item.setDefaultTextColor(pen_color)
                text_item.setPos(rect.x(), rect.y() - 15)
                text_item.setZValue(2)
    
    def finish_selection(self):
        """Finish the selection and calculate coordinates"""
        if self.graphics_view.start_point and self.graphics_view.end_point:
            # Convert to original image coordinates
            x1 = int(self.graphics_view.start_point.x() * self.scale_factor)
            y1 = int(self.graphics_view.start_point.y() * self.scale_factor)
            x2 = int(self.graphics_view.end_point.x() * self.scale_factor)
            y2 = int(self.graphics_view.end_point.y() * self.scale_factor)
            
            # Ensure proper order
            x1, x2 = min(x1, x2), max(x1, x2)
            y1, y2 = min(y1, y2), max(y1, y2)
            
            # Save coordinates to current region
            self.regions[self.current_region]['coordinates'] = (x1, y1, x2, y2)
            
            # Update display
            region = self.regions[self.current_region]
            self.coord_label.setText(f"{region['name']} ({region['color'].title()}): {region['coordinates']}")
            self.status_label.setText(f"Selection complete for {region['name']}! Switch to another region or save all regions.")
            
            # Enable buttons
            self.clear_btn.setEnabled(True)
            self.test_btn.setEnabled(True)
            
            # Update region status and redraw
            self.update_region_status()
            self.draw_all_regions()
    
    def update_selection_rect(self):
        """Update the selection rectangle display"""
        if self.graphics_view.start_point and self.graphics_view.end_point:
            # Remove previous rectangle
            for item in self.scene.items():
                if isinstance(item, QGraphicsRectItem) and item != self.pixmap_item:
                    self.scene.removeItem(item)
            
            # Create new rectangle with current region's color
            rect = QRectF(self.graphics_view.start_point, self.graphics_view.end_point)
            current_color = self.regions[self.current_region]['color']
            
            # Set color based on current region
            if current_color == 'red':
                pen_color = QColor(255, 0, 0, 200)
            elif current_color == 'blue':
                pen_color = QColor(0, 0, 255, 200)
            elif current_color == 'green':
                pen_color = QColor(0, 255, 0, 200)
            elif current_color == 'orange':
                pen_color = QColor(255, 165, 0, 200)  # Orange color
            elif current_color == 'purple':
                pen_color = QColor(128, 0, 128, 200)  # Purple color
            else:
                pen_color = QColor(128, 128, 128, 200)
            
            rect_item = self.scene.addRect(rect, QPen(pen_color, 2))
            rect_item.setZValue(1)  # Above the image
    
    def clear_current_region(self):
        """Clear the current region selection"""
        self.regions[self.current_region]['coordinates'] = None
        
        # Update display
        region = self.regions[self.current_region]
        self.coord_label.setText(f"{region['name']} ({region['color'].title()}): Not selected")
        self.status_label.setText(f"Select {region['name']} ({region['color'].title()}) and click and drag to select the OCR region")
        
        # Update buttons and redraw
        self.clear_btn.setEnabled(False)
        self.test_btn.setEnabled(False)
        self.update_region_status()
        self.draw_all_regions()
    
    def test_current_region(self):
        """Test OCR on the current region"""
        region = self.regions[self.current_region]
        if not region['coordinates']:
            return
        
        try:
            # Open PDF and get first page
            pdf_document = fitz.open(self.pdf_path)
            page = pdf_document[0]
            
            # Convert to image with same settings as main processing
            mat = fitz.Matrix(1.0, 1.0)  # 1.0x zoom to match coordinate selector
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            
            # Convert to PIL Image
            image = Image.open(io.BytesIO(img_data))
            
            # Crop to selected coordinates
            cropped_image = image.crop(region['coordinates'])
            
            # Test OCR with multiple configurations
            results = []
            psm_modes = [3, 6, 7, 8, 13]
            
            for psm_mode in psm_modes:
                try:
                    text = pytesseract.image_to_string(cropped_image, config=f'--psm {psm_mode}')
                    if text.strip():
                        results.append(f"PSM {psm_mode}: '{text.strip()}'")
                except Exception as e:
                    results.append(f"PSM {psm_mode}: Error - {str(e)}")
            
            pdf_document.close()
            
            # Show results
            result_text = f"OCR Test Results for {region['name']} ({region['color'].title()}):\n"
            result_text += f"Coordinates: {region['coordinates']}\n\n"
            if results:
                result_text += "\n".join(results)
            else:
                result_text += "No text detected with any OCR configuration."
            
            QMessageBox.information(
                self,
                f"OCR Test Results - {region['name']}",
                result_text
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "OCR Test Error",
                f"Error testing OCR: {str(e)}"
            )
    
    def show_all_regions(self):
        """Show all configured regions"""
        configured_regions = [region for region in self.regions.values() if region['coordinates']]
        if not configured_regions:
            QMessageBox.information(self, "No Regions", "No regions have been configured yet.")
            return
        
        # Create a summary dialog
        summary_dialog = QDialog(self)
        summary_dialog.setWindowTitle("All Configured OCR Regions")
        summary_dialog.setModal(True)
        summary_dialog.resize(600, 400)
        
        layout = QVBoxLayout(summary_dialog)
        
        # Title
        title = QLabel("Configured OCR Regions")
        title.setObjectName("sectionTitle")
        layout.addWidget(title)
        
        # Text area for results
        text_area = QTextEdit()
        text_area.setReadOnly(True)
        text_area.setFont(QFont("Consolas", 10))
        
        result_text = "All Configured OCR Regions\n"
        result_text += "=" * 40 + "\n\n"
        
        for region in configured_regions:
            result_text += f"{region['name']} ({region['color'].title()}):\n"
            result_text += f"  Coordinates: {region['coordinates']}\n\n"
        
        text_area.setPlainText(result_text)
        layout.addWidget(text_area)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(summary_dialog.accept)
        layout.addWidget(close_btn)
        
        # Show dialog
        summary_dialog.exec()
    
    def browse_save_location(self):
        """Browse for save location for region 5 data"""
        options = [
            "Column K (Region 5 Data) - Default",
            "Column L (Custom Field 1)",
            "Column M (Custom Field 2)",
            "Column N (Custom Field 3)",
            "Custom Excel Column",
            "Separate Text File",
            "Database Field"
        ]
        
        from PySide6.QtWidgets import QInputDialog
        choice, ok = QInputDialog.getItem(
            self,
            "Select Save Location for Region 5",
            "Choose where to save Region 5 extracted data:",
            options,
            0,
            False
        )
        
        if ok and choice:
            if choice == "Custom Excel Column":
                custom_col, ok = QInputDialog.getText(
                    self,
                    "Custom Excel Column",
                    "Enter custom column name (e.g., 'Custom Data', 'Notes', etc.):"
                )
                if ok and custom_col:
                    self.region_5_save_location.setText(f"Column: {custom_col}")
            else:
                self.region_5_save_location.setText(choice)
    
    def save_all_regions(self):
        """Save all configured regions"""
        configured_regions = [region for region in self.regions.values() if region['coordinates']]
        if not configured_regions:
            QMessageBox.warning(self, "No Regions", "No regions have been configured yet.")
            return
        
        # Save to config file
        config = {
            'ocr_regions': self.regions,
            'setup_completed': True,
            'setup_date': datetime.now().isoformat(),
            'region_5_save_location': self.region_5_save_location.text()
        }
        
        config_path = Path("app_data") / "ocr_config.json"
        config_path.parent.mkdir(exist_ok=True)
        
        with open(config_path, 'w') as f:
            json.dump(config, f, indent=2)
        
        # Update parent application's regions if available
        if hasattr(self.parent(), 'ocr_regions'):
            self.parent().ocr_regions = self.regions.copy()
        
        # Show success message
        region_list = "\n".join([f"• {region['name']} ({region['color'].title()}): {region['coordinates']}" 
                                for region in configured_regions])
        
        QMessageBox.information(
            self,
            "Configuration Complete",
            f"OCR regions saved successfully!\n\n"
            f"Configured regions:\n{region_list}\n\n"
            f"Config saved to: {config_path}"
        )
        
        self.accept()
    
    def show_ocr_region(self):
        """Show the current OCR region on the PDF"""
        if not hasattr(self, 'coordinates') or not self.coordinates:
            return
        
        try:
            # Create a new dialog to show the OCR region
            region_dialog = QDialog(self)
            region_dialog.setWindowTitle("Current OCR Region")
            region_dialog.setModal(True)
            region_dialog.resize(1200, 900)
            
            layout = QVBoxLayout(region_dialog)
            
            # Instructions
            instructions = QLabel(
                f"Current OCR Region: {self.coordinates}\n"
                "The red rectangle shows exactly where OCR will extract text from."
            )
            instructions.setObjectName("sectionTitle")
            instructions.setWordWrap(True)
            layout.addWidget(instructions)
            
            # Graphics view for PDF display
            graphics_view = QGraphicsView()
            layout.addWidget(graphics_view)
            
            # Scene for graphics
            scene = QGraphicsScene()
            graphics_view.setScene(scene)
            
            # Load PDF and display with OCR region highlighted
            pdf_document = fitz.open(self.pdf_path)
            page = pdf_document[0]
            
            # Convert to image
            mat = fitz.Matrix(1.0, 1.0)  # Normal size for better overview
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            
            # Convert to QPixmap
            pixmap = QPixmap()
            pixmap.loadFromData(img_data)
            
            # Add to scene
            pixmap_item = QGraphicsPixmapItem(pixmap)
            scene.addItem(pixmap_item)
            
            # Add OCR region rectangle
            x1, y1, x2, y2 = self.coordinates
            # Convert coordinates to display coordinates (divide by scale factor)
            scale_factor = 1.0  # Same as matrix
            rect = QRectF(
                x1 / scale_factor, 
                y1 / scale_factor, 
                (x2 - x1) / scale_factor, 
                (y2 - y1) / scale_factor
            )
            
            # Create red rectangle for OCR region
            rect_item = scene.addRect(rect, QPen(QColor(255, 0, 0, 255), 3))
            rect_item.setZValue(1)  # Above the image
            
            # Add text label
            text_item = scene.addText("OCR Region", QFont("Arial", 12, QFont.Bold))
            text_item.setDefaultTextColor(QColor(255, 0, 0))
            text_item.setPos(rect.x(), rect.y() - 20)
            text_item.setZValue(2)
            
            pdf_document.close()
            
            # Close button
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(region_dialog.accept)
            layout.addWidget(close_btn)
            
            # Show dialog
            region_dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error showing OCR region: {str(e)}"
            )
    
    def test_ocr(self):
        """Test OCR on the selected region"""
        if not hasattr(self, 'coordinates') or not self.coordinates:
            return
        
        try:
            # Open PDF and get first page
            pdf_document = fitz.open(self.pdf_path)
            page = pdf_document[0]
            
            # Convert to image with same settings as main processing
            mat = fitz.Matrix(1.0, 1.0)  # 1.0x zoom to match coordinate selector
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            
            # Convert to PIL Image
            image = Image.open(io.BytesIO(img_data))
            
            # Crop to selected coordinates
            cropped_image = image.crop(self.coordinates)
            
            # Test OCR with multiple configurations
            results = []
            psm_modes = [3, 6, 7, 8, 13]
            
            for psm_mode in psm_modes:
                try:
                    text = pytesseract.image_to_string(cropped_image, config=f'--psm {psm_mode}')
                    if text.strip():
                        results.append(f"PSM {psm_mode}: '{text.strip()}'")
                except Exception as e:
                    results.append(f"PSM {psm_mode}: Error - {str(e)}")
            
            pdf_document.close()
            
            # Show results
            result_text = f"OCR Test Results for coordinates {self.coordinates}:\n\n"
            if results:
                result_text += "\n".join(results)
            else:
                result_text += "No text detected with any OCR configuration."
            
            QMessageBox.information(
                self,
                "OCR Test Results",
                result_text
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "OCR Test Error",
                f"Error testing OCR: {str(e)}"
            )
    
    def save_coordinates(self):
        """Save the selected coordinates"""
        if hasattr(self, 'coordinates') and self.coordinates:
            # Save to a config file
            config = {
                'ocr_coordinates': self.coordinates,
                'setup_completed': True,
                'setup_date': datetime.now().isoformat()
            }
            
            config_path = Path("app_data") / "ocr_config.json"
            config_path.parent.mkdir(exist_ok=True)
            
            with open(config_path, 'w') as f:
                json.dump(config, f, indent=2)
            
            QMessageBox.information(
                self,
                "Setup Complete",
                f"OCR coordinates saved successfully!\n\n"
                f"Coordinates: {self.coordinates}\n"
                f"Config saved to: {config_path}"
            )
            
            self.accept()


class CrateCountDialog(QDialog):
    """Dialog for selecting number of crates to print"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Number of Crates")
        self.setModal(True)
        self.resize(400, 200)
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        
        # Set white background for the dialog
        self.setStyleSheet("""
            QDialog {
                background-color: white;
                color: black;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title_label = QLabel("How many crate labels do you want to print?")
        title_label.setObjectName("crateDialogTitle")
        title_label.setStyleSheet("""
            QLabel#crateDialogTitle {
                font-size: 16px;
                font-weight: bold;
                color: black;
                padding: 10px;
                background-color: white;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # Crate count selection
        self.crate_spinbox = QComboBox()
        self.crate_spinbox.setObjectName("crateDialogSpinbox")
        self.crate_spinbox.addItems([str(i) for i in range(1, 21)])  # 1-20 crates
        self.crate_spinbox.setCurrentText("1")
        self.crate_spinbox.setStyleSheet("""
            QComboBox#crateDialogSpinbox {
                background-color: white;
                color: black;
                border: 3px solid #3498db;
                border-radius: 8px;
                padding: 12px 16px;
                font-size: 18px;
                font-weight: bold;
                min-width: 120px;
                min-height: 30px;
            }
            QComboBox#crateDialogSpinbox:focus {
                border-color: #2980b9;
                background-color: white;
                color: black;
            }
            QComboBox#crateDialogSpinbox::drop-down {
                background-color: white;
                border: none;
            }
            QComboBox#crateDialogSpinbox QAbstractItemView {
                background-color: white;
                color: black;
                border: 1px solid #3498db;
            }
        """)
        layout.addWidget(self.crate_spinbox, alignment=Qt.AlignCenter)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        cancel_button = QPushButton("Cancel")
        cancel_button.setObjectName("cancelButton")
        cancel_button.setStyleSheet("""
            QPushButton#cancelButton {
                background-color: #f8f9fa;
                color: black;
                border: 2px solid #6c757d;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton#cancelButton:hover {
                background-color: #e9ecef;
            }
        """)
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        preview_button = QPushButton("Preview Labels")
        preview_button.setObjectName("previewButton")
        preview_button.setStyleSheet("""
            QPushButton#previewButton {
                background-color: #17a2b8;
                color: white;
                border: 2px solid #17a2b8;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton#previewButton:hover {
                background-color: #138496;
            }
        """)
        preview_button.clicked.connect(self.preview_labels)
        button_layout.addWidget(preview_button)
        
        print_button = QPushButton("Print Labels")
        print_button.setObjectName("printDialogButton")
        print_button.setStyleSheet("""
            QPushButton#printDialogButton {
                background-color: #28a745;
                color: white;
                border: 2px solid #28a745;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton#printDialogButton:hover {
                background-color: #218838;
            }
        """)
        print_button.clicked.connect(self.accept)
        print_button.setDefault(True)
        button_layout.addWidget(print_button)
        
        layout.addLayout(button_layout)
        
        # Set focus to spinbox
        self.crate_spinbox.setFocus()
    
    def keyPressEvent(self, event):
        """Handle key press events"""
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            self.accept()
        elif event.key() == Qt.Key_Escape:
            self.reject()
        else:
            super().keyPressEvent(event)
    
    def get_crate_count(self):
        """Get the selected crate count"""
        return int(self.crate_spinbox.currentText())
    
    def preview_labels(self):
        """Preview the labels that will be printed"""
        if not hasattr(self.parent(), 'current_order_data') or not self.parent().current_order_data:
            QMessageBox.warning(self, "Preview Error", "No order data available.")
            return
        
        try:
            crate_count = self.get_crate_count()
            order_data = self.parent().current_order_data
            
            order_number = order_data.get('ordernumber', 'N/A')
            site_name = order_data.get('sitename', 'N/A')
            route = order_data.get('route', 'N/A')
            dispatchcode = order_data.get('dispatchcode', 'N/A')
            
            # Create preview dialog showing all labels
            preview_dialog = LabelPreviewDialog(
                zpl_code="",  # We'll generate this in the dialog
                order_number=order_number,
                dispatchcode=dispatchcode,
                site_name=site_name,
                route=route,
                crate_number=1,  # Show first crate as example
                total_crates=crate_count,
                parent=self
            )
            
            # Update the preview dialog to show multiple labels
            preview_dialog.setWindowTitle(f"Label Preview - {crate_count} Labels for Order {order_number}")
            preview_dialog.update_preview_for_multiple_labels(crate_count, order_data)
            
            preview_dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Preview Error", f"Error generating preview: {str(e)}")


class LabelPreviewDialog(QDialog):
    """Dialog for previewing label layout"""
    
    def __init__(self, zpl_code, order_number, dispatchcode, site_name, route, crate_number, total_crates, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Label Preview")
        self.setModal(True)
        self.resize(600, 800)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Title
        title_label = QLabel("Label Preview")
        title_label.setObjectName("previewTitle")
        title_label.setStyleSheet("""
            QLabel#previewTitle {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
                margin-bottom: 10px;
            }
        """)
        layout.addWidget(title_label)
        
        # Label preview area
        preview_frame = QFrame()
        preview_frame.setObjectName("previewFrame")
        preview_frame.setStyleSheet("""
            QFrame#previewFrame {
                background-color: white;
                border: 2px solid #e2e8f0;
                border-radius: 8px;
                padding: 20px;
            }
        """)
        preview_layout = QVBoxLayout(preview_frame)
        
        # Create a visual representation of the label
        self.create_label_preview(preview_layout, order_number, dispatchcode, site_name, route, crate_number, total_crates)
        
        layout.addWidget(preview_frame)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        close_button = QPushButton("Close")
        close_button.setObjectName("closeButton")
        close_button.clicked.connect(self.accept)
        button_layout.addWidget(close_button)
        
        layout.addLayout(button_layout)
    
    def create_label_preview(self, layout, order_number, dispatchcode, site_name, route, crate_number, total_crates):
        """Create a visual preview of the label"""
        # Calculate label dimensions (6.5cm x 13.5cm scaled down for preview)
        label_width = 260  # Scaled down from 520 dots
        label_height = 540  # Scaled down from 1080 dots
        
        # Create a custom widget for the label preview
        label_widget = QWidget()
        label_widget.setFixedSize(label_width, label_height)
        label_widget.setStyleSheet("""
            QWidget {
                background-color: white;
                border: 1px solid #333;
            }
        """)
        
        # Add label content using QLabel widgets positioned absolutely
        label_widget.setLayout(QVBoxLayout())
        label_widget.layout().setContentsMargins(10, 10, 10, 10)
        label_widget.layout().setSpacing(5)
        
        # Add more empty space at top to match new ZPL positioning (moved content down)
        label_widget.layout().addStretch(1.5)  # Fine-tuned to match adjusted positioning
        
        # Separator line
        line1 = QFrame()
        line1.setFrameShape(QFrame.HLine)
        line1.setStyleSheet("color: #333;")
        label_widget.layout().addWidget(line1)
        
        # Site Name section (no header)
        
        # Handle long site names (matching ZPL logic - max 2 lines)
        site_name_lines = []
        max_chars_per_line = 25  # Match ZPL logic
        
        if len(site_name) > max_chars_per_line:
            words = site_name.split()
            current_line = ""
            for word in words:
                if len(current_line + " " + word) <= max_chars_per_line:
                    current_line += (" " + word) if current_line else word
                else:
                    if current_line:
                        site_name_lines.append(current_line)
                    current_line = word
                    # Limit to maximum 2 lines
                    if len(site_name_lines) >= 2:
                        # If we already have 2 lines, truncate the remaining text
                        remaining_text = " ".join([current_line] + words[words.index(word):])
                        if len(remaining_text) > 15:
                            remaining_text = remaining_text[:12] + "..."
                        site_name_lines.append(remaining_text)
                        break
            if current_line and len(site_name_lines) < 2:
                site_name_lines.append(current_line)
        else:
            site_name_lines = [site_name]
        
        for line in site_name_lines:
            site_value_label = QLabel(line)
            site_value_label.setStyleSheet("font-size: 18px; font-weight: bold; color: black;")
            label_widget.layout().addWidget(site_value_label)
        
        # Separator line
        line2 = QFrame()
        line2.setFrameShape(QFrame.HLine)
        line2.setStyleSheet("color: #333;")
        label_widget.layout().addWidget(line2)
        
        # Date (center aligned) - use selected date from parent
        current_date = self.parent().selected_date if hasattr(self.parent(), 'selected_date') else datetime.now().strftime("%d/%m/%Y")
        date_label = QLabel(current_date)
        date_label.setStyleSheet("font-size: 10px; color: black;")
        date_label.setAlignment(Qt.AlignCenter)
        label_widget.layout().addWidget(date_label)
        
        # Separator line
        line3 = QFrame()
        line3.setFrameShape(QFrame.HLine)
        line3.setStyleSheet("color: #333;")
        label_widget.layout().addWidget(line3)
        
        # Crate info (center aligned)
        crate_label = QLabel(f"{crate_number} of {total_crates}")
        crate_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #0066cc;")
        crate_label.setAlignment(Qt.AlignCenter)
        label_widget.layout().addWidget(crate_label)
        
        # Separator line
        line4 = QFrame()
        line4.setFrameShape(QFrame.HLine)
        line4.setStyleSheet("color: #333;")
        label_widget.layout().addWidget(line4)
        
        # Barcode representation (center aligned)
        barcode_label = QLabel("████████████████████████████████")
        barcode_label.setStyleSheet("font-size: 8px; color: black; font-family: monospace;")
        barcode_label.setAlignment(Qt.AlignCenter)
        label_widget.layout().addWidget(barcode_label)
        
        barcode_text_label = QLabel(order_number)
        barcode_text_label.setStyleSheet("font-size: 10px; color: black; font-family: monospace;")
        barcode_text_label.setAlignment(Qt.AlignCenter)
        label_widget.layout().addWidget(barcode_text_label)
        
        # Separator line
        line5 = QFrame()
        line5.setFrameShape(QFrame.HLine)
        line5.setStyleSheet("color: #333;")
        label_widget.layout().addWidget(line5)
        
        # Route section (center aligned)
        route_label = QLabel("Route")
        route_label.setStyleSheet("font-size: 10px; color: black;")
        route_label.setAlignment(Qt.AlignCenter)
        label_widget.layout().addWidget(route_label)
        
        # Handle long route names - split into maximum 2 lines with optimized breaking
        route_lines = []
        max_route_chars_per_line = 15
        
        if len(route) > max_route_chars_per_line:
            words = route.split()
            current_line = ""
            for word in words:
                if len(current_line + " " + word) <= max_route_chars_per_line:
                    current_line += (" " + word) if current_line else word
                else:
                    if current_line:
                        route_lines.append(current_line)
                    current_line = word
                    # Limit to maximum 2 lines
                    if len(route_lines) >= 2:
                        # If we already have 2 lines, truncate the remaining text
                        remaining_text = " ".join([current_line] + words[words.index(word):])
                        if len(remaining_text) > 15:
                            remaining_text = remaining_text[:12] + "..."
                        route_lines.append(remaining_text)
                        break
            if current_line and len(route_lines) < 2:
                route_lines.append(current_line)
        else:
            route_lines = [route]
        
        # Dynamic route font sizing based on number of lines and total length
        num_route_lines = len(route_lines)
        total_route_length = len(route)
        
        if num_route_lines == 1:
            if total_route_length <= 6:
                route_font_size = "24px"
            elif total_route_length <= 10:
                route_font_size = "22px"
            elif total_route_length <= 15:
                route_font_size = "20px"
            elif total_route_length <= 20:
                route_font_size = "18px"
            else:
                route_font_size = "16px"
        else:  # 2 lines
            if total_route_length <= 20:
                route_font_size = "18px"
            elif total_route_length <= 30:
                route_font_size = "16px"
            else:
                route_font_size = "14px"
        
        for line in route_lines:
            route_value_label = QLabel(line)
            route_value_label.setStyleSheet(f"font-size: {route_font_size}; font-weight: bold; color: black;")
        route_value_label.setAlignment(Qt.AlignCenter)
        label_widget.layout().addWidget(route_value_label)
        
        # Add some spacing at the bottom
        label_widget.layout().addStretch()
        
        layout.addWidget(label_widget, alignment=Qt.AlignCenter)
    
    def update_preview_for_multiple_labels(self, crate_count, order_data):
        """Update the preview to show multiple labels"""
        try:
            # Clear existing content
            for i in reversed(range(self.layout().count())):
                child = self.layout().itemAt(i).widget()
                if child:
                    child.setParent(None)
            
            # Get order data
            order_number = order_data.get('ordernumber', 'N/A')
            site_name = order_data.get('sitename', 'N/A')
            route = order_data.get('route', 'N/A')
            dispatchcode = order_data.get('dispatchcode', 'N/A')
            
            # Create scroll area for multiple labels
            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)
            scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            
            # Create container widget for all labels
            container_widget = QWidget()
            container_layout = QVBoxLayout(container_widget)
            container_layout.setSpacing(20)
            container_layout.setContentsMargins(20, 20, 20, 20)
            
            # Generate preview for each label
            for i in range(1, crate_count + 1):
                # Create label preview widget
                label_widget = self.create_single_label_preview(
                    order_number, site_name, route, dispatchcode, i, crate_count
                )
                container_layout.addWidget(label_widget)
            
            scroll_area.setWidget(container_widget)
            self.layout().addWidget(scroll_area)
            
        except Exception as e:
            QMessageBox.critical(self, "Preview Error", f"Error updating preview: {str(e)}")
    
    def create_single_label_preview(self, order_number, site_name, route, dispatchcode, crate_number, total_crates):
        """Create a single label preview widget"""
        # Create label container
        label_container = QFrame()
        label_container.setFrameStyle(QFrame.Box)
        label_container.setLineWidth(2)
        label_container.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 2px solid #333;
                border-radius: 8px;
                margin: 10px;
            }
        """)
        
        # Set label size (6.5cm x 13.5cm scaled down for preview)
        label_container.setFixedSize(260, 540)  # Scaled down by 0.5 for preview
        
        layout = QVBoxLayout(label_container)
        layout.setSpacing(5)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Add label number
        label_number = QLabel(f"Label {crate_number} of {total_crates}")
        label_number.setAlignment(Qt.AlignCenter)
        label_number.setStyleSheet("font-size: 12px; font-weight: bold; color: #666; margin-bottom: 5px;")
        layout.addWidget(label_number)
        
        # Add more empty space at top to match new ZPL positioning (moved content down)
        layout.addStretch(1.5)  # Fine-tuned to match adjusted positioning
        
        # Add separator line
        line1 = QFrame()
        line1.setFrameShape(QFrame.HLine)
        line1.setStyleSheet("color: #333;")
        layout.addWidget(line1)
        
        # Add site name (handle multiple lines)
        site_lines = site_name.split('\n') if '\n' in site_name else [site_name]
        for line in site_lines:
            if line.strip():
                site_label = QLabel(line.strip())
                site_label.setAlignment(Qt.AlignLeft)
                site_label.setStyleSheet("font-size: 20px; font-weight: bold; color: black;")
                site_label.setWordWrap(True)
                layout.addWidget(site_label)
        
        # Add separator line
        line2 = QFrame()
        line2.setFrameShape(QFrame.HLine)
        line2.setStyleSheet("color: #333;")
        layout.addWidget(line2)
        
        
        # Add crate info
        crate_text = f"{crate_number} of {total_crates}"
        crate_label = QLabel(crate_text)
        crate_label.setAlignment(Qt.AlignCenter)
        crate_label.setStyleSheet("font-size: 24px; font-weight: bold; color: black;")
        layout.addWidget(crate_label)
        
        # Add separator line
        line3 = QFrame()
        line3.setFrameShape(QFrame.HLine)
        line3.setStyleSheet("color: #333;")
        layout.addWidget(line3)
        
        # Add barcode representation (text-based)
        barcode_label = QLabel(f"Barcode: {order_number}")
        barcode_label.setAlignment(Qt.AlignCenter)
        barcode_label.setStyleSheet("font-size: 14px; font-weight: bold; color: black; background-color: #f0f0f0; padding: 5px;")
        layout.addWidget(barcode_label)
        
        # Add separator line
        line4 = QFrame()
        line4.setFrameShape(QFrame.HLine)
        line4.setStyleSheet("color: #333;")
        layout.addWidget(line4)
        
        
        route_value_label = QLabel(route)
        route_value_label.setAlignment(Qt.AlignCenter)
        route_value_label.setStyleSheet("font-size: 18px; font-weight: bold; color: black;")
        layout.addWidget(route_value_label)
        
        # Add stretch to push everything to top
        layout.addStretch()
        
        return label_container

        self.order_number_input = QLineEdit()
        self.order_number_input.setPlaceholderText("Enter order number")
        self.order_number_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                color: black;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 20px 16px;
                font-size: 16px;
                font-weight: normal;
                min-height: 25px;
                selection-background-color: #3b82f6;
                selection-color: white;
            }
            QLineEdit:focus {
                border-color: #3b82f6;
                background-color: white;
                color: black;
                outline: none;
            }
            QLineEdit:hover {
                border-color: #9ca3af;
            }
        """)
        form_layout.addWidget(self.order_number_input)
        
        # Site Name
        self.site_name_input = QLineEdit()
        self.site_name_input.setPlaceholderText("Enter site name")
        self.site_name_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                color: black;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 20px 16px;
                font-size: 16px;
                font-weight: normal;
                min-height: 25px;
                selection-background-color: #3b82f6;
                selection-color: white;
            }
            QLineEdit:focus {
                border-color: #3b82f6;
                background-color: white;
                color: black;
                outline: none;
            }
            QLineEdit:hover {
                border-color: #9ca3af;
            }
        """)
        form_layout.addWidget(self.site_name_input)
        
        # Route
        self.route_input = QLineEdit()
        self.route_input.setPlaceholderText("Enter route")
        self.route_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                color: black;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 20px 16px;
                font-size: 16px;
                font-weight: normal;
                min-height: 25px;
                selection-background-color: #3b82f6;
                selection-color: white;
            }
            QLineEdit:focus {
                border-color: #3b82f6;
                background-color: white;
                color: black;
                outline: none;
            }
            QLineEdit:hover {
                border-color: #9ca3af;
            }
        """)
        form_layout.addWidget(self.route_input)
        
        # Dispatch Code
        self.dispatch_code_input = QLineEdit()
        self.dispatch_code_input.setPlaceholderText("Enter dispatch code")
        self.dispatch_code_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                color: black;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 20px 16px;
                font-size: 16px;
                font-weight: normal;
                min-height: 25px;
                selection-background-color: #3b82f6;
                selection-color: white;
            }
            QLineEdit:focus {
                border-color: #3b82f6;
                background-color: white;
                color: black;
                outline: none;
            }
            QLineEdit:hover {
                border-color: #9ca3af;
            }
        """)
        form_layout.addWidget(self.dispatch_code_input)
        
        # Crate Count
        self.crate_count_input = QLineEdit()
        self.crate_count_input.setPlaceholderText("Enter number of crates")
        self.crate_count_input.setText("1")
        self.crate_count_input.setStyleSheet("""
            QLineEdit {
                background-color: white;
                color: black;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 20px 16px;
                font-size: 16px;
                font-weight: normal;
                min-height: 25px;
                selection-background-color: #3b82f6;
                selection-color: white;
            }
            QLineEdit:focus {
                border-color: #3b82f6;
                background-color: white;
                color: black;
                outline: none;
            }
            QLineEdit:hover {
                border-color: #9ca3af;
            }
        """)
        form_layout.addWidget(self.crate_count_input)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(20)
        button_layout.setContentsMargins(0, 20, 0, 0)  # Add top margin for spacing
        button_layout.addStretch()
        
        cancel_button = QPushButton("Cancel")
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 12px 28px;
                font-size: 14px;
                font-weight: 500;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
            QPushButton:pressed {
                background-color: #545b62;
            }
        """)
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        print_button = QPushButton("Print Labels")
        print_button.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 12px 28px;
                font-size: 14px;
                font-weight: 500;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #004085;
            }
        """)
        print_button.clicked.connect(self.accept)
        button_layout.addWidget(print_button)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        # Set focus to first input
        self.order_number_input.setFocus()
    
    def get_label_data(self):
        """Get the entered label data"""
        return {
            'order_number': self.order_number_input.text().strip(),
            'site_name': self.site_name_input.text().strip(),
            'route': self.route_input.text().strip(),
            'dispatch_code': self.dispatch_code_input.text().strip(),
            'crate_count': int(self.crate_count_input.text()) if self.crate_count_input.text().isdigit() else 1
        }


class OCRDebugDialog(QDialog):
    """Dialog for displaying OCR region debugging information"""
    
    def __init__(self, ocr_regions, parent=None):
        super().__init__(parent)
        self.ocr_regions = ocr_regions
        self.setWindowTitle("OCR Regions Debug Information")
        self.setModal(True)
        self.resize(800, 600)
        
        self.init_ui()
        self.apply_styling()
    
    def init_ui(self):
        """Initialize the user interface"""
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header
        header_label = QLabel("🔍 OCR Regions Debug Information")
        header_label.setObjectName("headerTitle")
        layout.addWidget(header_label)
        
        # Instructions
        instructions = QLabel(
            "This dialog shows all configured OCR regions and their coordinates. "
            "Use this information to verify that your OCR regions are set up correctly.\n\n"
            "<b>To view all regions on a PDF:</b> Click 'Select PDF to View All Regions' to choose a PDF file and see all configured regions overlaid on it."
        )
        instructions.setObjectName("infoText")
        instructions.setWordWrap(True)
        layout.addWidget(instructions)
        
        # OCR Regions Table
        self.regions_table = QTableWidget()
        self.regions_table.setColumnCount(4)
        self.regions_table.setHorizontalHeaderLabels(["Region", "Color", "Coordinates", "Status"])
        self.regions_table.horizontalHeader().setStretchLastSection(True)
        self.regions_table.setAlternatingRowColors(True)
        self.regions_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        # Populate table with OCR regions
        self.populate_regions_table()
        
        layout.addWidget(self.regions_table)
        
        # Summary information
        summary_label = QLabel()
        configured_count = len([r for r in self.ocr_regions.values() if r.get('coordinates')])
        total_count = len(self.ocr_regions)
        summary_label.setText(f"<b>Summary:</b> {configured_count} of {total_count} regions configured")
        summary_label.setObjectName("infoText")
        layout.addWidget(summary_label)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.test_region_btn = QPushButton("Select PDF to View All Regions")
        self.test_region_btn.setObjectName("secondaryButton")
        self.test_region_btn.clicked.connect(self.select_pdf_to_view_regions)
        self.test_region_btn.setEnabled(True)
        self.test_region_btn.setToolTip("Click to select a PDF file and view all configured OCR regions overlaid on it")
        
        self.close_btn = QPushButton("Close")
        self.close_btn.setObjectName("primaryButton")
        self.close_btn.clicked.connect(self.accept)
        
        button_layout.addWidget(self.test_region_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.close_btn)
        
        layout.addLayout(button_layout)
        
    
    def populate_regions_table(self):
        """Populate the regions table with OCR region data"""
        self.regions_table.setRowCount(len(self.ocr_regions))
        
        for row, (region_id, region_data) in enumerate(self.ocr_regions.items()):
            # Region name
            region_name = region_data.get('name', region_id)
            self.regions_table.setItem(row, 0, QTableWidgetItem(region_name))
            
            # Color
            color = region_data.get('color', 'unknown')
            color_item = QTableWidgetItem(color.title())
            color_item.setBackground(self.get_color_brush(color))
            self.regions_table.setItem(row, 1, color_item)
            
            # Coordinates
            coordinates = region_data.get('coordinates')
            if coordinates:
                coord_text = f"[{', '.join(map(str, coordinates))}]"
                status = "Configured"
            else:
                coord_text = "Not set"
                status = "Not configured"
            
            self.regions_table.setItem(row, 2, QTableWidgetItem(coord_text))
            self.regions_table.setItem(row, 3, QTableWidgetItem(status))
    
    def get_color_brush(self, color):
        """Get a QBrush for the given color"""
        color_map = {
            'red': QColor(255, 200, 200),
            'blue': QColor(200, 200, 255),
            'green': QColor(200, 255, 200),
            'orange': QColor(255, 220, 200),
            'purple': QColor(220, 200, 255)
        }
        return QBrush(color_map.get(color, QColor(240, 240, 240)))
    
    
    def select_pdf_to_view_regions(self):
        """Select a PDF file and show all configured regions on it"""
        # Let user select a PDF file
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select PDF File to View All OCR Regions",
            "",
            "PDF Files (*.pdf)"
        )
        
        if file_path:
            try:
                # Show all regions on the selected PDF
                self.show_all_regions_on_pdf(file_path)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to show regions on PDF: {str(e)}")
    
    def show_all_regions_on_pdf(self, pdf_path):
        """Show all configured OCR regions overlaid on the PDF"""
        dialog = QDialog(self)
        dialog.setWindowTitle("All OCR Regions on PDF")
        dialog.setModal(True)
        dialog.resize(1200, 900)
        
        layout = QVBoxLayout(dialog)
        
        # Graphics view for PDF display
        graphics_view = QGraphicsView()
        scene = QGraphicsScene()
        graphics_view.setScene(scene)
        
        try:
            # Load PDF with PyMuPDF
            import fitz
            pdf_document = fitz.open(pdf_path)
            page = pdf_document[0]
            
            # Convert to image
            mat = fitz.Matrix(1.0, 1.0)
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            
            # Create QPixmap from image data
            pixmap = QPixmap()
            pixmap.loadFromData(img_data)
            
            # Add pixmap to scene
            pixmap_item = scene.addPixmap(pixmap)
            scene.setSceneRect(pixmap.rect())
            
            # Add all OCR region rectangles
            regions_added = 0
            for region_id, region_data in self.ocr_regions.items():
                coordinates = region_data.get('coordinates')
                if coordinates and len(coordinates) == 4:
                    x1, y1, x2, y2 = coordinates
                    rect = QRectF(x1, y1, x2 - x1, y2 - y1)
                    
                    # Set color based on region
                    color = region_data.get('color', 'red')
                    if color == 'red':
                        pen_color = QColor(255, 0, 0, 200)
                    elif color == 'blue':
                        pen_color = QColor(0, 0, 255, 200)
                    elif color == 'green':
                        pen_color = QColor(0, 255, 0, 200)
                    elif color == 'orange':
                        pen_color = QColor(255, 165, 0, 200)
                    elif color == 'purple':
                        pen_color = QColor(128, 0, 128, 200)
                    else:
                        pen_color = QColor(128, 128, 128, 200)
                    
                    pen = QPen(pen_color, 3)
                    rect_item = scene.addRect(rect, pen)
                    rect_item.setZValue(1)
                    
                    # Add label
                    text_item = scene.addText(region_data['name'], QFont("Arial", 12, QFont.Bold))
                    text_item.setDefaultTextColor(pen_color)
                    text_item.setPos(rect.x(), rect.y() - 20)
                    text_item.setZValue(2)
                    
                    regions_added += 1
            
            pdf_document.close()
            
            # Add status label
            status_label = QLabel(f"Showing {regions_added} configured regions on PDF: {Path(pdf_path).name}")
            status_label.setStyleSheet("font-weight: bold; color: #2c3e50; padding: 10px;")
            layout.addWidget(status_label)
            
        except Exception as e:
            error_label = QLabel(f"Error loading PDF: {str(e)}")
            layout.addWidget(error_label)
        
        layout.addWidget(graphics_view)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec()
    
    def show_region_on_pdf(self, pdf_path, region_data):
        """Show the OCR region overlaid on the PDF"""
        dialog = QDialog(self)
        dialog.setWindowTitle(f"OCR Region: {region_data['name']}")
        dialog.setModal(True)
        dialog.resize(1000, 800)
        
        layout = QVBoxLayout(dialog)
        
        # Graphics view for PDF display
        graphics_view = QGraphicsView()
        scene = QGraphicsScene()
        graphics_view.setScene(scene)
        
        try:
            # Load PDF with PyMuPDF
            import fitz
            pdf_document = fitz.open(pdf_path)
            page = pdf_document[0]
            
            # Convert to image
            mat = fitz.Matrix(1.0, 1.0)
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            
            # Create QPixmap from image data
            pixmap = QPixmap()
            pixmap.loadFromData(img_data)
            
            # Add pixmap to scene
            pixmap_item = scene.addPixmap(pixmap)
            scene.setSceneRect(pixmap.rect())
            
            # Add OCR region rectangle
            coordinates = region_data['coordinates']
            x1, y1, x2, y2 = coordinates
            rect = QRectF(x1, y1, x2 - x1, y2 - y1)
            
            # Set color based on region
            color = region_data.get('color', 'red')
            if color == 'red':
                pen_color = QColor(255, 0, 0, 200)
            elif color == 'blue':
                pen_color = QColor(0, 0, 255, 200)
            elif color == 'green':
                pen_color = QColor(0, 255, 0, 200)
            elif color == 'orange':
                pen_color = QColor(255, 165, 0, 200)
            elif color == 'purple':
                pen_color = QColor(128, 0, 128, 200)
            else:
                pen_color = QColor(128, 128, 128, 200)
            
            pen = QPen(pen_color, 3)
            rect_item = scene.addRect(rect, pen)
            rect_item.setZValue(1)
            
            # Add label
            text_item = scene.addText(region_data['name'], QFont("Arial", 12, QFont.Bold))
            text_item.setDefaultTextColor(pen_color)
            text_item.setPos(rect.x(), rect.y() - 20)
            text_item.setZValue(2)
            
            pdf_document.close()
            
        except Exception as e:
            error_label = QLabel(f"Error loading PDF: {str(e)}")
            layout.addWidget(error_label)
        
        layout.addWidget(graphics_view)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec()
    
    def apply_styling(self):
        """Apply styling to the dialog"""
        self.setStyleSheet("""
            QDialog {
                background-color: white;
            }
            QLabel#headerTitle {
                font-size: 18px;
                font-weight: bold;
                color: #1f2937;
                margin-bottom: 10px;
            }
            QLabel#infoText {
                color: #6b7280;
                font-size: 14px;
                margin-bottom: 10px;
            }
            QTableWidget {
                gridline-color: #e5e7eb;
                background-color: white;
                alternate-background-color: #f9fafb;
                selection-background-color: #dbeafe;
                border: 1px solid #e5e7eb;
                border-radius: 6px;
            }
            QTableWidget::item {
                padding: 8px;
                border: none;
            }
            QHeaderView::section {
                background-color: #f3f4f6;
                padding: 8px;
                border: none;
                border-right: 1px solid #e5e7eb;
                font-weight: bold;
            }
            QPushButton#primaryButton {
                background-color: #2563eb;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-weight: 500;
                font-size: 14px;
                min-width: 100px;
            }
            QPushButton#primaryButton:hover {
                background-color: #1d4ed8;
            }
            QPushButton#secondaryButton {
                background-color: #6b7280;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-weight: 500;
                font-size: 14px;
                min-width: 100px;
            }
            QPushButton#secondaryButton:hover {
                background-color: #4b5563;
            }
        """)


class DateFilterDialog(QDialog):
    """Dialog for filtering print history by date range"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Filter Print History by Date")
        self.setModal(True)
        self.setFixedSize(400, 200)
        
        # Set dialog background to white
        self.setStyleSheet("""
            QDialog {
                background-color: white;
                border: 1px solid #e0e0e0;
                border-radius: 8px;
            }
        """)
        
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(30, 25, 30, 25)
        
        # Title
        title_label = QLabel("Filter Print History by Date Range")
        title_label.setObjectName("dateFilterTitle")
        title_label.setStyleSheet("""
            QLabel#dateFilterTitle {
                font-size: 18px;
                font-weight: 600;
                color: #2c3e50;
                padding: 0px;
                background-color: transparent;
            }
        """)
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # Date range selection
        date_layout = QHBoxLayout()
        date_layout.setSpacing(15)
        
        # Start date
        start_label = QLabel("From:")
        start_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        date_layout.addWidget(start_label)
        
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setDate(QDate.currentDate().addDays(-7))  # Default to 7 days ago
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDisplayFormat("dd/MM/yyyy")
        self.start_date_edit.setStyleSheet("""
            QDateEdit {
                background-color: white;
                color: black;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 14px;
                min-height: 20px;
            }
            QDateEdit:focus {
                border-color: #3b82f6;
            }
        """)
        date_layout.addWidget(self.start_date_edit)
        
        # End date
        end_label = QLabel("To:")
        end_label.setStyleSheet("font-weight: bold; color: #2c3e50;")
        date_layout.addWidget(end_label)
        
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setDate(QDate.currentDate())  # Default to today
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDisplayFormat("dd/MM/yyyy")
        self.end_date_edit.setStyleSheet("""
            QDateEdit {
                background-color: white;
                color: black;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 8px 12px;
                font-size: 14px;
                min-height: 20px;
            }
            QDateEdit:focus {
                border-color: #3b82f6;
            }
        """)
        date_layout.addWidget(self.end_date_edit)
        
        layout.addLayout(date_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(12)
        button_layout.addStretch()
        
        # Cancel button
        cancel_button = QPushButton("Cancel")
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 24px;
                font-size: 14px;
                font-weight: 500;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        # Filter button
        filter_button = QPushButton("Filter")
        filter_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px 24px;
                font-size: 14px;
                font-weight: 500;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        filter_button.clicked.connect(self.accept)
        button_layout.addWidget(filter_button)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        # Set focus to start date
        self.start_date_edit.setFocus()
    
    def get_date_range(self):
        """Get the selected date range"""
        return self.start_date_edit.date(), self.end_date_edit.date()


class DispatchScanningApp(QMainWindow):
    """Upload Excel Files, Process PDFs, Generate Barcodes"""
    
    def __init__(self):
        super().__init__()
        
        # Application data
        self.delivery_data_values = []
        self.delivery_data_with_drivers = {}
        self.delivery_json_file = "delivery_sequence_data.json"
        self.selected_picking_pdf_files = []
        self.selected_excel_file = ""  # For backward compatibility
        self.excel_order_numbers = []  # For backward compatibility
        self.excel_dataframe = None  # For backward compatibility

        self.order_barcodes = {}
        self.processing_thread = None
        
        # Track processing state
        self.picking_dockets_processed = False
        
        # Excel Generation data
        self.excel_selected_output_folder = ""
        self.excel_selected_pdf_files = []
        
        # Unified flow data
        self.internal_excel_data = []  # Store Excel data internally instead of generating file
        self.picking_sheet_files = []  # Store picking sheet PDF files
        
        
        # OCR Configuration - Multiple regions (hardcoded coordinates)
        self.ocr_regions = {
            'region_1': {'coordinates': [387, 765, 590, 795], 'color': 'red', 'name': 'Region 1'},
            'region_2': {'coordinates': [432, 44, 591, 65], 'color': 'blue', 'name': 'Region 2'},
            'region_3': {'coordinates': [23, 47, 326, 73], 'color': 'green', 'name': 'Region 3'},
            'region_4': {'coordinates': [28, 772, 183, 799], 'color': 'orange', 'name': 'Region 4'},
            'region_5': {'coordinates': [28, 72, 328, 92], 'color': 'purple', 'name': 'Region 5'}  # Hardcoded to address line location
        }
        self.region_5_save_location = 'Column K (Region 5 Data)'  # Default save location
        self.ocr_setup_completed = True  # Mark as completed since coordinates are hardcoded
        
        # Print hardcoded OCR configuration
        configured_regions = [region for region in self.ocr_regions.values() if region['coordinates']]
        print(f"OCR configuration loaded: {len(configured_regions)} regions configured (hardcoded)")
        for region in configured_regions:
            print(f"  {region['name']} ({region['color']}): {region['coordinates']}")
        
        # Initialize UI
        self.init_ui()
        self.apply_clean_styling()
        
        # Load existing data
        self.load_existing_delivery_data()
        
        # Load OCR configuration
        self.load_ocr_config()
        
        self.update_status("Ready")
    
    def init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle("Dispatch Scanning - Streamlined Processing")
        self.setGeometry(100, 100, 1000, 600)
        
        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 8)
        
        # Header
        header_frame = self.create_header()
        main_layout.addWidget(header_frame)
        
        # Content area - main processing interface
        main_processing_content = self.create_main_processing_content()
        main_layout.addWidget(main_processing_content)
        
        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.update_status("Ready")
        
        # Progress bar (initially hidden)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.status_bar.addPermanentWidget(self.progress_bar)
    
    def create_header(self):
        """Create modern application header"""
        header_frame = QFrame()
        header_frame.setObjectName("headerFrame")
        header_frame.setFixedHeight(50)
        
        layout = QHBoxLayout(header_frame)
        layout.setContentsMargins(16, 8, 16, 8)
        
        # Main title with icon
        title_container = QHBoxLayout()
        title_container.setSpacing(12)
        
        

        
        # Title and subtitle
        title_text_container = QVBoxLayout()
        title_text_container.setSpacing(4)
        
        title_label = QLabel("Dispatch Picking Sheet Upload")
        title_label.setObjectName("headerTitle")
        title_text_container.addWidget(title_label)
        

      
        
        
        title_container.addLayout(title_text_container)
        layout.addLayout(title_container)
        
        layout.addStretch()
        
        
        return header_frame
    
    def create_main_processing_content(self):
        """Create the main processing content with tabbed interface"""
        # Create tab widget
        tab_widget = QTabWidget()
        tab_widget.setObjectName("mainTabWidget")
        
        # Tab 1: Main Processing (existing functionality)
        main_processing_tab = self.create_main_processing_tab()
        tab_widget.addTab(main_processing_tab, "Main Processing")
        
        # Tab 2: Order Management (new functionality)
        order_management_tab = self.create_order_management_tab()
        tab_widget.addTab(order_management_tab, "Order Management")
        
        # Tab 3: Label Printing (new functionality)
        label_printing_tab = self.create_label_printing_tab()
        tab_widget.addTab(label_printing_tab, "Label Printing")
        
        # Tab 4: Print History (new functionality)
        print_history_tab = self.create_print_history_tab()
        tab_widget.addTab(print_history_tab, "Print History")
        
        return tab_widget
    
    def create_main_processing_tab(self):
        """Create the main processing tab with existing functionality"""
        # Content widget - no scroll area
        content_widget = QWidget()
        content_layout = QHBoxLayout(content_widget)
        content_layout.setSpacing(12)
        content_layout.setContentsMargins(12, 12, 12, 12)
        
        # Left column - File Selection
        left_column = self.create_unified_file_selection_column()
        content_layout.addWidget(left_column)
        
        # Right column - Processing Section
        right_column = self.create_unified_processing_column()
        content_layout.addWidget(right_column)
        
        # Set column proportions (50% left, 50% right)
        content_layout.setStretch(0, 5)
        content_layout.setStretch(1, 5)
        
        return content_widget
    
    def create_order_management_tab(self):
        """Create the Order Management tab with dispatch_orders table view"""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        layout.setSpacing(12)
        layout.setContentsMargins(12, 12, 12, 12)
        
        # Header section
        header_frame = QFrame()
        header_frame.setObjectName("orderManagementHeader")
        header_frame.setStyleSheet("""
            QFrame#orderManagementHeader {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin: 0px;
            }
        """)
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(16, 12, 16, 12)
        header_layout.setSpacing(10)
        
        # Title
        title_label = QLabel("Order Management")
        title_label.setObjectName("orderManagementTitle")
        title_label.setStyleSheet("""
            QLabel#orderManagementTitle {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
            }
        """)
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        
        # Refresh button
        refresh_button = QPushButton("Refresh Data")
        refresh_button.setObjectName("refreshButton")
        refresh_button.clicked.connect(self.refresh_order_data)
        header_layout.addWidget(refresh_button)
        layout.addWidget(header_frame)
        
        # Table section
        table_frame = QFrame()
        table_frame.setObjectName("orderTableFrame")
        table_layout = QVBoxLayout(table_frame)
        table_layout.setContentsMargins(0, 0, 0, 0)
        
        # Create table widget
        self.order_table = QTableWidget()
        self.order_table.setObjectName("orderTable")
        self.order_table.setAlternatingRowColors(True)
        self.order_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.order_table.setSelectionMode(QTableWidget.SingleSelection)
        self.order_table.setSortingEnabled(True)
        
        # Set table properties
        self.order_table.horizontalHeader().setStretchLastSection(True)
        self.order_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.order_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)  # Order Number
        self.order_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)  # Customer Type
        self.order_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)  # Picking Date
        self.order_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)  # Site Name (stretches)
        self.order_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Fixed)  # Route
        self.order_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.Fixed)  # PDF File Name
        self.order_table.verticalHeader().setVisible(False)
        
        table_layout.addWidget(self.order_table)
        layout.addWidget(table_frame)
        
        # Load initial data
        self.load_order_data()
        
        return tab_widget
    
    def load_order_data(self):
        """Load order data from Supabase dispatch_orders table"""
        if not SUPABASE_AVAILABLE:
            return
        
        try:
            # Get Supabase client
            supabase_client = get_supabase_client()
            
            # Fetch all data from dispatch_orders table
            result = supabase_client.table('dispatch_orders').select("*").order('created_at', desc=True).execute()
            
            if result.data:
                self.populate_order_table(result.data)
            else:
                self.order_table.setRowCount(0)
                self.order_table.setColumnCount(0)
                
        except Exception as e:
            print(f"Error loading order data: {e}")
    
    def populate_order_table(self, data):
        """Populate the order table with data from Supabase"""
        if not data:
            return
        
        # Define column mapping - only show requested columns
        columns = [
            'ordernumber', 'customer_type', 'created_at', 'sitename', 'route', 'pdf_file_name'
        ]
        
        # User-friendly column headers
        column_headers = [
            'Order Number', 'Customer Type', 'Picking Date', 'Site Name', 'Route', 'PDF File Name'
        ]
        
        # Set table dimensions
        self.order_table.setRowCount(len(data))
        self.order_table.setColumnCount(len(columns))
        self.order_table.setHorizontalHeaderLabels(column_headers)
        
        # Populate table with data
        for row_idx, record in enumerate(data):
            for col_idx, column in enumerate(columns):
                value = record.get(column, '')
                
                # Format the value for display
                if value is None:
                    display_value = ''
                elif isinstance(value, bool):
                    display_value = 'Yes' if value else 'No'
                elif isinstance(value, (dict, list)):
                    display_value = str(value)
                elif column == 'created_at' and value:
                    # Format date for better readability
                    try:
                        from datetime import datetime
                        if isinstance(value, str):
                            # Parse ISO format date
                            dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
                            display_value = dt.strftime('%d/%m/%Y')
                        else:
                            display_value = str(value)
                    except:
                        display_value = str(value)
                else:
                    display_value = str(value)
                
                # Create table item
                item = QTableWidgetItem(display_value)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Make read-only
                
                # Set special formatting for certain columns
                if column in ['created_at']:
                    item.setTextAlignment(Qt.AlignCenter)
                elif column in ['ordernumber', 'dispatchcode']:
                    item.setTextAlignment(Qt.AlignCenter)
                    if column == 'ordernumber':
                        item.setFont(QFont("Consolas", 11, QFont.Weight.Bold))  # Make order numbers more prominent
                    else:
                        item.setFont(QFont("Consolas", 10, QFont.Weight.Normal))  # Dispatch code in monospace
                
                self.order_table.setItem(row_idx, col_idx, item)
        
        # Resize columns to content
        self.order_table.resizeColumnsToContents()
        
        # Set appropriate column widths for better readability
        column_widths = {
            'ordernumber': 200,      # Order Number - increased width for better visibility
            'customer_type': 150,    # Customer Type - wider for text content
            'created_at': 140,       # Picking Date - needs space for date/time format
            'sitename': 400,         # Site Name - increased width to show all text
            'route': 120,            # Route - similar to order number
            'pdf_file_name': 250     # PDF File Name - wider for file names
        }
        
        for col in range(self.order_table.columnCount()):
            column_name = columns[col]
            if column_name in column_widths:
                self.order_table.setColumnWidth(col, column_widths[column_name])
            else:
                # Fallback minimum width
                current_width = self.order_table.columnWidth(col)
                self.order_table.setColumnWidth(col, max(current_width, 120))
        
        # Ensure the table stretches to fill available space
        self.order_table.horizontalHeader().setStretchLastSection(True)
    
    def refresh_order_data(self):
        """Refresh the order data from Supabase"""
        self.load_order_data()
    
    def create_label_printing_tab(self):
        """Create the Label Printing tab with barcode scanner and Zebra printer functionality"""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        layout.setSpacing(12)
        layout.setContentsMargins(12, 12, 12, 12)
        
        # Header section - matching Order Management style
        header_frame = QFrame()
        header_frame.setObjectName("labelPrintingHeader")
        header_frame.setStyleSheet("""
            QFrame#labelPrintingHeader {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin: 0px;
            }
        """)
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(16, 12, 16, 12)
        header_layout.setSpacing(10)
        
        # Title
        title_label = QLabel("Label Printing")
        title_label.setObjectName("labelPrintingTitle")
        title_label.setStyleSheet("""
            QLabel#labelPrintingTitle {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
            }
        """)
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        
        # Printer status
        self.printer_status_label = QLabel("Connected to: Not Connected")
        self.printer_status_label.setObjectName("printerStatusLabel")
        self.printer_status_label.setStyleSheet("""
            QLabel#printerStatusLabel {
                color: #e74c3c;
                font-weight: bold;
                font-size: 12px;
            }
        """)
        header_layout.addWidget(self.printer_status_label)
        
        # Printer scanning status indicator
        self.printer_scanning_label = QLabel("🔍 Scanning...")
        self.printer_scanning_label.setObjectName("printerScanningLabel")
        self.printer_scanning_label.setStyleSheet("""
            QLabel#printerScanningLabel {
                color: #f39c12;
                font-weight: bold;
                font-size: 12px;
            }
        """)
        header_layout.addWidget(self.printer_scanning_label)
        
        layout.addWidget(header_frame)
        
        # Main content area - matching Order Management style
        content_frame = QFrame()
        content_frame.setObjectName("labelPrintingContent")
        content_layout = QHBoxLayout(content_frame)
        content_layout.setSpacing(12)
        content_layout.setContentsMargins(0, 0, 0, 0)
        
        # Left column - Scanner and Order Info
        left_column = QFrame()
        left_column.setObjectName("leftColumn")
        left_layout = QVBoxLayout(left_column)
        left_layout.setSpacing(12)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        # Barcode Scanner Section
        scanner_group = QGroupBox("Barcode Scanner")
        scanner_group.setObjectName("scannerGroup")
        scanner_group.setStyleSheet("""
            QGroupBox#scannerGroup {
                font-weight: bold;
                border: 2px solid #e2e8f0;
                border-radius: 6px;
                margin-top: 0px;
                padding-top: 10px;
            }
            QGroupBox#scannerGroup::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        scanner_layout = QVBoxLayout(scanner_group)
        scanner_layout.setSpacing(8)
        
        # Order number input (for barcode scanner)
        order_input_label = QLabel("Scan Order Number")
        order_input_label.setObjectName("orderInputLabel")
        order_input_label.setStyleSheet("""
            QLabel#orderInputLabel {
                font-weight: bold;
                color: #27ae60;
            }
        """)
        scanner_layout.addWidget(order_input_label)
        
        self.order_number_input = QLineEdit()
        self.order_number_input.setObjectName("orderNumberInput")
        self.order_number_input.setPlaceholderText("Click here to begin scanning")
        # Connect returnPressed signal for manual entry
        self.order_number_input.returnPressed.connect(self.on_order_number_entered)
        scanner_layout.addWidget(self.order_number_input)
        
        # Order info display
        self.order_info_label = QLabel("No order selected")
        self.order_info_label.setObjectName("orderInfoLabel")
        self.order_info_label.setWordWrap(True)
        self.order_info_label.setStyleSheet("""
            QLabel#orderInfoLabel {
                background-color: #f8f9fa;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                padding: 8px;
                min-height: 60px;
            }
        """)
        scanner_layout.addWidget(self.order_info_label)
        
        left_layout.addWidget(scanner_group)
        left_layout.addStretch()
        
        content_layout.addWidget(left_column)
        
        # Right column - Printer Status and Log
        right_column = QFrame()
        right_column.setObjectName("rightColumn")
        right_layout = QVBoxLayout(right_column)
        right_layout.setSpacing(12)
        
        # Print Log Section
        log_group = QGroupBox("Print Log")
        log_group.setObjectName("logGroup")
        log_group.setStyleSheet("""
            QGroupBox#logGroup {
                font-weight: bold;
                border: 2px solid #e2e8f0;
                border-radius: 6px;
                margin-top: 0px;
                padding-top: 10px;
            }
            QGroupBox#logGroup::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
            }
        """)
        log_layout = QVBoxLayout(log_group)
        log_layout.setSpacing(8)
        
        self.print_log = QTextEdit()
        self.print_log.setObjectName("printLog")
        self.print_log.setMaximumHeight(200)
        self.print_log.setReadOnly(True)
        self.print_log.setStyleSheet("""
            QTextEdit#printLog {
                background-color: #f8f9fa;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                font-family: 'Consolas', monospace;
                font-size: 11px;
                color: black;
            }
        """)
        log_layout.addWidget(self.print_log)
        
        right_layout.addWidget(log_group)
        right_layout.addStretch()
        
        content_layout.addWidget(right_column)
        layout.addWidget(content_frame)
        
        # Initialize printer connection
        self.zebra_printer = None
        self.current_order_data = None
        
        # Auto-connect to printer
        self.auto_connect_printer()
        
        # Set up continuous printer scanning
        self.setup_printer_scanning()
        
        # Set up barcode scanner monitoring
        self.setup_barcode_scanner()
        
        return tab_widget
    
    def create_print_history_tab(self):
        """Create the Print History tab to show print records"""
        tab_widget = QWidget()
        layout = QVBoxLayout(tab_widget)
        layout.setSpacing(12)
        layout.setContentsMargins(12, 12, 12, 12)
        
        # Header section
        header_frame = QFrame()
        header_frame.setObjectName("printHistoryHeader")
        header_frame.setStyleSheet("""
            QFrame#printHistoryHeader {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                margin: 0px;
            }
        """)
        header_layout = QHBoxLayout(header_frame)
        header_layout.setSpacing(10)
        
        # Title
        title_label = QLabel("Print History")
        title_label.setObjectName("printHistoryTitle")
        title_label.setStyleSheet("""
            QLabel#printHistoryTitle {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
            }
        """)
        header_layout.addWidget(title_label)
        
        header_layout.addStretch()
        
        # Date filter button
        self.date_filter_button = QPushButton("Filter by Date")
        self.date_filter_button.setObjectName("dateFilterButton")
        self.date_filter_button.setStyleSheet("""
            QPushButton#dateFilterButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton#dateFilterButton:hover {
                background-color: #2980b9;
            }
        """)
        self.date_filter_button.clicked.connect(self.show_date_filter_dialog)
        header_layout.addWidget(self.date_filter_button)
        
        # Refresh button
        self.refresh_history_button = QPushButton("Refresh")
        self.refresh_history_button.setObjectName("refreshHistoryButton")
        self.refresh_history_button.setStyleSheet("""
            QPushButton#refreshHistoryButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton#refreshHistoryButton:hover {
                background-color: #2980b9;
            }
        """)
        self.refresh_history_button.clicked.connect(self.refresh_print_history)
        header_layout.addWidget(self.refresh_history_button)
        layout.addWidget(header_frame)
        
        # Print history table
        self.print_history_table = QTableWidget()
        self.print_history_table.setObjectName("printHistoryTable")
        self.print_history_table.setAlternatingRowColors(True)
        self.print_history_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.print_history_table.setSortingEnabled(True)
        
        # Set table headers
        headers = ["Order Number", "Site Name", "Crate Quantity", "Route", "Printed At"]
        self.print_history_table.setColumnCount(len(headers))
        self.print_history_table.setHorizontalHeaderLabels(headers)
        
        # Configure table appearance
        self.print_history_table.horizontalHeader().setStretchLastSection(True)
        
        # Auto-resize columns to fit content
        self.print_history_table.resizeColumnsToContents()
        self.print_history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        
        # Set minimum column widths to ensure readability
        self.print_history_table.setColumnWidth(0, 120)  # Order Number
        self.print_history_table.setColumnWidth(1, 200)  # Site Name
        self.print_history_table.setColumnWidth(2, 100)  # Crate Quantity
        self.print_history_table.setColumnWidth(3, 100)  # Route
        self.print_history_table.setColumnWidth(4, 150)  # Printed At
        
        self.print_history_table.setStyleSheet("""
            QTableWidget#printHistoryTable {
                gridline-color: #e0e0e0;
                background-color: white;
                alternate-background-color: #f8f9fa;
                selection-background-color: #e3f2fd;
                border: 1px solid #e0e0e0;
                border-radius: 6px;
            }
            QTableWidget#printHistoryTable::item {
                padding: 8px;
                border: none;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                padding: 8px;
                border: none;
                border-right: 1px solid #e0e0e0;
                font-weight: bold;
            }
        """)
        
        layout.addWidget(self.print_history_table)
        
        # Load initial data
        self.load_print_history()
        
        return tab_widget
    
    def load_print_history(self):
        """Load print history from Supabase"""
        if not SUPABASE_AVAILABLE:
            self.print_history_table.setRowCount(0)
            return
        
        try:
            # Get Supabase client
            supabase_client = get_supabase_client()
            
            # Query the print_history table
            response = supabase_client.table('print_history').select('*').order('printed_at', desc=True).execute()
            
            if response.data:
                self.print_history_table.setRowCount(len(response.data))
                
                for row_idx, record in enumerate(response.data):
                    # Order Number
                    self.print_history_table.setItem(row_idx, 0, QTableWidgetItem(record.get('order_number', 'N/A')))
                    
                    # Site Name
                    self.print_history_table.setItem(row_idx, 1, QTableWidgetItem(record.get('site_name', 'N/A')))
                    
                    # Crate Quantity
                    self.print_history_table.setItem(row_idx, 2, QTableWidgetItem(str(record.get('crate_quantity', 0))))
                    
                    # Route
                    self.print_history_table.setItem(row_idx, 3, QTableWidgetItem(record.get('route', 'N/A')))
                    
                    # Printed At
                    printed_at = record.get('printed_at', '')
                    if printed_at:
                        # Format timestamp for display (dd/MM/yyyy)
                        try:
                            from datetime import datetime
                            dt = datetime.fromisoformat(printed_at.replace('Z', '+00:00'))
                            formatted_time = dt.strftime('%d/%m/%Y')
                        except:
                            formatted_time = printed_at
                    else:
                        formatted_time = 'N/A'
                    self.print_history_table.setItem(row_idx, 4, QTableWidgetItem(formatted_time))
                
                # Auto-resize columns after data is loaded
                self.print_history_table.resizeColumnsToContents()
            else:
                self.print_history_table.setRowCount(0)
                
        except Exception as e:
            print(f"Error loading print history: {e}")
            self.print_history_table.setRowCount(0)
    
    def refresh_print_history(self):
        """Refresh the print history table"""
        self.load_print_history()
    
    def show_date_filter_dialog(self):
        """Show date filter dialog for print history"""
        dialog = DateFilterDialog(self)
        if dialog.exec() == QDialog.Accepted:
            start_date, end_date = dialog.get_date_range()
            self.filter_print_history_by_date(start_date, end_date)
    
    def filter_print_history_by_date(self, start_date, end_date):
        """Filter print history by date range"""
        if not SUPABASE_AVAILABLE:
            self.print_history_table.setRowCount(0)
            return
        
        try:
            # Get Supabase client
            supabase_client = get_supabase_client()
            
            # Build query with date filter
            query = supabase_client.table('print_history').select('*')
            
            if start_date:
                # Convert to ISO format for Supabase
                start_iso = start_date.toString('yyyy-MM-dd') + 'T00:00:00Z'
                query = query.gte('printed_at', start_iso)
            
            if end_date:
                # Convert to ISO format for Supabase (end of day)
                end_iso = end_date.toString('yyyy-MM-dd') + 'T23:59:59Z'
                query = query.lte('printed_at', end_iso)
            
            # Execute query
            response = query.order('printed_at', desc=True).execute()
            
            if response.data:
                self.print_history_table.setRowCount(len(response.data))
                
                for row_idx, record in enumerate(response.data):
                    # Order Number
                    self.print_history_table.setItem(row_idx, 0, QTableWidgetItem(record.get('order_number', 'N/A')))
                    
                    # Site Name
                    self.print_history_table.setItem(row_idx, 1, QTableWidgetItem(record.get('site_name', 'N/A')))
                    
                    # Crate Quantity
                    self.print_history_table.setItem(row_idx, 2, QTableWidgetItem(str(record.get('crate_quantity', 0))))
                    
                    # Route
                    self.print_history_table.setItem(row_idx, 3, QTableWidgetItem(record.get('route', 'N/A')))
                    
                    # Printed At
                    printed_at = record.get('printed_at', '')
                    if printed_at:
                        # Format timestamp for display (dd/MM/yyyy)
                        try:
                            from datetime import datetime
                            dt = datetime.fromisoformat(printed_at.replace('Z', '+00:00'))
                            formatted_time = dt.strftime('%d/%m/%Y')
                        except:
                            formatted_time = printed_at
                    else:
                        formatted_time = 'N/A'
                    self.print_history_table.setItem(row_idx, 4, QTableWidgetItem(formatted_time))
                
                # Auto-resize columns after data is loaded
                self.print_history_table.resizeColumnsToContents()
            else:
                self.print_history_table.setRowCount(0)
                
        except Exception as e:
            print(f"Error filtering print history: {e}")
            self.print_history_table.setRowCount(0)
    
    def record_print_event(self, order_number, site_name, crate_quantity, route, printed_by):
        """Record a print event in the database"""
        if not SUPABASE_AVAILABLE:
            print("Supabase not available - cannot record print event")
            return
        
        try:
            from datetime import datetime
            
            # Get Supabase client
            supabase_client = get_supabase_client()
            
            # Prepare the record data
            record_data = {
                'order_number': order_number,
                'site_name': site_name,
                'crate_quantity': crate_quantity,
                'route': route,
                'printed_by': printed_by,
                'printed_at': datetime.now().isoformat()
            }
            
            # Insert the record
            response = supabase_client.table('print_history').insert(record_data).execute()
            
            if response.data:
                print(f"Print event recorded: {order_number} - {crate_quantity} crates")
            else:
                print("Failed to record print event")
                
        except Exception as e:
            print(f"Error recording print event: {e}")
    
    def auto_connect_printer(self):
        """Automatically connect to printer on startup"""
        self.log_print_message("Auto-connecting to printer...")
        self.connect_printer()
    
    def setup_printer_scanning(self):
        """Set up continuous printer scanning"""
        # Create a timer to continuously scan for printers
        self.printer_scan_timer = QTimer()
        self.printer_scan_timer.timeout.connect(self.scan_for_printer)
        self.printer_scan_timer.start(5000)  # Check every 5 seconds
        
        self.log_print_message("Continuous printer scanning started")
    
    def scan_for_printer(self):
        """Continuously scan for available printers"""
        # Only scan if we don't have a printer connected
        if not self.zebra_printer:
            self.log_print_message("Scanning for printer...")
            self.connect_printer()
        else:
            # Verify the current printer is still available
            if not self.verify_printer_connection():
                self.log_print_message("Printer connection lost, scanning for new printer...")
                self.zebra_printer = None
                self.update_printer_status("Not Connected", "#e74c3c")
                self.connect_printer()
    
    def verify_printer_connection(self):
        """Verify that the current printer is still connected and responsive"""
        try:
            if self.zebra_printer and not self.zebra_printer.startswith("USB:"):
                # For system printers, check if they're still available
                return self.test_printer_connectivity(self.zebra_printer)
            elif self.zebra_printer and self.zebra_printer.startswith("USB:"):
                # For USB printers, test the connection more thoroughly
                port = self.zebra_printer.replace("USB:", "")
                try:
                    ser = serial.Serial(port, 9600, timeout=2)
                    
                    # Send a status query to verify it's still responsive
                    status_query = "~HS\r\n"
                    ser.write(status_query.encode('utf-8'))
                    time.sleep(0.5)  # Give printer time to respond
                    
                    # Try to read response
                    response = ser.read(100).decode('utf-8', errors='ignore')
                    ser.close()
                    
                    # Check if we got a valid response
                    if len(response) > 0 and ('ZEBRA' in response.upper() or 'ZT' in response.upper() or 'ZPL' in response.upper()):
                        return True
                    else:
                        return False
                except:
                    return False
            return False
        except Exception as e:
            self.log_print_message(f"Error verifying printer connection: {str(e)}")
            return False
    
    def update_printer_status(self, status_text, color):
        """Update printer status display"""
        self.printer_status_label.setText(f"Connected to: {status_text}")
        self.printer_status_label.setStyleSheet(f"""
            QLabel#printerStatusLabel {{
                color: {color};
                font-weight: bold;
                font-size: 12px;
            }}
        """)
        
        # Update scanning indicator
        if status_text == "Not Connected":
            self.printer_scanning_label.setText("🔍 Scanning...")
            self.printer_scanning_label.setStyleSheet("""
                QLabel#printerScanningLabel {
                    color: #f39c12;
                    font-weight: bold;
                    font-size: 12px;
                }
            """)
        else:
            self.printer_scanning_label.setText("✅ Connected")
            self.printer_scanning_label.setStyleSheet("""
                QLabel#printerScanningLabel {
                    color: #27ae60;
                    font-weight: bold;
                    font-size: 12px;
                }
            """)
    
    def setup_barcode_scanner(self):
        """Set up continuous barcode scanner monitoring"""
        # Create a timer to check for barcode input
        self.scanner_timer = QTimer()
        self.scanner_timer.timeout.connect(self.check_barcode_input)
        self.scanner_timer.start(100)  # Check every 100ms
        
        # Store previous input to detect new scans
        self.previous_input = ""
        self.processing_scan = False  # Flag to prevent duplicate processing
        self.current_processing_order = None  # Track currently processing order
        self.dialog_open = False  # Flag to prevent multiple dialogs
        
        # Barcode input stabilization
        self.input_stable_timer = QTimer()
        self.input_stable_timer.timeout.connect(self.process_stable_input)
        self.input_stable_timer.setSingleShot(True)
        self.pending_input = ""  # Store input waiting to be processed
        
        self.log_print_message("Barcode scanner monitoring started - scanning in background")
    
    def check_barcode_input(self):
        """Check for new barcode input with stabilization"""
        current_input = self.order_number_input.text().strip()
        
        # If input has changed and is not empty, start stabilization timer
        if current_input != self.previous_input and current_input and not self.processing_scan:
            self.pending_input = current_input
            self.previous_input = current_input
            # Restart the stabilization timer (300ms delay)
            self.input_stable_timer.start(300)
    
    def process_stable_input(self):
        """Process input after it has stabilized (no changes for 300ms)"""
        if self.pending_input and not self.processing_scan:
            self.processing_scan = True
            self.process_barcode_scan(self.pending_input)
    
    def process_barcode_scan(self, order_number):
        """Process a scanned barcode"""
        self.log_print_message(f"Barcode scanned: {order_number}")
        
        # Store the order number we're processing to prevent duplicate processing
        self.current_processing_order = order_number
        
        # Clear the input field after a delay to allow scanner to finish inputting
        QTimer.singleShot(500, self.order_number_input.clear)  # 500ms delay
        
        # Fetch order data
        self.fetch_order_data(order_number)
        
        # Reset the processing flag and clear current order after a longer delay
        QTimer.singleShot(1000, self.reset_processing_flags)
    
    def reset_processing_flags(self):
        """Reset processing flags after scan is complete"""
        self.processing_scan = False
        self.current_processing_order = None
    
    def highlight_order_input(self, is_processing):
        """Provide visual feedback for order input processing"""
        if is_processing:
            self.order_number_input.setStyleSheet("""
                QLineEdit#orderNumberInput {
                    border: 2px solid #f39c12;
                    background-color: #fefbf3;
                }
            """)
        else:
            self.order_number_input.setStyleSheet("""
                QLineEdit#orderNumberInput {
                    border: 1px solid #d1d5db;
                    background-color: white;
                }
                QLineEdit#orderNumberInput:focus {
                    border-color: #27ae60;
                }
            """)
    
    def on_order_number_entered(self):
        """Handle order number input from barcode scanner or manual entry"""
        order_number = self.order_number_input.text().strip()
        if not order_number:
            return
        
        # Skip if we're already processing this same order number
        if hasattr(self, 'current_processing_order') and order_number == self.current_processing_order:
            self.log_print_message(f"Skipping duplicate processing of order: {order_number}")
            return
        
        # Skip if we're already processing a scan
        if self.processing_scan:
            return
        
        # Visual feedback - highlight the input field
        self.highlight_order_input(True)
        
        self.log_print_message(f"Scanning order number: {order_number}")
        self.fetch_order_data(order_number)
    
    def fetch_order_data(self, order_number):
        """Fetch order data from Supabase using order number"""
        if not SUPABASE_AVAILABLE:
            self.order_info_label.setText("Error: Supabase not available")
            self.log_print_message("Error: Supabase not available")
            return
        
        try:
            # Get Supabase client
            supabase_client = get_supabase_client()
            
            # Store original case for display
            original_order_number = order_number
            # Convert to uppercase for database search (database likely stores in uppercase)
            search_order_number = order_number.upper()
            
            # Query for the specific order
            result = supabase_client.table('dispatch_orders').select(
                "ordernumber, sitename, route, customer_type, created_at, dispatchcode"
            ).eq('ordernumber', search_order_number).execute()
            
            if result.data and len(result.data) > 0:
                order_data = result.data[0]
                # Store the original case order number for display and printing
                order_data['ordernumber'] = original_order_number
                self.current_order_data = order_data
                
                # Display order info
                site_name = order_data.get('sitename', 'N/A')
                route = order_data.get('route', 'N/A')
                customer_type = order_data.get('customer_type', 'N/A')
                created_at = order_data.get('created_at', 'N/A')
                dispatchcode = order_data.get('dispatchcode', 'N/A')
                
                # Format date
                if created_at != 'N/A':
                    try:
                        from datetime import datetime
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        created_at = dt.strftime('%d/%m/%Y')
                    except:
                        pass
                
                order_info = f"""Order: {original_order_number}
Dispatch Code: {dispatchcode}
Site: {site_name}
Route: {route}
Customer: {customer_type}
Date: {created_at}"""
                
                self.order_info_label.setText(order_info)
                self.log_print_message(f"Order found: {original_order_number} - {site_name}")
                
                # Reset visual feedback
                self.highlight_order_input(False)
                
                # Show crate count dialog immediately after scanning
                self.show_crate_count_dialog()
                
            else:
                self.order_info_label.setText(f"Order {original_order_number} not found in database")
                self.current_order_data = None
                self.log_print_message(f"Order {original_order_number} not found")
                
                # Reset visual feedback
                self.highlight_order_input(False)
                
        except Exception as e:
            self.order_info_label.setText(f"Error fetching order data: {str(e)}")
            self.current_order_data = None
            self.log_print_message(f"Error fetching order data: {str(e)}")
            
            # Reset visual feedback
            self.highlight_order_input(False)
    
    def connect_printer(self):
        """Connect to Zebra ZT411 printer"""
        try:
            # Get available printers using Windows wmic command
            result = subprocess.run(['wmic', 'printer', 'get', 'name'], 
                                  capture_output=True, text=True, shell=True)
            
            if result.returncode == 0:
                printers = [line.strip() for line in result.stdout.split('\n') 
                           if line.strip() and line.strip() != 'Name']
                
                # Look for Zebra printer with expanded search patterns
                zebra_printer = None
                for printer in printers:
                    printer_upper = printer.upper()
                    if ('ZT411' in printer_upper or 
                        'ZT421' in printer_upper or
                        'ZT231' in printer_upper or
                        'ZT230' in printer_upper or
                        'ZT220' in printer_upper or
                        'ZT420' in printer_upper or
                        'ZD420' in printer_upper or
                        'ZD620' in printer_upper or
                        'ZEBRA' in printer_upper or 
                        'ZDESIGNER' in printer_upper or
                        'ZPL' in printer_upper or
                        'ZEBRA ZT' in printer_upper or
                        'ZEBRA ZD' in printer_upper or
                        'ZEBRA GC' in printer_upper or
                        'ZEBRA GK' in printer_upper):
                        zebra_printer = printer
                        break
                
                if zebra_printer:
                    # Test if the printer is actually available and responsive
                    if self.test_printer_connectivity(zebra_printer):
                        self.zebra_printer = zebra_printer
                        self.update_printer_status(zebra_printer, "#27ae60")
                        self.log_print_message(f"Successfully connected to Zebra printer: {zebra_printer}")
                        return
                    else:
                        self.log_print_message(f"Found Zebra printer '{zebra_printer}' but it's not responding")
                else:
                    self.log_print_message("No Zebra printer found in system printers")
            else:
                self.log_print_message(f"Could not get printer list: {result.stderr}")
            
            # If no Zebra printer found in system, try USB connection with Zebra verification
            self.try_usb_connection()
                
        except Exception as e:
            self.log_print_message(f"Printer detection failed: {str(e)}")
            self.try_usb_connection()
    
    def test_printer_connectivity(self, printer_name):
        """Test if a system printer is actually available and responsive"""
        try:
            import win32print
            import win32api
            
            # Get printer handle
            handle = win32print.OpenPrinter(printer_name)
            if handle:
                try:
                    # Get printer info to test connectivity
                    printer_info = win32print.GetPrinter(handle, 2)
                    
                    # Check if printer is online and ready
                    if printer_info and printer_info.get('Status') == 0:  # 0 means ready
                        # Additional test: Try to send a small test job to verify physical connectivity
                        try:
                            # Create a minimal ZPL test command
                            test_zpl = "^XA^XZ"  # Minimal ZPL command that should be accepted by any Zebra printer
                            
                            # Try to send the test command
                            job_id = win32print.StartDocPrinter(handle, 1, ("Test", None, "RAW"))
                            if job_id:
                                win32print.StartPagePrinter(handle)
                                win32print.WritePrinter(handle, test_zpl.encode('utf-8'))
                                win32print.EndPagePrinter(handle)
                                win32print.EndDocPrinter(handle)
                                win32print.ClosePrinter(handle)
                                return True
                            else:
                                win32print.ClosePrinter(handle)
                                return False
                        except Exception as test_error:
                            # If test print fails, printer is likely not physically connected
                            self.log_print_message(f"Printer '{printer_name}' test print failed: {str(test_error)}")
                            win32print.ClosePrinter(handle)
                            return False
                    else:
                        win32print.ClosePrinter(handle)
                        self.log_print_message(f"Printer '{printer_name}' is not ready (Status: {printer_info.get('Status', 'Unknown') if printer_info else 'No info'})")
                        return False
                        
                except Exception as e:
                    self.log_print_message(f"Error testing printer connectivity: {str(e)}")
                    try:
                        win32print.ClosePrinter(handle)
                    except:
                        pass
                    return False
            else:
                self.log_print_message(f"Could not open printer '{printer_name}'")
                return False
                
        except ImportError:
            # Fallback method if win32print is not available
            self.log_print_message("win32print not available, using fallback connectivity test")
            return self.test_printer_connectivity_fallback(printer_name)
        except Exception as e:
            self.log_print_message(f"Printer connectivity test failed: {str(e)}")
            return False
    
    def test_printer_connectivity_fallback(self, printer_name):
        """Fallback method to test printer connectivity without win32print"""
        try:
            # Use Windows command to check printer status
            result = subprocess.run(['wmic', 'printer', 'where', f'name="{printer_name}"', 'get', 'workoffline,printerstatus'], 
                                  capture_output=True, text=True, shell=True)
            
            if result.returncode == 0:
                output = result.stdout.lower()
                # Check if printer is offline
                if 'workoffline' in output and 'true' in output:
                    self.log_print_message(f"Printer '{printer_name}' is offline")
                    return False
                # Check printer status (3 = ready, 4 = printing, 5 = error)
                elif 'printerstatus' in output:
                    if '3' in output or '4' in output:  # Ready or printing
                        return True
                    else:
                        self.log_print_message(f"Printer '{printer_name}' status indicates not ready")
                        return False
                else:
                    # If we can't determine status, be conservative and assume not available
                    self.log_print_message(f"Could not determine printer '{printer_name}' status")
                    return False
            else:
                self.log_print_message(f"Could not check printer status: {result.stderr}")
                return False
                
        except Exception as e:
            self.log_print_message(f"Fallback connectivity test failed: {str(e)}")
            return False
    
    def try_usb_connection(self):
        """Try to connect via USB port and verify it's a Zebra printer"""
        try:
            # Try common USB ports for Zebra printers (expanded range)
            usb_ports = ['COM1', 'COM2', 'COM3', 'COM4', 'COM5', 'COM6', 'COM7', 'COM8', 'COM9', 'COM10', 'COM11', 'COM12', 'COM13', 'COM14', 'COM15', 'COM16']
            
            for port in usb_ports:
                try:
                    # Test if port is available and verify it's a Zebra printer
                    ser = serial.Serial(port, 9600, timeout=2)
                    
                    # Send a status query to verify it's a Zebra printer
                    # ZPL command to get printer status
                    status_query = "~HS\r\n"
                    ser.write(status_query.encode('utf-8'))
                    time.sleep(0.5)  # Give printer time to respond
                    
                    # Try to read response
                    response = ser.read(100).decode('utf-8', errors='ignore')
                    ser.close()
                    
                    # Check if response indicates a Zebra printer (be more strict)
                    if ('ZEBRA' in response.upper() or 
                        'ZT' in response.upper() or
                        'ZPL' in response.upper() or
                        'ZDESIGNER' in response.upper()):
                        
                        # Store the port for printing
                        self.zebra_printer = f"USB:{port}"
                        
                        self.update_printer_status(f"Zebra USB {port}", "#27ae60")
                        self.log_print_message(f"Successfully connected to Zebra printer via USB: {port}")
                        return
                    
                except Exception as port_error:
                    # Port might be in use or not a printer, continue to next port
                    continue
            
            # If no Zebra printer found on USB ports, try to find any available printer
            self.try_fallback_usb_connection()
            
        except Exception as e:
            self.update_printer_status("Connection Failed", "#e74c3c")
            self.log_print_message(f"USB connection failed: {str(e)}")
            self.zebra_printer = None
    
    def try_fallback_usb_connection(self):
        """Fallback method - do not connect to unverified ports"""
        self.log_print_message("No verified Zebra printer found on any USB ports")
        self.update_printer_status("No Printer Found", "#e74c3c")
        self.zebra_printer = None
    
    def show_crate_count_dialog(self):
        """Show crate count dialog after scanning order number"""
        if not self.current_order_data:
            QMessageBox.warning(self, "Order Error", "No order data available.")
            return
        
        # Check if dialog is already open
        if self.dialog_open:
            self.log_print_message("Dialog already open, skipping duplicate")
            return
        
        # Set dialog flag
        self.dialog_open = True
        
        # Create and show the crate count dialog
        dialog = CrateCountDialog(self)
        dialog.setWindowTitle(f"Print Labels for Order {self.current_order_data.get('ordernumber', 'N/A')}")
        
        # Center the dialog on screen
        dialog.move(self.x() + (self.width() - dialog.width()) // 2, 
                   self.y() + (self.height() - dialog.height()) // 2)
        
        if dialog.exec() == QDialog.Accepted:
            crate_count = dialog.get_crate_count()
            self.log_print_message(f"Printing {crate_count} labels for order {self.current_order_data.get('ordernumber', 'N/A')}")
            self.print_labels_with_count(crate_count)
            
            # Move cursor back to order number input ONLY after successful printing
            self.order_number_input.setFocus()
            self.order_number_input.clear()
            self.log_print_message("Ready for next scan - cursor positioned in order number field")
        else:
            self.log_print_message("Print cancelled by user")
            # Do NOT move cursor back if cancelled - user can manually click if needed
        
        # Reset dialog flag
        self.dialog_open = False
    
    
    def print_labels_with_count(self, crate_count):
        """Print labels with specified crate count"""
        if not self.zebra_printer:
            QMessageBox.warning(self, "Printer Error", "Please connect to printer first.")
            return
        
        if not self.current_order_data:
            self.log_print_message("Error: No order data available")
            QMessageBox.warning(self, "Order Error", "Please scan an order number first.")
            return
        
        # Check if more than 15 labels are being printed
        if crate_count > 15:
            reply = QMessageBox.question(
                self, 
                "Confirm Label Count", 
                f"You are trying to print {crate_count} labels. Is 15 labels the correct required amount?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                crate_count = 15  # Limit to 15 labels
                self.log_print_message(f"Label count limited to 15 as confirmed by user")
            else:
                self.log_print_message(f"Print cancelled by user - {crate_count} labels requested")
            return
        
        try:
            order_number = self.current_order_data.get('ordernumber', 'N/A')
            site_name = self.current_order_data.get('sitename', 'N/A')
            route = self.current_order_data.get('route', 'N/A')
            dispatchcode = self.current_order_data.get('dispatchcode', 'N/A')
            
            if not order_number or order_number == 'N/A':
                raise Exception("Order number is missing or invalid")
            
            self.log_print_message(f"Printing {crate_count} labels for order {order_number}")
            
            # Print each label
            for i in range(1, crate_count + 1):
                zpl_code = self.generate_label_zpl(order_number, site_name, route, i, crate_count, dispatchcode)
                self.send_to_printer(zpl_code)
                self.log_print_message(f"Printed label {i}/{crate_count}")
            
            self.log_print_message(f"Successfully printed {crate_count} labels for order {order_number}")
            
            # Record the print event in the database
            import os
            printed_by = os.getenv('USERNAME', 'Unknown User')
            self.record_print_event(order_number, site_name, crate_count, route, printed_by)
            
        except Exception as e:
            error_msg = f"Print error: {str(e)}"
            self.log_print_message(error_msg)
            QMessageBox.critical(self, "Print Error", error_msg)
    
    def print_labels(self):
        """Print labels for the current order using spinbox value"""
        if not self.current_order_data:
            self.log_print_message("Error: No order data available")
            QMessageBox.warning(self, "Order Error", "Please scan an order number first.")
            return
        
        crate_count = 1  # Default to 1 crate since UI element was removed
        self.print_labels_with_count(crate_count)
    
    def send_to_printer(self, zpl_code):
        """Send ZPL code to the printer"""
        try:
            if self.zebra_printer.startswith("USB:"):
                # USB connection - send via serial port
                port = self.zebra_printer.replace("USB:", "")
                with serial.Serial(port, 9600, timeout=5) as ser:
                    ser.write(zpl_code.encode('utf-8'))
                    time.sleep(0.5)  # Give printer time to process
            else:
                # Windows printer - try multiple methods
                try:
                    # Method 1: Try Windows API (most reliable)
                    self.send_via_windows_api(zpl_code)
                except Exception as api_error:
                    self.log_print_message(f"Windows API method failed: {str(api_error)}")
                    
                    # Method 2: Try PowerShell
                    try:
                        self.send_via_powershell(zpl_code)
                    except Exception as ps_error:
                        self.log_print_message(f"PowerShell method failed: {str(ps_error)}")
                        
                        # Method 3: Try file-based approach
                        self.send_via_file(zpl_code)
                        
        except Exception as e:
            raise Exception(f"Failed to send to printer: {str(e)}")
    
    def send_via_windows_api(self, zpl_code):
        """Send ZPL using Windows API"""
        try:
            # Open printer
            hprinter = win32print.OpenPrinter(self.zebra_printer)
            
            try:
                # Start document
                job_info = win32print.StartDocPrinter(hprinter, 1, ("ZPL Label", None, "RAW"))
                
                try:
                    # Start page
                    win32print.StartPagePrinter(hprinter)
                    
                    # Write ZPL data
                    win32print.WritePrinter(hprinter, zpl_code.encode('utf-8'))
                    
                    # End page and document
                    win32print.EndPagePrinter(hprinter)
                    win32print.EndDocPrinter(hprinter)
                    
                except Exception as e:
                    win32print.AbortPrinter(hprinter)
                    raise e
                    
            finally:
                win32print.ClosePrinter(hprinter)
                
        except Exception as e:
            raise Exception(f"Windows API failed: {str(e)}")
    
    def send_via_powershell(self, zpl_code):
        """Send ZPL using PowerShell"""
        try:
            # Escape the ZPL code for PowerShell
            escaped_zpl = zpl_code.replace('"', '`"').replace('$', '`$')
            ps_command = f'''
            $printer = "{self.zebra_printer}"
            $zpl = @"
{escaped_zpl}
"@
            $zpl | Out-Printer -Name $printer
            '''
            
            result = subprocess.run([
                'powershell', '-Command', ps_command
            ], capture_output=True, text=True, shell=True)
            
            if result.returncode != 0:
                raise Exception(f"PowerShell failed: {result.stderr}")
                
        except Exception as e:
            raise Exception(f"PowerShell method failed: {str(e)}")
    
    def send_via_file(self, zpl_code):
        """Send ZPL using temporary file"""
        try:
            # Create temporary file with ZPL code
            with tempfile.NamedTemporaryFile(mode='w', suffix='.zpl', delete=False) as temp_file:
                temp_file.write(zpl_code)
                temp_file_path = temp_file.name
            
            try:
                # Try using Windows copy command to printer port
                result = subprocess.run([
                    'copy', '/B', temp_file_path, f'\\\\{self.zebra_printer}'
                ], capture_output=True, text=True, shell=True)
                
                if result.returncode != 0:
                    # Try alternative method with print command
                    result = subprocess.run([
                        'print', '/D:', self.zebra_printer, temp_file_path
                    ], capture_output=True, text=True, shell=True)
                    
                    if result.returncode != 0:
                        raise Exception(f"File-based print failed: {result.stderr}")
                
            finally:
                # Clean up temporary file
                try:
                    os.unlink(temp_file_path)
                except:
                    pass
                    
        except Exception as e:
            raise Exception(f"File method failed: {str(e)}")
    
    def generate_label_zpl(self, order_number, site_name, route, crate_number, total_crates, dispatchcode):
        """Generate ZPL code for the label (6.5cm x 13.5cm)"""
        # Convert cm to dots (203 DPI = 80 dots per cm)
        width_dots = int(6.5 * 80)  # 520 dots
        height_dots = int(13.5 * 80)  # 1080 dots
        
        
        # Ensure all variables have default values if None
        order_number = order_number or "N/A"
        site_name = site_name or "N/A"
        route = route or "N/A"
        dispatchcode = dispatchcode or "N/A"
        barcode_data = order_number
        
        # Handle long site names - allow up to 3 lines to accommodate Region 3 + Region 5 text
        site_name_lines = []
        # Use longer line length to reduce number of lines while ensuring fit
        max_chars_per_line = 28  # Increased from 25 to accommodate longer combined text
        
        if len(site_name) > max_chars_per_line:
            words = site_name.split()
            current_line = ""
            for word in words:
                if len(current_line + " " + word) <= max_chars_per_line:
                    current_line += (" " + word) if current_line else word
                else:
                    if current_line:
                        site_name_lines.append(current_line)
                    current_line = word
                    # Allow up to 3 lines to fit the complete site name
                    if len(site_name_lines) >= 3:
                        # If we already have 3 lines, truncate the remaining text
                        remaining_text = " ".join([current_line] + words[words.index(word):])
                        if len(remaining_text) > 20:
                            remaining_text = remaining_text[:17] + "..."
                        site_name_lines.append(remaining_text)
                        break
            if current_line and len(site_name_lines) < 3:
                site_name_lines.append(current_line)
        else:
            site_name_lines = [site_name]
        
        # Site name font sizing - adjust based on number of lines and total length (increased sizes)
        num_lines = len(site_name_lines)
        total_length = len(site_name)
        
        if num_lines == 1:
            if total_length <= 15:
                site_font_size = 70
            elif total_length <= 20:
                site_font_size = 65
            else:
                site_font_size = 60
        elif num_lines == 2:
            if total_length <= 30:
                site_font_size = 60
            else:
                site_font_size = 55
        elif num_lines == 3:
            if total_length <= 45:
                site_font_size = 50
            else:
                site_font_size = 45
        else:  # 4+ lines (shouldn't happen with our 3-line limit)
            site_font_size = 40
        
        # Handle long route names - split into maximum 2 lines with optimized breaking
        route_lines = []
        # Use 15 character limit to wrap sooner
        max_route_chars_per_line = 15
        
        if len(route) > max_route_chars_per_line:
            words = route.split()
            current_line = ""
            for word in words:
                if len(current_line + " " + word) <= max_route_chars_per_line:
                    current_line += (" " + word) if current_line else word
                else:
                    if current_line:
                        route_lines.append(current_line)
                    current_line = word
                    # Limit to maximum 2 lines
                    if len(route_lines) >= 2:
                        # If we already have 2 lines, truncate the remaining text
                        remaining_text = " ".join([current_line] + words[words.index(word):])
                        if len(remaining_text) > 15:
                            remaining_text = remaining_text[:12] + "..."
                        route_lines.append(remaining_text)
                        break
            if current_line and len(route_lines) < 2:
                route_lines.append(current_line)
        else:
            route_lines = [route]
        
        # Dynamic barcode sizing based on order number length - bigger for shorter numbers
        barcode_height = 60 if len(order_number) <= 10 else 50 if len(order_number) <= 15 else 40
        barcode_width = 3 if len(order_number) <= 10 else 2 if len(order_number) <= 15 else 1
        
        # Dynamic text sizing based on content length
        crate_text = f"{crate_number} of {total_crates}"
        
        # Crate text sizing - adjust based on length (shorter text now)
        if len(crate_text) <= 8:
            crate_font_size = 65
        elif len(crate_text) <= 12:
            crate_font_size = 55
        else:
            crate_font_size = 45
            
        # Route text sizing - adjust based on number of lines and total length
        num_route_lines = len(route_lines)
        total_route_length = len(route)
        
        if num_route_lines == 1:
            if total_route_length <= 6:
                route_font_size = 80
            elif total_route_length <= 10:
                route_font_size = 70
            elif total_route_length <= 15:
                route_font_size = 60
            elif total_route_length <= 20:
                route_font_size = 50
            else:
                route_font_size = 40
        else:  # 2 lines
            if total_route_length <= 20:
                route_font_size = 50
            elif total_route_length <= 30:
                route_font_size = 45
            else:
                route_font_size = 40
            
        
        # Start position with more empty space at top to move content down
        start_y = 350  # Fine-tuned positioning - moved up slightly from 400
        
        # Build ZPL code with proper positioning and center alignment
        zpl = f"""^XA
^PW{width_dots}
^LL{height_dots}
^FO30,{start_y + 80}^GB480,2,2^FS"""
        
        # Add site name lines (left aligned) with increased spacing and bold font
        # Note: site_name now contains combined Region 3 + Region 5 text
        y_pos = start_y + 100
        for line in site_name_lines:
            zpl += f"\n^FO30,{y_pos}^A0N,{site_font_size},{site_font_size}^FD{line}^FS"
            y_pos += site_font_size + 25  # Increased from 15 to 25
        
        # Add crate info (center aligned) with increased spacing
        y_pos += 30  # Increased from 20 to 30
        zpl += f"""
^FO30,{y_pos}^GB480,2,2^FS
^FO{width_dots//2 - len(crate_text)*12},{y_pos + 25}^A0N,{crate_font_size},{crate_font_size}^FD^CI28^FD{crate_text}^FS
^FO30,{y_pos + 90}^GB480,2,2^FS
^FO{max(20, width_dots//2 - (len(barcode_data) * barcode_width * 7))},{y_pos + 110}^BY{barcode_width}
^BCN,{barcode_height},Y,N,N
^FD{barcode_data}^FS
^FO30,{y_pos + 110 + barcode_height + 50}^GB480,2,2^FS"""
        
        # Add route lines (center aligned) with increased spacing
        route_y_pos = y_pos + 110 + barcode_height + 70
        for line in route_lines:
            zpl += f"\n^FO{max(30, width_dots//2 - len(line)*18)},{route_y_pos}^A0N,{route_font_size},{route_font_size}^FD{line}^FS"
            route_y_pos += route_font_size + 15  # Space between lines
        
        zpl += "\n^XZ"""
        
        return zpl
    
    def preview_label(self):
        """Show a preview of the label"""
        if not self.current_order_data:
            QMessageBox.warning(self, "Preview Error", "Please scan an order number first.")
            return
        
        try:
            crate_count = 1  # Default to 1 crate since UI element was removed
            order_number = self.current_order_data.get('ordernumber', 'N/A')
            site_name = self.current_order_data.get('sitename', 'N/A')
            route = self.current_order_data.get('route', 'N/A')
            dispatchcode = self.current_order_data.get('dispatchcode', 'N/A')
            
            # Validate that we have the essential data
            if not order_number or order_number == 'N/A':
                raise Exception("Order number is missing or invalid")
            
            # Generate ZPL code for preview
            zpl_code = self.generate_label_zpl(order_number, site_name, route, 1, crate_count, dispatchcode)
            
            # Show preview dialog
            preview_dialog = LabelPreviewDialog(zpl_code, order_number, dispatchcode, site_name, route, 1, crate_count, self)
            preview_dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(self, "Preview Error", f"Failed to generate preview: {str(e)}")
    
    def log_print_message(self, message):
        """Add message to print log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        self.print_log.append(log_message)
        # Auto-scroll to bottom
        self.print_log.verticalScrollBar().setValue(self.print_log.verticalScrollBar().maximum())
    
    def show_logs_dialog(self):
        """Show logs in a separate dialog window"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Print Logs")
        dialog.setModal(True)
        dialog.resize(800, 600)
        
        layout = QVBoxLayout(dialog)
        
        # Log text area
        log_text = QTextEdit()
        log_text.setReadOnly(True)
        log_text.setPlainText(self.print_log.toPlainText())
        log_text.setStyleSheet("""
            QTextEdit {
                background-color: #f8f9fa;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                font-family: 'Consolas', monospace;
                font-size: 11px;
            }
        """)
        layout.addWidget(log_text)
        
        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        clear_button = QPushButton("Clear Logs")
        clear_button.clicked.connect(lambda: self.clear_logs(dialog))
        button_layout.addWidget(clear_button)
        
        close_button = QPushButton("Close")
        close_button.clicked.connect(dialog.accept)
        button_layout.addWidget(close_button)
        
        layout.addLayout(button_layout)
        dialog.exec()
    
    def clear_logs(self, dialog):
        """Clear all logs"""
        self.print_log.clear()
    
    def clear_print_log(self):
        """Clear the main print log widget"""
        self.print_log.clear()
        self.log_print_message("Log cleared by user")
    
    
    def create_unified_file_selection_column(self):
        """Create unified left column with file selection controls"""
        column = QFrame()
        column.setObjectName("columnFrame")
        column.setFixedWidth(480)
        
        layout = QVBoxLayout(column)
        layout.setSpacing(8)
        
        # Output folder section
        output_section = self.create_output_folder_section()
        layout.addWidget(output_section)
        
        # Picking sheet PDF files section
        picking_section = self.create_picking_sheet_section()
        layout.addWidget(picking_section)
        
        # Date picker section
        date_section = self.create_date_section()
        layout.addWidget(date_section)
        
        # Process button section
        process_section = self.create_process_button_section()
        layout.addWidget(process_section)
        
        layout.addStretch()
        return column
    
    def create_process_button_section(self):
        """Create process button section"""
        section = QFrame()
        section.setObjectName("section")
        
        layout = QVBoxLayout(section)
        layout.setSpacing(8)
        
        # Title with icon
        title_container = QHBoxLayout()
        title_container.setSpacing(8)
        
        icon_label = QLabel("4")
        icon_label.setObjectName("sectionIcon")
        icon_label.setAlignment(Qt.AlignCenter)
        icon_label.setStyleSheet("""
            QLabel#sectionIcon {
                font-size: 12px;
                background-color: #3498db;
                color: white;
                border-radius: 3px;
                padding: 4px;
                min-width: 20px;
                max-width: 20px;
                min-height: 20px;
                max-height: 20px;
                text-align: center;
                qproperty-alignment: AlignCenter;
            }
        """)
        title_container.addWidget(icon_label)
        
        title_label = QLabel("Process Picking Sheets")
        title_label.setObjectName("sectionTitle")
        title_container.addWidget(title_label)
        
        title_container.addStretch()
        layout.addLayout(title_container)
        
        # Process button
        self.unified_process_btn = QPushButton("Process Picking Sheets")
        self.unified_process_btn.setObjectName("primaryButton")
        self.unified_process_btn.clicked.connect(self.process_unified_flow)
        self.unified_process_btn.setEnabled(False)
        layout.addWidget(self.unified_process_btn)
        
        # Progress bar (initially hidden)
        self.unified_progress_bar = QProgressBar()
        self.unified_progress_bar.setVisible(False)
        layout.addWidget(self.unified_progress_bar)
        
        return section
    
    def create_unified_processing_column(self):
        """Create unified right column with processing section"""
        column = QFrame()
        column.setObjectName("columnFrame")
        
        layout = QVBoxLayout(column)
        layout.setSpacing(8)
        
        # Workflow information section
        workflow_section = self.create_workflow_info_section()
        layout.addWidget(workflow_section)
        
        # Excel column requirements section
        requirements_section = self.create_requirements_section()
        layout.addWidget(requirements_section)
        
        layout.addStretch()
        return column
    
    def create_picking_sheet_section(self):
        """Create picking sheet PDF files selection section"""
        section = QFrame()
        section.setObjectName("section")
        
        layout = QVBoxLayout(section)
        layout.setSpacing(12)
        
        # Title with icon
        title_container = QHBoxLayout()
        title_container.setSpacing(8)
        
        icon_label = QLabel("2")
        icon_label.setObjectName("sectionIcon")
        icon_label.setAlignment(Qt.AlignCenter)
        icon_label.setStyleSheet("""
            QLabel#sectionIcon {
                font-size: 12px;
                background-color: #3498db;
                color: white;
                border-radius: 3px;
                padding: 4px;
                min-width: 20px;
                max-width: 20px;
                min-height: 20px;
                max-height: 20px;
                text-align: center;
                qproperty-alignment: AlignCenter;
            }
        """)
        title_container.addWidget(icon_label)
        
        title_label = QLabel("Picking Sheet PDF Files")
        title_label.setObjectName("sectionTitle")
        title_container.addWidget(title_label)
        
        title_container.addStretch()
        layout.addLayout(title_container)
        
      


        
        # File selection controls
        file_controls_layout = QHBoxLayout()
        
        self.picking_sheet_btn = QPushButton("Select Picking Sheets")
        self.picking_sheet_btn.setObjectName("fileButton")
        self.picking_sheet_btn.clicked.connect(self.browse_picking_sheet_files)
        file_controls_layout.addWidget(self.picking_sheet_btn)
        
        self.clear_picking_sheet_btn = QPushButton("Clear")
        self.clear_picking_sheet_btn.setObjectName("clearButton")
        self.clear_picking_sheet_btn.clicked.connect(self.clear_picking_sheet_files)
        file_controls_layout.addWidget(self.clear_picking_sheet_btn)
        
        layout.addLayout(file_controls_layout)
        
        # File status label
        self.picking_sheet_label = QLabel("No picking sheet files selected")
        self.picking_sheet_label.setObjectName("fileStatusLabel")
        layout.addWidget(self.picking_sheet_label)
        
        return section
    
    
    
    




    
    
    def create_file_selection_column(self):
        """Create left column with file selection controls"""
        column = QFrame()
        column.setObjectName("columnFrame")
        column.setFixedWidth(550)
        
        layout = QVBoxLayout(column)
        layout.setSpacing(8)
        
        # Output folder section
        output_section = self.create_output_folder_section()
        layout.addWidget(output_section)
        
        # Excel file section
        excel_section = self.create_excel_file_section()
        layout.addWidget(excel_section)
        
        # Date picker section
        date_section = self.create_date_section()
        layout.addWidget(date_section)
        
        # PDF files section
        pdf_section = self.create_pdf_files_section()
        layout.addWidget(pdf_section)
        
        layout.addStretch()
        
        return column
    
    def create_output_folder_section(self):
        """Create output folder selection section"""
        section = QFrame()
        section.setObjectName("section")
        
        layout = QVBoxLayout(section)
        layout.setSpacing(12)
        
        # Title with icon
        title_container = QHBoxLayout()
        title_container.setSpacing(8)
        
        icon_label = QLabel("1")
        icon_label.setObjectName("sectionIcon")
        icon_label.setAlignment(Qt.AlignCenter)
        icon_label.setStyleSheet("""
            QLabel#sectionIcon {
                font-size: 12px;
                background-color: #3498db;
                color: white;
                border-radius: 3px;
                padding: 4px;
                min-width: 20px;
                max-width: 20px;
                min-height: 20px;
                max-height: 20px;
                text-align: center;
                qproperty-alignment: AlignCenter;
            }
        """)
        title_container.addWidget(icon_label)
        
        title = QLabel("Output Folder")
        title.setObjectName("sectionTitle")
        title_container.addWidget(title)
        
        title_container.addStretch()
        layout.addLayout(title_container)
        
        # Button row
        btn_layout = QHBoxLayout()
        self.browse_output_btn = QPushButton("Select Output Folder")
        self.browse_output_btn.clicked.connect(self.browse_output_folder)
        
        self.clear_output_btn = QPushButton("Clear")
        self.clear_output_btn.setObjectName("secondaryButton")
        self.clear_output_btn.clicked.connect(self.clear_output_folder)
        
        btn_layout.addWidget(self.browse_output_btn)
        btn_layout.addWidget(self.clear_output_btn)
        layout.addLayout(btn_layout)
        
        # Status display
        self.output_folder_label = QLabel("Files will be saved in a date-based subfolder (YYYY-MM-DD)")
        self.output_folder_label.setObjectName("infoText")
        self.output_folder_label.setWordWrap(True)
        layout.addWidget(self.output_folder_label)
        
        return section
    
    def create_excel_file_section(self):
        """Create Excel file selection section"""
        section = QFrame()
        section.setObjectName("section")
        
        layout = QVBoxLayout(section)
        layout.setSpacing(0)
        
        # Section title
        title = QLabel("📊 Store Order File (Excel/CSV)")
        title.setObjectName("sectionTitle")
        layout.addWidget(title)
        
        # Subtitle
        
        
        
        
        # Button row
        btn_layout = QHBoxLayout()
        self.browse_excel_btn = QPushButton("Select File (Excel/CSV)")
        self.browse_excel_btn.clicked.connect(self.browse_excel_file)
        
        self.clear_excel_btn = QPushButton("Clear")
        self.clear_excel_btn.setObjectName("secondaryButton")
        self.clear_excel_btn.clicked.connect(self.clear_excel_file)
        
        btn_layout.addWidget(self.browse_excel_btn)
        btn_layout.addWidget(self.clear_excel_btn)
        layout.addLayout(btn_layout)
        
        # Status display
        self.excel_file_label = QLabel("No Excel file selected")
        self.excel_file_label.setObjectName("infoText")
        self.excel_file_label.setWordWrap(True)
        layout.addWidget(self.excel_file_label)
        
        return section
    

    
    def create_date_section(self):
        """Create date picker section"""
        section = QFrame()
        section.setObjectName("section")
        
        layout = QVBoxLayout(section)
        layout.setSpacing(12)
        
        # Title with icon
        title_container = QHBoxLayout()
        title_container.setSpacing(8)
        
        icon_label = QLabel("3")
        icon_label.setObjectName("sectionIcon")
        icon_label.setAlignment(Qt.AlignCenter)
        icon_label.setStyleSheet("""
            QLabel#sectionIcon {
                font-size: 12px;
                background-color: #3498db;
                color: white;
                border-radius: 3px;
                padding: 4px;
                min-width: 20px;
                max-width: 20px;
                min-height: 20px;
                max-height: 20px;
                text-align: center;
                qproperty-alignment: AlignCenter;
            }
        """)
        title_container.addWidget(icon_label)
        
        title = QLabel("Order Picking Date")
        title.setObjectName("sectionTitle")
        title_container.addWidget(title)
        
        title_container.addStretch()
        layout.addLayout(title_container)
        
        # Date picker
        self.delivery_date_edit = QDateEdit()
        self.delivery_date_edit.setCalendarPopup(True)
        self.delivery_date_edit.setDate(QDate.currentDate())
        self.delivery_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.delivery_date_edit.setFixedHeight(35)
        layout.addWidget(self.delivery_date_edit)
        
        return section
    
    def create_pdf_files_section(self):
        """Create PDF files selection section"""
        section = QFrame()
        section.setObjectName("section")
        
        layout = QVBoxLayout(section)
        layout.setSpacing(0)
        
        # Section title
        title = QLabel("📄 PDF Files to Process")
        title.setObjectName("sectionTitle")
        layout.addWidget(title)
        
        # Button row
        btn_layout = QHBoxLayout()
        add_picking_pdf_btn = QPushButton("Add PDF Files")
        add_picking_pdf_btn.clicked.connect(self.browse_picking_pdf_files)
        
        clear_picking_pdf_btn = QPushButton("Clear")
        clear_picking_pdf_btn.setObjectName("secondaryButton")
        clear_picking_pdf_btn.clicked.connect(self.clear_picking_pdf_files)
        
        btn_layout.addWidget(add_picking_pdf_btn)
        btn_layout.addWidget(clear_picking_pdf_btn)
        layout.addLayout(btn_layout)
        
        # PDF list
        self.picking_pdf_list = QListWidget()
        self.picking_pdf_list.setMaximumHeight(120)
        layout.addWidget(self.picking_pdf_list)
        
        return section
    
    def create_processing_column(self):
        """Create right column with processing section"""
        column = QFrame()
        column.setObjectName("columnFrame")
        
        layout = QVBoxLayout(column)
        layout.setSpacing(8)
        
        # Main processing section
        processing_section = self.create_main_processing_section()
        layout.addWidget(processing_section)
        
        # Workflow information section
        workflow_section = self.create_workflow_info_section()
        layout.addWidget(workflow_section)
        
        # Excel column requirements section
        requirements_section = self.create_requirements_section()
        layout.addWidget(requirements_section)
        
        layout.addStretch()
        
        return column
    
    def create_main_processing_section(self):
        """Create main processing section with action button"""
        section = QFrame()
        section.setObjectName("section")
        
        layout = QVBoxLayout(section)
        layout.setSpacing(0)
        
        # Section title
        title = QLabel("🚀 Process PDFs & Add Barcodes")
        title.setObjectName("sectionTitle")
        layout.addWidget(title)
        
        # Process button
        self.process_picking_btn = QPushButton("Process PDFs & Add Barcodes")
        self.process_picking_btn.setObjectName("primaryButton")
        self.process_picking_btn.clicked.connect(self.process_picking_dockets)
        self.process_picking_btn.setFixedHeight(45)
        layout.addWidget(self.process_picking_btn)
        

        
        return section
    
    def create_workflow_info_section(self):
        """Create workflow information section"""
        section = QFrame()
        section.setObjectName("section")
        
        layout = QVBoxLayout(section)
        layout.setSpacing(12)
        
        # Title
        title = QLabel("Step by Step Process")
        title.setObjectName("sectionTitle")
        layout.addWidget(title)
        
        # Workflow steps
        workflow_text = QLabel("""
        <b>Follow these steps</b>
        <br>1. Select the output folder where proccessed picking sheets will be saved
        <br>2. Select the picking sheets from customer service
        <br>3. Select the order picking date
        <br>4. Click the 'Proccess Picking Sheets' button to start the process
        <br>5. Check the output folder for the processed picking sheets
      
       
        """)
        workflow_text.setObjectName("workflowText")
        workflow_text.setWordWrap(True)
        layout.addWidget(workflow_text)
        
        # Debug OCR Regions button
        debug_btn = QPushButton("🔍 Debug OCR Regions")
        debug_btn.setObjectName("secondaryButton")
        debug_btn.clicked.connect(self.show_ocr_debug_dialog)
        debug_btn.setToolTip("View and test OCR region configurations")
        layout.addWidget(debug_btn)
        
        return section
    
    def show_ocr_debug_dialog(self):
        """Show the OCR debug dialog"""
        dialog = OCRDebugDialog(self.ocr_regions, self)
        dialog.exec()
    
    def create_requirements_section(self):
        """Create Excel column requirements section"""
        section = QFrame()
        section.setObjectName("section")
        
        layout = QVBoxLayout(section)
        layout.setSpacing(12)
        
        # Title
        title = QLabel("Expected Excel Columns")
        title.setObjectName("sectionTitle")
        layout.addWidget(title)
        
        # Requirements text
        requirements_text = QLabel("""
        <b>Required columns in your Excel file:</b>
        <br>• Column A: Order Number (→ ordernumber)
        <br>• Column B: Item Code (→ itemcode)
        <br>• Column C: Product Description (→ product_description)
        <br>• Column D: Barcode (→ barcode)
        <br>• Column E: Customer Type (→ customer_type)
        <br>• Column F: Quantity (→ quantity)
        <br>• Column G: Site Name (→ sitename)
        <br>• Column H: Account Code (→ accountcode)
        <br>• Column I: Dispatch Code (→ dispatchcode)
        <br>• Column J: Route (→ route)
                                   
        """)
        requirements_text.setObjectName("requirementsText")
        requirements_text.setWordWrap(True)
        layout.addWidget(requirements_text)
        
        return section
    

    
    # File handling methods
    def browse_picking_pdf_files(self):
        """Browse for picking PDF files to process"""
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Picking Docket PDF Files",
            str(Path.home()),
            "PDF files (*.pdf);;All files (*.*)"
        )
        if file_paths:
            for file_path in file_paths:
                if file_path not in self.selected_picking_pdf_files:
                    self.selected_picking_pdf_files.append(file_path)
                    self.picking_pdf_list.addItem(Path(file_path).name)
    
    def clear_picking_pdf_files(self):
        """Clear selected picking PDF files"""
        self.selected_picking_pdf_files.clear()
        self.picking_pdf_list.clear()
        
        # Reset processing state since picking PDFs are cleared
        self.picking_dockets_processed = False
    
    # NEW: Excel file handling methods
    def browse_excel_file(self):
        """Browse for Excel or CSV file containing store orders for database upload and barcode generation"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Store Order File (Excel or CSV for database upload & barcodes)",
            str(Path.home()),
            "Excel files (*.xlsx *.xls);;CSV files (*.csv);;All files (*.*)"
        )
        if file_path:
            self.selected_excel_file = file_path
            try:
                # Determine file type and load accordingly
                file_extension = Path(file_path).suffix.lower()
                print(f"📋 Loading file: {file_path} (type: {file_extension})")
                
                if file_extension in ['.csv']:
                    # Load CSV file
                    df = pd.read_csv(file_path)
                    file_type = "CSV"
                elif file_extension in ['.xlsx', '.xls']:
                    # Load Excel file
                    df = pd.read_excel(file_path)
                    file_type = "Excel"
                else:
                    raise ValueError(f"Unsupported file type: {file_extension}")
                
                if df.empty or len(df.columns) == 0:
                    raise ValueError(f"{file_type} file is empty or has no columns")
                
                print(f"📋 {file_type} file loaded successfully:")
                print(f"   - Total rows: {len(df)}")
                print(f"   - Total columns: {len(df.columns)}")
                print(f"   - Column names: {list(df.columns)}")
                
                # Show first few rows for debugging
                print(f"📋 First 3 rows of data:")
                for i, row in df.head(3).iterrows():
                    print(f"   Row {i+1}: {dict(row)}")
                
                # Get column A values (first column)
                column_a_values = df.iloc[:, 0].dropna().astype(str).tolist()
                self.excel_order_numbers = [str(val).strip() for val in column_a_values if str(val).strip()]
                
                print(f"📋 Extracted {len(self.excel_order_numbers)} order numbers from Column A")
                if self.excel_order_numbers:
                    print(f"   First 5 order numbers: {self.excel_order_numbers[:5]}")
                
                filename = Path(file_path).name
                self.excel_file_label.setText(f"Selected: {filename} ({len(self.excel_order_numbers)} order numbers)")
                self.excel_file_label.setObjectName("successText")
                self.update_status(f"Loaded {len(self.excel_order_numbers)} order numbers from {file_type} file")
                
                # Store the full DataFrame for later use in upload
                self.excel_dataframe = df
                
            except Exception as e:
                print(f"❌ Error reading file: {str(e)}")
                print(f"❌ Error type: {type(e).__name__}")
                QMessageBox.critical(self, "File Error", f"Error reading file: {str(e)}")
                self.selected_excel_file = ""
                self.excel_order_numbers = []
                self.excel_dataframe = None
                self.excel_file_label.setText("Error reading file")
                self.excel_file_label.setObjectName("warningText")
    
    def clear_excel_file(self):
        """Clear selected file (Excel or CSV)"""
        self.selected_excel_file = ""
        self.excel_order_numbers = []
        self.excel_dataframe = None
        self.excel_file_label.setText("No file selected")
        self.excel_file_label.setObjectName("infoText")
        self.update_status("File cleared")

    
    # NEW: Output folder handling methods
    def browse_output_folder(self):
        """Browse for output folder"""
        folder_path = QFileDialog.getExistingDirectory(
            self,
            "Select Output Folder",
            str(Path.home())
        )
        if folder_path:
            self.selected_output_folder = folder_path
            self.output_folder_label.setText(f"Selected: {folder_path}")
            self.output_folder_label.setObjectName("successText")
            self.update_status(f"Output folder set to: {folder_path}")
            # Update unified status if the method exists
            if hasattr(self, 'update_unified_status'):
                self.update_unified_status()
    
    def clear_output_folder(self):
        """Clear selected output folder"""
        self.selected_output_folder = ""
        self.output_folder_label.setText("No output folder selected (will use default: picking_dockets_output)\nFiles will be saved in a date-based subfolder (YYYY-MM-DD)")
        self.output_folder_label.setObjectName("infoText")
        self.update_status("Output folder cleared - will use default location")
        # Update unified status if the method exists
        if hasattr(self, 'update_unified_status'):
            self.update_unified_status()

    # Excel Generation file handling methods
    

    def browse_picking_sheet_files(self):
        """Browse for picking sheet PDF files"""
        file_paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Select Picking Sheet PDF Files",
            "",
            "PDF Files (*.pdf);;All Files (*)"
        )
        
        if file_paths:
            self.picking_sheet_files = file_paths
            file_count = len(file_paths)
            self.picking_sheet_label.setText(f"Selected {file_count} picking sheet file(s)")
            self.picking_sheet_label.setObjectName("successText")
            self.update_unified_status()
            self.update_status(f"Selected {file_count} picking sheet files")
    
    def clear_picking_sheet_files(self):
        """Clear selected picking sheet files"""
        self.picking_sheet_files = []
        self.picking_sheet_label.setText("No picking sheet files selected")
        self.picking_sheet_label.setObjectName("infoText")
        self.update_unified_status()
        self.update_status("Picking sheet files cleared")
    
    def update_unified_status(self):
        """Update the unified processing status and enable/disable process button"""
        has_output_folder = hasattr(self, 'selected_output_folder') and self.selected_output_folder
        has_picking_sheets = hasattr(self, 'picking_sheet_files') and self.picking_sheet_files
        
        if has_output_folder and has_picking_sheets:
            self.unified_process_btn.setEnabled(True)
        else:
            self.unified_process_btn.setEnabled(False)
    
    def process_unified_flow(self):
        """Process the unified flow: extract data from picking sheets, create internal Excel data, and process with barcodes"""
        if not hasattr(self, 'selected_output_folder') or not self.selected_output_folder:
            QMessageBox.warning(self, "Missing Output Folder", "Please select an output folder first.")
            return
        
        if not hasattr(self, 'picking_sheet_files') or not self.picking_sheet_files:
            QMessageBox.warning(self, "Missing Picking Sheets", "Please select picking sheet PDF files to process first.")
            return
        
        try:
            # Check if OCR setup is completed
            if not self.check_ocr_setup():
                self.unified_process_btn.setEnabled(True)
                return
            
            # Update status and show progress bar
            self.unified_process_btn.setEnabled(False)
            self.unified_progress_bar.setVisible(True)
            self.unified_progress_bar.setValue(0)
            self.update_status("Starting unified processing...")
            
            # Step 1: Extract data from picking sheets (same as Excel generation)
            debug_results = []
            configured_regions = [region for region in self.ocr_regions.values() if region['coordinates']]
            
            # Debug: Print OCR region information
            print("="*80)
            print("🔍 OCR REGIONS DEBUG INFORMATION")
            print("="*80)
            print(f"Total OCR regions configured: {len(configured_regions)}")
            for region_id, region_data in self.ocr_regions.items():
                if region_data.get('coordinates'):
                    print(f"  ✅ {region_data['name']} ({region_data['color']}): {region_data['coordinates']}")
                    # Special debug for Region 5
                    if region_data['name'] == 'Region 5':
                        print(f"      🟣 REGION 5 COORDINATES: {region_data['coordinates']}")
                else:
                    print(f"  ❌ {region_data['name']} ({region_data['color']}): Not configured")
            print("="*80)
            
            # Calculate total work
            total_pdfs = len(self.picking_sheet_files)
            total_work = 0
            pdf_page_counts = []
            
            for pdf_path in self.picking_sheet_files:
                try:
                    pdf_document = fitz.open(pdf_path)
                    page_count = len(pdf_document)
                    pdf_page_counts.append(page_count)
                    total_work += page_count * len(configured_regions)
                    pdf_document.close()
                except Exception as e:
                    pdf_page_counts.append(0)
                    print(f"Error counting pages in {pdf_path}: {e}")
            
            current_work = 0
            
            for pdf_index, pdf_path in enumerate(self.picking_sheet_files):
                self.update_status(f"Processing: {Path(pdf_path).name}")
                
                try:
                    pdf_document = fitz.open(pdf_path)
                    
                    for page_num in range(len(pdf_document)):
                        page = pdf_document[page_num]
                        
                        # Debug: Print page processing info
                        print(f"  📄 Processing page {page_num + 1} of {Path(pdf_path).name}")
                        
                        for region in configured_regions:
                            coordinates = region['coordinates']
                            rect = fitz.Rect(coordinates[0], coordinates[1], coordinates[2], coordinates[3])
                            
                            # Debug: Print region processing info
                            print(f"    🔍 Extracting from {region['name']} ({region['color']}) at {coordinates}")
                            
                            extracted_text = self.extract_text_from_exact_coordinates(page, rect)
                            
                            # Debug: Print extraction results
                            if extracted_text.strip():
                                print(f"      ✅ Extracted: '{extracted_text.strip()}'")
                                # Special debug for Region 5
                                if region['name'] == 'Region 5':
                                    print(f"      🟣 REGION 5 DEBUG: Raw text = '{extracted_text}'")
                                    print(f"      🟣 REGION 5 DEBUG: Cleaned text = '{self.clean_extracted_text(extracted_text)}'")
                            else:
                                print(f"      ❌ No text extracted from {region['name']}")
                            
                            if not extracted_text.strip():
                                print(f"      🔄 Trying OCR fallback for {region['name']}...")
                                try:
                                    mat = fitz.Matrix(3.0, 3.0)
                                    pix = page.get_pixmap(matrix=mat, clip=rect)
                                    img_data = pix.tobytes("png")
                                    image = Image.open(io.BytesIO(img_data))
                                    
                                    psm_modes = [6, 3, 7, 8, 13]
                                    for psm_mode in psm_modes:
                                        try:
                                            ocr_text = pytesseract.image_to_string(image, config=f'--psm {psm_mode}')
                                            if ocr_text.strip():
                                                extracted_text = ocr_text
                                                print(f"      ✅ OCR extracted: '{ocr_text.strip()}' (PSM {psm_mode})")
                                                break
                                        except Exception:
                                            continue
                                    else:
                                        print(f"      ❌ OCR failed for {region['name']} - no text detected")
                                except Exception as ocr_error:
                                    print(f"      ❌ OCR error for {region['name']}: {str(ocr_error)}")
                            
                            cleaned_text = self.clean_extracted_text(extracted_text)
                            
                            result = {
                                'file': Path(pdf_path).name,
                                'page': page_num + 1,
                                'region': region['name'],
                                'color': region['color'],
                                'coordinates': coordinates,
                                'extracted_text': cleaned_text,
                                'raw_text': extracted_text
                            }
                            debug_results.append(result)
                            
                            current_work += 1
                            if total_work > 0:
                                progress = int((current_work / total_work) * 50)  # First half for data extraction
                                self.unified_progress_bar.setValue(progress)
                            
                            self.update_status(f"Page {page_num + 1}, {region['name']}: '{cleaned_text}'")
                    
                    pdf_document.close()
                        
                except Exception as e:
                    error_result = {
                        'file': Path(pdf_path).name,
                        'error': str(e),
                        'coordinates': coordinates
                    }
                    debug_results.append(error_result)
                    self.update_status(f"Error processing {Path(pdf_path).name}: {str(e)}")
            
            # Step 2: Create internal Excel data structure (instead of generating Excel file)
            self.unified_progress_bar.setValue(60)
            
            self.internal_excel_data = self.create_internal_excel_data(debug_results)
            
            if not self.internal_excel_data:
                QMessageBox.warning(self, "No Data", "No valid data found in the picking sheets. Please check the PDF files and try again.")
                self.unified_process_btn.setEnabled(True)
                self.unified_progress_bar.setVisible(False)
                return
            
            # Step 2.5: Generate Excel backup file
            self.unified_progress_bar.setValue(70)
            
            try:
                self.generate_excel_backup_file(self.internal_excel_data)
                self.update_status("✅ Excel backup file generated successfully")
            except Exception as e:
                self.update_status(f"⚠️ Warning: Could not generate Excel backup: {str(e)}")
            
            # DEBUG: Show the generated table data
            self.show_debug_table(self.internal_excel_data)
            
            # Also print to console for easy debugging
            self.print_debug_data(self.internal_excel_data)
            
            # Step 3: Set up for barcode generation and database upload
            self.unified_progress_bar.setValue(70)
            
            # Set the internal data as the Excel data for the existing processing flow
            self.excel_order_numbers = [row.get('ordernumber', '') for row in self.internal_excel_data if row.get('ordernumber')]
            
            # Set the picking sheet files as the PDF files for processing
            self.selected_picking_pdf_files = self.picking_sheet_files
            
            # Step 4: Continue with the existing barcode generation and database upload process
            self.unified_progress_bar.setValue(80)
            
            # Start the existing processing flow
            self.show_progress(True)
            self.update_status("Starting barcode generation and database upload...")
            
            # Start background processing
            self.processing_thread = ProcessingThread(self)
            self.processing_thread.progress_signal.connect(self.update_status)
            self.processing_thread.finished_signal.connect(self.on_unified_processing_finished)
            self.processing_thread.start()
            
        except Exception as e:
            QMessageBox.critical(self, "Unified Processing Error", f"An error occurred during unified processing:\n{str(e)}")
            self.unified_process_btn.setEnabled(True)
            self.unified_progress_bar.setVisible(False)
            self.update_status(f"Unified processing failed: {str(e)}")
    
    def create_internal_excel_data(self, debug_results):
        """Create internal Excel data structure from debug results (same logic as generate_excel_files but returns data instead of saving)"""
        if not debug_results:
            return []
        
        try:
            # Group results by file and page, but only include pages with "Total Items Delivered:" in Region 4
            results_by_file_page = {}
            pages_to_skip = set()
            
            # First pass: identify pages that should be skipped
            for result in debug_results:
                if 'error' in result:
                    continue
                    
                file_name = result.get('file', 'Unknown')
                page_num = result.get('page', 1)
                region_name = result.get('region', 'Unknown')
                extracted_text = result.get('extracted_text', '')
                
                key = (file_name, page_num)
                
                if 'Region 4' in region_name:
                    cleaned_text = extracted_text.strip()
                    if 'Total Items Delivered:' not in cleaned_text:
                        pages_to_skip.add(key)
                    else:
                        self.update_status(f"Processing page {page_num} - 'Total Items Delivered:' found in Region 4")
            
            # Second pass: collect data only from pages that should be processed
            for result in debug_results:
                if 'error' in result:
                    continue
                    
                file_name = result.get('file', 'Unknown')
                page_num = result.get('page', 1)
                region_name = result.get('region', 'Unknown')
                extracted_text = result.get('extracted_text', '')
                
                key = (file_name, page_num)
                
                if key in pages_to_skip:
                    continue
                
                if key not in results_by_file_page:
                    results_by_file_page[key] = {
                        'file': file_name,
                        'page': page_num,
                        'region_1': '',  # Column J (Route)
                        'region_2': '',  # Column A (Order Number)
                        'region_3': '',  # Column G (Site Name)
                        'region_4': '',  # For trigger text verification
                        'region_5': ''   # Additional data field
                    }
                
                if 'Region 1' in region_name:
                    results_by_file_page[key]['region_1'] = extracted_text
                elif 'Region 2' in region_name:
                    results_by_file_page[key]['region_2'] = extracted_text.upper()  # Convert to uppercase
                elif 'Region 3' in region_name:
                    results_by_file_page[key]['region_3'] = extracted_text
                elif 'Region 4' in region_name:
                    results_by_file_page[key]['region_4'] = extracted_text
                elif 'Region 5' in region_name:
                    # Append Region 5 text to Region 3 text
                    if 'region_3' in results_by_file_page[key]:
                        # If Region 3 already has text, append Region 5 with a space
                        results_by_file_page[key]['region_3'] += ' ' + extracted_text
                    else:
                        # If Region 3 doesn't have text yet, just set Region 5 text
                        results_by_file_page[key]['region_3'] = extracted_text
                    # Also store Region 5 separately for reference
                    results_by_file_page[key]['region_5'] = extracted_text
            
            # Convert to Excel data structure
            excel_data = []
            for key, data in results_by_file_page.items():
                # Map regions to Excel columns
                order_number = data['region_2'].strip()
                site_name = data['region_3'].strip()
                route = data['region_1'].strip()
                
                if order_number:  # Only add rows with order numbers
                    excel_row = {
                        'ordernumber': order_number,  # Database expects lowercase field names
                        'itemcode': 'DEFAULT',  # Default item code for picking sheets
                        'product_description': 'Picking Sheet Order',  # Default description
                        'barcode': '',  # Will be generated later
                        'customer_type': 'PICKUP',  # Default customer type
                        'quantity': 1,  # Default quantity (as integer)
                        'sitename': site_name,  # Database expects lowercase field names
                        'accountcode': 'PICKUP',  # Default account code
                        'dispatchcode': 'PICKUP',  # Default dispatch code
                        'route': route,  # Database expects lowercase field names
                        'pdf_file_name': data['file']  # Add PDF file name to each row
                    }
                    excel_data.append(excel_row)
            
            self.update_status(f"Created {len(excel_data)} rows of internal Excel data")
            return excel_data
            
        except Exception as e:
            self.update_status(f"Error creating internal Excel data: {str(e)}")
            return []
    
    def on_unified_processing_finished(self, success, result):
        """Handle unified processing completion"""
        self.show_progress(False)
        self.unified_process_btn.setEnabled(True)
        self.unified_progress_bar.setVisible(False)
        
        if success:
            self.update_status("Unified processing completed successfully")
            
            # Mark as processed
            self.picking_dockets_processed = True
            
            # Show professional results dialog
            results_dialog = ProcessingResultsDialog(result, self)
            results_dialog.setWindowTitle("Unified Processing Results")
            results_dialog.exec()
        else:
            error_msg = result.get("error", "Unknown error occurred")
            self.update_status(f"Unified processing failed: {error_msg}")
            
            # Show professional error dialog with any partial results
            if result.get('processed_files', 0) > 0:
                results_dialog = ProcessingResultsDialog(result, self)
                results_dialog.setWindowTitle("Unified Processing Completed with Issues")
                results_dialog.exec()
            else:
                QMessageBox.critical(self, "Processing Error", f"Error during unified processing: {error_msg}")

    def show_debug_table(self, data):
        """Show a debug table with the generated data"""
        try:
            from PyQt5.QtWidgets import QDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QHeaderView, QPushButton, QLabel
            from PyQt5.QtCore import Qt
            
            dialog = QDialog(self)
            dialog.setWindowTitle("DEBUG: Generated Table Data")
            dialog.setModal(True)
            dialog.resize(1200, 600)
            
            layout = QVBoxLayout(dialog)
            
            # Title
            title_label = QLabel(f"Generated Table Data ({len(data)} rows)")
            title_label.setStyleSheet("font-size: 14px; font-weight: bold; margin: 10px;")
            layout.addWidget(title_label)
            
            # Table
            table = QTableWidget()
            table.setRowCount(len(data))
            
            if data:
                # Get column names from first row
                columns = list(data[0].keys())
                table.setColumnCount(len(columns))
                table.setHorizontalHeaderLabels(columns)
                
                # Populate table
                for row_idx, row_data in enumerate(data):
                    for col_idx, column in enumerate(columns):
                        value = str(row_data.get(column, ''))
                        item = QTableWidgetItem(value)
                        table.setItem(row_idx, col_idx, item)
                
                # Resize columns to content
                table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
                table.horizontalHeader().setStretchLastSection(True)
            
            layout.addWidget(table)
            
            # Close button
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(dialog.accept)
            layout.addWidget(close_btn)
            
            # Show dialog
            dialog.exec()
            
        except ImportError:
            # Fallback: Show data in console only (no PyQt5 dependencies)
            print("\n" + "="*80)
            print("DEBUG: Generated Table Data")
            print("="*80)
            print(f"Total rows: {len(data)}")
            print()
            
            if data:
                # Show column headers
                columns = list(data[0].keys())
                print("Columns:", " | ".join(columns))
                print("-" * 80)
                
                # Show first 10 rows
                for i, row in enumerate(data[:10]):
                    row_str = " | ".join([str(row.get(col, ''))[:15] for col in columns])
                    print(f"Row {i+1:2d}: {row_str}")
                
                if len(data) > 10:
                    print(f"... and {len(data) - 10} more rows")
            
            print("="*80)
            print("✅ Debug data displayed in console above")

    def print_debug_data(self, data):
        """Print debug data to console in a formatted way"""
        print("\n" + "="*100)
        print("🔍 DEBUG: Generated Internal Excel Data")
        print("="*100)
        print(f"📊 Total rows extracted: {len(data)}")
        print()
        
        if data:
            # Show column headers
            columns = list(data[0].keys())
            print("📋 Columns found:")
            for i, col in enumerate(columns, 1):
                print(f"   {i:2d}. {col}")
            print()
            
            # Show sample data
            print("📄 Sample data (first 5 rows):")
            print("-" * 100)
            
            # Create header row
            header = " | ".join([f"{col[:12]:<12}" for col in columns])
            print(f"Row | {header}")
            print("-" * 100)
            
            # Show data rows
            for i, row in enumerate(data[:5]):
                row_data = []
                for col in columns:
                    value = str(row.get(col, ''))[:12]
                    row_data.append(f"{value:<12}")
                row_str = " | ".join(row_data)
                print(f"{i+1:3d} | {row_str}")
            
            if len(data) > 5:
                print(f"... and {len(data) - 5} more rows")
            
            print("-" * 100)
            
            # Show summary statistics
            print("📈 Summary:")
            print(f"   • Orders found: {len(set(row.get('ordernumber', '') for row in data if row.get('ordernumber')))}")
            print(f"   • Routes found: {len(set(row.get('route', '') for row in data if row.get('route')))}")
            print(f"   • Sites found: {len(set(row.get('sitename', '') for row in data if row.get('sitename')))}")
            
        else:
            print("❌ No data extracted!")
        
        print("="*100)

    def check_ocr_setup(self):
        """Check if OCR setup is completed, prompt for setup if not"""
        configured_regions = [region for region in self.ocr_regions.values() if region['coordinates']]
        if not self.ocr_setup_completed or not configured_regions:
            if not self.excel_selected_pdf_files:
                QMessageBox.warning(
                    self,
                    "No PDF Files Selected",
                    "Please select PDF files first before setting up OCR coordinates."
                )
                return False
            
            # Use the first PDF file for setup
            setup_pdf = self.excel_selected_pdf_files[0]
            
            reply = QMessageBox.question(
                self,
                "First Time Setup Required",
                "OCR coordinates need to be configured for the first time.\n\n"
                "This will open a dialog where you can visually select the region "
                "where OCR should extract text (e.g., 'Route Cork 1').\n\n"
                "Would you like to proceed with setup?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes:
                # Open coordinate selector dialog
                dialog = CoordinateSelectorDialog(setup_pdf, self)
                if dialog.exec() == QDialog.Accepted:
                    # Reload configuration
                    self.load_ocr_config()
                    return True
                else:
                    return False
            else:
                return False
        
        return True

    def quick_ocr_setup(self):
        """Quick OCR setup - select a PDF file and configure OCR region"""
        # First, let user select a PDF file
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select PDF File for OCR Configuration",
            str(Path.home()),
            "PDF files (*.pdf);;All files (*.*)"
        )
        
        if not file_path:
            return  # User cancelled
        
        # Show current coordinates if they exist
        configured_regions = [region for region in self.ocr_regions.values() if region['coordinates']]
        current_coords = ""
        if configured_regions:
            region_list = "\n".join([f"• {region['name']} ({region['color'].title()}): {region['coordinates']}" 
                                    for region in configured_regions])
            current_coords = f"\n\nCurrent regions:\n{region_list}"
        
        reply = QMessageBox.question(
            self,
            "Configure OCR Region",
            f"Selected PDF: {Path(file_path).name}\n\n"
            f"This will open a dialog where you can visually select the region "
            f"where OCR should extract text (e.g., 'Route Cork 1').\n\n"
            f"You can draw a rectangle around the text area to set the coordinates.{current_coords}\n\n"
            f"Would you like to proceed?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply == QMessageBox.Yes:
            # Open multi-region coordinate selector dialog
            dialog = MultiRegionCoordinateSelectorDialog(file_path, self)
            if dialog.exec() == QDialog.Accepted:
                # Reload configuration
                self.load_ocr_config()
                
                # Update status
                if self.ocr_coordinates:
                    self.excel_status_label.setText(f"OCR region configured: {self.ocr_coordinates}")
                    self.excel_status_label.setObjectName("successText")
                    self.update_status(f"OCR coordinates updated: {self.ocr_coordinates}")
                    
                    QMessageBox.information(
                        self,
                        "Configuration Complete",
                        f"OCR region has been configured successfully!\n\n"
                        f"Coordinates: {self.ocr_coordinates}\n\n"
                        f"You can now select your PDF files for processing and use "
                        f"'Show Current OCR Region' to verify the selection."
                    )
                else:
                    QMessageBox.warning(
                        self,
                        "Configuration Failed",
                        "OCR coordinates were not saved. Please try again."
                    )

    def configure_ocr_region(self):
        """Configure OCR region by opening PDF and allowing visual selection"""
        if not self.excel_selected_pdf_files:
            QMessageBox.warning(
                self,
                "No PDF Files Selected",
                "Please select PDF files first before configuring OCR region."
            )
            return
        
        # Use the first PDF file for configuration
        setup_pdf = self.excel_selected_pdf_files[0]
        
        # Show current coordinates if they exist
        configured_regions = [region for region in self.ocr_regions.values() if region['coordinates']]
        current_coords = ""
        if configured_regions:
            region_list = "\n".join([f"• {region['name']} ({region['color'].title()}): {region['coordinates']}" 
                                    for region in configured_regions])
            current_coords = f"\n\nCurrent regions:\n{region_list}"
        
        reply = QMessageBox.question(
            self,
            "Configure OCR Region",
            f"This will open a dialog where you can visually select the region "
            f"where OCR should extract text (e.g., 'Route Cork 1').\n\n"
            f"You can draw a rectangle around the text area to set the coordinates.{current_coords}\n\n"
            f"Would you like to proceed?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply == QMessageBox.Yes:
            # Open multi-region coordinate selector dialog
            dialog = MultiRegionCoordinateSelectorDialog(setup_pdf, self)
            if dialog.exec() == QDialog.Accepted:
                # Reload configuration
                self.load_ocr_config()
                
                # Update status
                configured_regions = [region for region in self.ocr_regions.values() if region['coordinates']]
                if configured_regions:
                    region_count = len(configured_regions)
                    self.excel_status_label.setText(f"OCR regions configured: {region_count} regions")
                    self.excel_status_label.setObjectName("successText")
                    self.update_status(f"OCR regions updated: {region_count} regions configured")
                    
                    QMessageBox.information(
                        self,
                        "Configuration Complete",
                        f"OCR regions have been configured successfully!\n\n"
                        f"Configured {region_count} regions:\n" +
                        "\n".join([f"• {region['name']} ({region['color'].title()}): {region['coordinates']}" 
                                  for region in configured_regions]) +
                        f"\n\nYou can now process your PDF files or use 'Show Current OCR Region' "
                        f"to verify the selections."
                    )
                else:
                    QMessageBox.warning(
                        self,
                        "Configuration Failed",
                        "OCR regions were not saved. Please try again."
                    )

    def show_current_ocr_region(self):
        """Show the current OCR region on a PDF"""
        configured_regions = [region for region in self.ocr_regions.values() if region['coordinates']]
        if not configured_regions:
            QMessageBox.information(
                self,
                "No OCR Regions Set",
                "No OCR regions have been configured yet.\n\n"
                "Please run the first-time setup by clicking 'Configure OCR Region' "
                "and following the setup process."
            )
            return
        
        if not self.excel_selected_pdf_files:
            QMessageBox.warning(
                self,
                "No PDF Files Selected",
                "Please select PDF files first to view the OCR region."
            )
            return
        
        try:
            # Use the first PDF file
            pdf_path = self.excel_selected_pdf_files[0]
            
            # Create a dialog to show the OCR region
            region_dialog = QDialog(self)
            region_dialog.setWindowTitle("Current OCR Region")
            region_dialog.setModal(True)
            region_dialog.resize(1200, 900)
            
            layout = QVBoxLayout(region_dialog)
            
            # Instructions
            region_count = len(configured_regions)
            instructions = QLabel(
                f"Current OCR Regions: {region_count} regions configured\n"
                "The colored rectangles show exactly where OCR will extract text from.\n"
                "Red = Region 1, Blue = Region 2, Green = Region 3, Orange = Region 4, Purple = Region 5"
            )
            instructions.setObjectName("sectionTitle")
            instructions.setWordWrap(True)
            layout.addWidget(instructions)
            
            # Graphics view for PDF display
            graphics_view = QGraphicsView()
            layout.addWidget(graphics_view)
            
            # Scene for graphics
            scene = QGraphicsScene()
            graphics_view.setScene(scene)
            
            # Load PDF and display with OCR region highlighted
            pdf_document = fitz.open(pdf_path)
            page = pdf_document[0]
            
            # Convert to image
            mat = fitz.Matrix(1.0, 1.0)  # Normal size for better overview
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")
            
            # Convert to QPixmap
            pixmap = QPixmap()
            pixmap.loadFromData(img_data)
            
            # Add to scene
            pixmap_item = QGraphicsPixmapItem(pixmap)
            scene.addItem(pixmap_item)
            
            # Add all OCR region rectangles
            scale_factor = 1.0  # Same as matrix
            for region in configured_regions:
                x1, y1, x2, y2 = region['coordinates']
                rect = QRectF(
                    x1 / scale_factor, 
                    y1 / scale_factor, 
                    (x2 - x1) / scale_factor, 
                    (y2 - y1) / scale_factor
                )
                
                # Set color based on region
                if region['color'] == 'red':
                    pen_color = QColor(255, 0, 0, 255)
                elif region['color'] == 'blue':
                    pen_color = QColor(0, 0, 255, 255)
                elif region['color'] == 'green':
                    pen_color = QColor(0, 255, 0, 255)
                elif region['color'] == 'orange':
                    pen_color = QColor(255, 165, 0, 255)  # Orange color
                elif region['color'] == 'purple':
                    pen_color = QColor(128, 0, 128, 255)  # Purple color
                else:
                    pen_color = QColor(128, 128, 128, 255)
                
                # Create rectangle for OCR region
                rect_item = scene.addRect(rect, QPen(pen_color, 3))
                rect_item.setZValue(1)  # Above the image
                
                # Add text label
                text_item = scene.addText(region['name'], QFont("Arial", 12, QFont.Bold))
                text_item.setDefaultTextColor(pen_color)
                text_item.setPos(rect.x(), rect.y() - 20)
                text_item.setZValue(2)
            
            pdf_document.close()
            
            # Buttons
            button_layout = QHBoxLayout()
            
            # Reconfigure button
            reconfigure_btn = QPushButton("Reconfigure Coordinates")
            reconfigure_btn.clicked.connect(lambda: self.reconfigure_ocr_coordinates(region_dialog))
            reconfigure_btn.setObjectName("primaryButton")
            
            # Close button
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(region_dialog.accept)
            
            button_layout.addWidget(reconfigure_btn)
            button_layout.addStretch()
            button_layout.addWidget(close_btn)
            
            layout.addLayout(button_layout)
            
            # Show dialog
            region_dialog.exec()
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error showing OCR region: {str(e)}"
            )

    def reconfigure_ocr_coordinates(self, parent_dialog):
        """Reconfigure OCR coordinates"""
        parent_dialog.accept()  # Close the current dialog
        
        if not self.excel_selected_pdf_files:
            QMessageBox.warning(
                self,
                "No PDF Files Selected",
                "Please select PDF files first before reconfiguring OCR coordinates."
            )
            return
        
        # Use the first PDF file for setup
        setup_pdf = self.excel_selected_pdf_files[0]
        
        # Open multi-region coordinate selector dialog
        dialog = MultiRegionCoordinateSelectorDialog(setup_pdf, self)
        if dialog.exec() == QDialog.Accepted:
            # Reload configuration
            self.load_ocr_config()
            QMessageBox.information(
                self,
                "Configuration Updated",
                "OCR regions have been updated successfully!"
            )

    def display_debug_results(self, debug_results):
        """Display OCR debug results in a dialog"""
        if not debug_results:
            QMessageBox.information(self, "Debug Results", "No results to display.")
            return
        
        # Create a dialog to display results
        dialog = QDialog(self)
        dialog.setWindowTitle("OCR Debug Results")
        dialog.setModal(True)
        dialog.resize(800, 600)
        
        layout = QVBoxLayout(dialog)
        
        # Title
        title = QLabel("OCR Extraction Results")
        title.setObjectName("sectionTitle")
        layout.addWidget(title)
        
        # Create text area for results
        text_area = QTextEdit()
        text_area.setReadOnly(True)
        text_area.setFont(QFont("Consolas", 10))
        
        # Format results
        result_text = "OCR Debug Results\n"
        result_text += "=" * 50 + "\n\n"
        result_text += f"Coordinates: {debug_results[0]['coordinates']}\n"
        result_text += f"Total files processed: {len(set(r.get('file', 'Unknown') for r in debug_results))}\n"
        result_text += f"Total pages processed: {len([r for r in debug_results if 'page' in r])}\n\n"
        
        for i, result in enumerate(debug_results, 1):
            result_text += f"Result {i}:\n"
            result_text += f"  File: {result.get('file', 'Unknown')}\n"
            
            if 'error' in result:
                result_text += f"  Error: {result['error']}\n"
            else:
                result_text += f"  Page: {result.get('page', 'Unknown')}\n"
                result_text += f"  Extracted Text: '{result.get('extracted_text', '')}'\n"
                result_text += f"  Raw Text: '{result.get('raw_text', '')}'\n"
            
            result_text += "\n"
        
        text_area.setPlainText(result_text)
        layout.addWidget(text_area)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        # Show dialog
        dialog.exec()

    def generate_excel_files(self, debug_results):
        """Generate Excel files with OCR results in specific columns"""
        if not debug_results:
            QMessageBox.information(self, "No Results", "No OCR results to generate Excel files from.")
            return
        
        try:
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "OCR Results"
            
            # Set up headers in the correct columns according to the required format
            # Column A: Order Number (→ ordernumber)
            # Column B: Item Code (→ itemcode)  
            # Column C: Product Description (→ product_description)
            # Column D: Barcode (→ barcode)
            # Column E: Customer Type (→ customer_type)
            # Column F: Quantity (→ quantity)
            # Column G: Site Name (→ sitename)
            # Column H: Account Code (→ accountcode)
            # Column I: Dispatch Code (→ dispatchcode)
            # Column J: Route (→ route)
            # Note: Region 5 data is combined with Region 3 in Column G
            ws.cell(row=1, column=1, value="Order Number").font = Font(bold=True)  # Column A
            ws.cell(row=1, column=2, value="Item Code").font = Font(bold=True)  # Column B
            ws.cell(row=1, column=3, value="Product Description").font = Font(bold=True)  # Column C
            ws.cell(row=1, column=4, value="Barcode").font = Font(bold=True)  # Column D
            ws.cell(row=1, column=5, value="Customer Type").font = Font(bold=True)  # Column E
            ws.cell(row=1, column=6, value="Quantity").font = Font(bold=True)  # Column F
            ws.cell(row=1, column=7, value="Site Name").font = Font(bold=True)  # Column G (Region 3 + Region 5)
            ws.cell(row=1, column=8, value="Account Code").font = Font(bold=True)  # Column H
            ws.cell(row=1, column=9, value="Dispatch Code").font = Font(bold=True)  # Column I
            ws.cell(row=1, column=10, value="Route").font = Font(bold=True)  # Column J
            
            # Style all headers
            for col in range(1, 11):  # Style all columns A through J
                cell = ws.cell(row=1, column=col)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Group results by file and page, but only include pages with "Total Items Delivered:" in Region 4
            results_by_file_page = {}
            pages_to_skip = set()
            
            # First pass: identify pages that should be skipped (no "Total Items Delivered:" in Region 4)
            for result in debug_results:
                if 'error' in result:
                    continue
                    
                file_name = result.get('file', 'Unknown')
                page_num = result.get('page', 1)
                region_name = result.get('region', 'Unknown')
                extracted_text = result.get('extracted_text', '')
                
                key = (file_name, page_num)
                
                # Check if this is Region 4 and if it contains the trigger text
                if 'Region 4' in region_name:
                    # Clean the extracted text for better matching
                    cleaned_text = extracted_text.strip()
                    if 'Total Items Delivered:' not in cleaned_text:
                        pages_to_skip.add(key)
                        self.update_status(f"Skipping page {page_num} - no 'Total Items Delivered:' found in Region 4. Found: '{cleaned_text}'")
                    else:
                        self.update_status(f"Processing page {page_num} - 'Total Items Delivered:' found in Region 4")
            
            # Second pass: collect data only from pages that should be processed
            for result in debug_results:
                if 'error' in result:
                    continue
                    
                file_name = result.get('file', 'Unknown')
                page_num = result.get('page', 1)
                region_name = result.get('region', 'Unknown')
                extracted_text = result.get('extracted_text', '')
                
                key = (file_name, page_num)
                
                # Skip this page if it doesn't have the trigger text
                if key in pages_to_skip:
                    continue
                
                if key not in results_by_file_page:
                    results_by_file_page[key] = {
                        'file': file_name,
                        'page': page_num,
                        'region_1': '',  # Column J
                        'region_2': '',  # Column A  
                        'region_3': '',  # Column G
                        'region_4': '',  # For trigger text verification
                        'region_5': ''   # Additional data field
                    }
                
                # Map regions to columns based on new requirements
                if 'Region 1' in region_name:
                    results_by_file_page[key]['region_1'] = extracted_text
                elif 'Region 2' in region_name:
                    results_by_file_page[key]['region_2'] = extracted_text.upper()  # Convert to uppercase
                elif 'Region 3' in region_name:
                    results_by_file_page[key]['region_3'] = extracted_text
                elif 'Region 4' in region_name:
                    results_by_file_page[key]['region_4'] = extracted_text
                elif 'Region 5' in region_name:
                    # Append Region 5 text to Region 3 text
                    if 'region_3' in results_by_file_page[key]:
                        # If Region 3 already has text, append Region 5 with a space
                        results_by_file_page[key]['region_3'] += ' ' + extracted_text
                    else:
                        # If Region 3 doesn't have text yet, just set Region 5 text
                        results_by_file_page[key]['region_3'] = extracted_text
                    # Also store Region 5 separately for reference
                    results_by_file_page[key]['region_5'] = extracted_text
            
            # Write data to Excel in the correct columns according to the required format
            row = 2
            for (file_name, page_num), data in results_by_file_page.items():
                # Only write rows for pages that have the trigger text and contain data from regions 1, 2, 3
                if data['region_4']:  # Ensure Region 4 had the trigger text
                    # Clean ordernumber (Region 2) - keep only letters and numbers
                    cleaned_ordernumber = re.sub(r'[^a-zA-Z0-9]', '', data['region_2'])
                    
                    # Populate all required columns according to the format
                    ws.cell(row=row, column=1, value=cleaned_ordernumber)  # Column A: Order Number (→ ordernumber)
                    ws.cell(row=row, column=2, value="")  # Column B: Item Code (→ itemcode) - empty for now
                    ws.cell(row=row, column=3, value="")  # Column C: Product Description (→ product_description) - empty for now
                    ws.cell(row=row, column=4, value="")  # Column D: Barcode (→ barcode) - empty for now
                    ws.cell(row=row, column=5, value="")  # Column E: Customer Type (→ customer_type) - empty for now
                    ws.cell(row=row, column=6, value="")  # Column F: Quantity (→ quantity) - empty for now
                    ws.cell(row=row, column=7, value=data['region_3'])  # Column G: Site Name (→ sitename) - Region 3 + Region 5 combined
                    ws.cell(row=row, column=8, value="")  # Column H: Account Code (→ accountcode) - empty for now
                    ws.cell(row=row, column=9, value="")  # Column I: Dispatch Code (→ dispatchcode) - empty for now
                    ws.cell(row=row, column=10, value=data['region_1'])  # Column J: Route (→ route) - Region 1
                    
                    row += 1
                    self.update_status(f"Added row for page {page_num}: Order Number='{cleaned_ordernumber}', Site Name='{data['region_3']}', Route='{data['region_1']}'")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generate filename with timestamp
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"OCR_Results_{timestamp}.xlsx"
            excel_path = Path(self.excel_selected_output_folder) / excel_filename
            
            # Save the workbook
            wb.save(excel_path)
            
            # Upload the generated Excel file to Supabase if available
            if SUPABASE_AVAILABLE:
                try:
                    self.processing_thread.progress_signal.emit("📤 Uploading generated Excel file to Supabase...")
                    
                    # Read the generated Excel file
                    df_generated = pd.read_excel(excel_path)
                    generated_data = df_generated.to_dict('records')
                    
                    # Upload to Supabase
                    # Extract PDF file name from the first row if available
                    pdf_file_name = None
                    if generated_data and 'pdf_file_name' in generated_data[0]:
                        pdf_file_name = generated_data[0]['pdf_file_name']
                    
                    success = upload_store_orders_from_excel(generated_data, excel_filename, pdf_file_name=pdf_file_name)
                    if success:
                        self.processing_thread.progress_signal.emit(f"✅ Successfully uploaded generated Excel file to Supabase!")
                    else:
                        self.processing_thread.progress_signal.emit(f"⚠️ Failed to upload generated Excel file to Supabase")
                except Exception as e:
                    self.processing_thread.progress_signal.emit(f"⚠️ Error uploading generated Excel file: {str(e)}")
            else:
                self.processing_thread.progress_signal.emit("⚠️ Supabase not available - Excel file saved locally only")
            
            # Show success message
            total_pages_scanned = len(set((r.get('file', ''), r.get('page', 0)) for r in debug_results if 'error' not in r))
            pages_processed = len(results_by_file_page)
            pages_skipped = total_pages_scanned - pages_processed
            
            QMessageBox.information(
                self,
                "Excel File Generated",
                f"Excel file has been generated successfully!\n\n"
                f"File: {excel_filename}\n"
                f"Location: {self.excel_selected_output_folder}\n\n"
                f"Processing Results:\n"
                f"• {total_pages_scanned} pages scanned\n"
                f"• {pages_processed} pages processed (contained 'Total Items Delivered:')\n"
                f"• {pages_skipped} pages skipped (no trigger text)\n\n"
                f"Column Mapping:\n"
                f"• Region 1 (Red) → Column H (route)\n"
                f"• Region 2 (Blue) → Column A (ordernumber)\n"
                f"• Region 3 (Green) → Column E (sitename)"
            )
            
            self.update_status(f"Excel file generated: {excel_filename}")
            
            # Final summary
            self.update_status(f"STRICT RULE APPLIED: Only pages with 'Total Items Delivered:' in Region 4 were processed")
            self.update_status(f"Final result: {pages_processed} pages processed, {pages_skipped} pages skipped")
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Excel Generation Error",
                f"Failed to generate Excel file:\n\n{str(e)}"
            )
            self.update_status(f"Excel generation failed: {str(e)}")
    
    def generate_excel_backup_file(self, excel_data):
        """Generate Excel backup file from processed table data"""
        if not excel_data:
            self.update_status("No data to generate Excel backup")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from datetime import datetime
            import os
            
            # Create a new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Processed Data Backup"
            
            # Set up headers in the correct columns
            headers = [
                "Order Number", "Item Code", "Product Description", "Barcode", 
                "Customer Type", "Quantity", "Site Name", "Account Code", 
                "Dispatch Code", "Route"
            ]
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows
            for row_idx, data in enumerate(excel_data, 2):
                ws.cell(row=row_idx, column=1, value=data.get('ordernumber', ''))
                ws.cell(row=row_idx, column=2, value=data.get('itemcode', ''))
                ws.cell(row=row_idx, column=3, value=data.get('product_description', ''))
                ws.cell(row=row_idx, column=4, value=data.get('barcode', ''))
                ws.cell(row=row_idx, column=5, value=data.get('customer_type', ''))
                ws.cell(row=row_idx, column=6, value=data.get('quantity', ''))
                ws.cell(row=row_idx, column=7, value=data.get('sitename', ''))
                ws.cell(row=row_idx, column=8, value=data.get('accountcode', ''))
                ws.cell(row=row_idx, column=9, value=data.get('dispatchcode', ''))
                ws.cell(row=row_idx, column=10, value=data.get('route', ''))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Determine output directory
            if hasattr(self, 'selected_output_folder') and self.selected_output_folder:
                base_output_dir = Path(self.selected_output_folder)
            else:
                base_output_dir = Path.cwd() / "picking_dockets_output"
            
            # Create date-based subfolder
            current_date = datetime.now().strftime("%Y-%m-%d")
            output_dir = base_output_dir / current_date
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_filename = f"Processed_Data_Backup_{timestamp}.xlsx"
            excel_path = output_dir / excel_filename
            
            # Save the workbook
            wb.save(excel_path)
            
            self.update_status(f"✅ Excel backup saved: {excel_path}")
            
        except Exception as e:
            self.update_status(f"❌ Error generating Excel backup: {str(e)}")
            raise e
    
    def add_more_orders(self):
        """Create in-app table for user to add more orders"""
        try:
            # Create and show the order entry dialog
            dialog = OrderEntryDialog(self)
            if dialog.exec() == QDialog.Accepted:
                # Get the data from the dialog
                order_data = dialog.get_order_data()
                if order_data:
                    # Convert to DataFrame format expected by the upload function
                    df = pd.DataFrame(order_data)
                    
                    # Get desktop path for easy access
                    desktop_path = str(Path.home() / "Desktop")
                    template_path = Path(desktop_path) / f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    
                    # Save the filled template
                    df.to_excel(template_path, index=False)
                    
                    # Show success message
                    QMessageBox.information(
                        self,
                        "Orders Saved Successfully",
                        f"Your orders have been saved to:\n{template_path}\n\n"
                        "You can now use 'Select Excel File' to upload this file."
                    )
                    
                    # Update status
                    self.update_status(f"Orders saved to: {template_path}")
                else:
                    QMessageBox.information(
                        self,
                        "No Orders Added",
                        "No orders were added to the table."
                    )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Error",
                f"Error creating order entry table: {str(e)}"
            )
            self.update_status("Failed to create order entry table")
    

    

    
    def open_output_directory(self, directory_path):
        """Open the output directory in file explorer"""
        try:
            import subprocess
            import platform
            
            if platform.system() == "Windows":
                subprocess.run(["explorer", directory_path], check=True)
            elif platform.system() == "Darwin":  # macOS
                subprocess.run(["open", directory_path], check=True)
            else:  # Linux
                subprocess.run(["xdg-open", directory_path], check=True)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Could not open directory: {str(e)}")
    
    # Data handling methods
    def load_existing_delivery_data(self):
        """Load existing delivery data from JSON file"""
        try:
            if os.path.exists(self.delivery_json_file):
                with open(self.delivery_json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                self.delivery_data_values = data.get("delivery_sequences", [])
                self.delivery_data_with_drivers = data.get("delivery_data_with_drivers", {})
                
                if self.delivery_data_values:
                    self.update_status(f"Loaded {len(self.delivery_data_values)} delivery sequences from file")
                else:
                    self.update_status("No delivery sequence data found - use OptimoRoute Sorter to load data first")
        except Exception as e:
            print(f"Error loading existing data: {e}")
            self.update_status("Ready - use OptimoRoute Sorter to load delivery sequence data first")
    
    def load_ocr_config(self):
        """Load OCR configuration from file, but keep hardcoded Region 5 coordinates"""
        try:
            config_path = Path("app_data") / "ocr_config.json"
            if config_path.exists():
                with open(config_path, 'r') as f:
                    config = json.load(f)
                
                if 'ocr_regions' in config:
                    # Load config but override Region 5 with hardcoded coordinates
                    self.ocr_regions = config['ocr_regions']
                    # Force Region 5 to use hardcoded coordinates
                    self.ocr_regions['region_5'] = {'coordinates': [28, 72, 328, 92], 'color': 'purple', 'name': 'Region 5'}
                    
                    self.ocr_setup_completed = config.get('setup_completed', False)
                    self.region_5_save_location = config.get('region_5_save_location', 'Column K (Region 5 Data)')
                    
                    # Print loaded configuration
                    configured_regions = [region for region in self.ocr_regions.values() if region['coordinates']]
                    print(f"OCR configuration loaded: {len(configured_regions)} regions configured")
                    for region in configured_regions:
                        print(f"  {region['name']} ({region['color']}): {region['coordinates']}")
                        if region['name'] == 'Region 5':
                            print(f"    🟣 Region 5 using HARDCODED coordinates: {region['coordinates']}")
                    
                    if hasattr(self, 'region_5_save_location'):
                        print(f"  Region 5 save location: {self.region_5_save_location}")
                    
                    return True
            else:
                print("OCR config file not found, using hardcoded configuration")
                return False
        except Exception as e:
            print(f"Error loading OCR config: {e}")
            return False
    
    # Processing methods
    def process_picking_dockets(self):
        """Process picking dockets with reversed page order and upload Excel to database"""
        # Check for store orders (either from Excel file or internal data)
        has_excel_file = hasattr(self, 'selected_excel_file') and self.selected_excel_file
        has_internal_data = hasattr(self, 'internal_excel_data') and self.internal_excel_data
        has_order_numbers = hasattr(self, 'excel_order_numbers') and self.excel_order_numbers
        
        has_store_orders = (has_excel_file or has_internal_data) and has_order_numbers
        
        if not has_store_orders:
            QMessageBox.warning(
                self, 
                "No Store Order Data", 
                "Please select a Store Order Excel File or use the unified flow:\n\n"
                "The application needs store order data to:\n"
                "• Upload data to database in exact order\n"
                "• Generate barcodes for order numbers\n"
                "• Match picking dockets to order numbers\n\n"
                "Use the unified flow to process picking sheets directly."
            )
            return
        
        if not self.selected_picking_pdf_files:
            QMessageBox.warning(self, "No Picking PDFs", "Please select picking docket PDF files to process.")
            return
        
        # Check Supabase availability for database upload
        if not SUPABASE_AVAILABLE:
            reply = QMessageBox.question(
                self, 
                "Supabase Unavailable", 
                "Supabase configuration is not available. The Excel file cannot be uploaded to the database.\n\n"
                "Do you want to continue with picking docket processing only (no database upload)?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply == QMessageBox.No:
                return
        
        self.show_progress(True)
        self.update_status("Starting processing...")
        self.process_picking_btn.setEnabled(False)
        
        # Start background processing
        self.processing_thread = ProcessingThread(self)
        self.processing_thread.progress_signal.connect(self.update_status)
        self.processing_thread.finished_signal.connect(self.on_picking_processing_finished)
        self.processing_thread.start()
    
    def on_picking_processing_finished(self, success, result):
        """Handle picking processing completion"""
        self.show_progress(False)
        self.process_picking_btn.setEnabled(True)
        
        if success:
            self.update_status("Picking dockets processing completed successfully")
            
            # Mark as processed
            self.picking_dockets_processed = True
            
            # Show professional results dialog
            results_dialog = ProcessingResultsDialog(result, self)
            results_dialog.setWindowTitle("Picking Dockets Processing Results")
            results_dialog.exec()
        else:
            error_msg = result.get("error", "Unknown error occurred")
            self.update_status(f"Picking dockets processing failed: {error_msg}")
            
            # Show professional error dialog with any partial results
            if result.get('processed_files', 0) > 0:
                # Some processing was done, show results dialog but with error status
                results_dialog = ProcessingResultsDialog(result, self)
                results_dialog.setWindowTitle("Picking Dockets Processing Completed with Issues")
                results_dialog.exec()
            else:
                QMessageBox.critical(self, "Processing Error", f"Error during picking dockets processing: {error_msg}")

    def generate_ocr_variants(self, order_id):
        """
        Generate common OCR variants of an order ID to handle character recognition errors
        """
        variants = [order_id]  # Include original
        
        # Common OCR character substitutions
        substitutions = {
            '0': ['O', 'o', 'Q', 'D'],
            'O': ['0', 'o', 'Q', 'D'],
            '1': ['l', 'I', '|', 'i'],
            'l': ['1', 'I', '|', 'i'],
            'I': ['1', 'l', '|', 'i'],
            '5': ['S', 's'],
            'S': ['5', 's'],
            '8': ['B', 'b'],
            'B': ['8', 'b'],
            '6': ['G', 'g'],
            'G': ['6', 'g'],
            '2': ['Z', 'z'],
            'Z': ['2', 'z'],
            '9': ['g', 'q'],
            'g': ['9', 'q'],
            'q': ['9', 'g'],
            'C': ['c', 'G'],
            'c': ['C', 'G'],
            'P': ['p', 'R'],
            'p': ['P', 'R'],
            'R': ['P', 'p'],
            'U': ['u', 'V'],
            'u': ['U', 'V'],
            'V': ['U', 'u'],
            'N': ['n', 'M'],
            'n': ['N', 'M'],
            'M': ['N', 'n'],
            'K': ['k', 'X'],
            'k': ['K', 'X'],
            'X': ['K', 'k'],
            'F': ['f', 'E'],
            'f': ['F', 'E'],
            'E': ['F', 'f'],
            'T': ['t', 'Y'],
            't': ['T', 'Y'],
            'Y': ['T', 't'],
            'W': ['w', 'VV'],
            'w': ['W', 'VV'],
            'H': ['h', 'A'],
            'h': ['H', 'A'],
            'A': ['H', 'h'],
            'J': ['j', 'I'],
            'j': ['J', 'I'],
            'L': ['l', 'I'],
            'D': ['d', 'O'],
            'd': ['D', 'O']
        }
        
        # Generate variants by substituting each character
        for i, char in enumerate(order_id):
            if char in substitutions:
                for sub in substitutions[char]:
                    variant = order_id[:i] + sub + order_id[i+1:]
                    if variant not in variants:
                        variants.append(variant)
        
        # Generate variants with missing characters (common OCR issue)
        for i in range(len(order_id)):
            variant = order_id[:i] + order_id[i+1:]
            if variant not in variants:
                variants.append(variant)
        
        # Generate variants with extra characters (common OCR issue)
        for i in range(len(order_id) + 1):
            for char in 'ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789':
                variant = order_id[:i] + char + order_id[i:]
                if variant not in variants:
                    variants.append(variant)
        
        return variants

    def clean_extracted_text(self, text):
        """
        Clean extracted text to improve accuracy and readability
        """
        if not text:
            return ""
        
        # Basic cleaning
        cleaned = text.strip()
        
        # Remove extra whitespace and normalize line breaks
        cleaned = ' '.join(cleaned.split())
        
        # Remove common OCR artifacts and control characters
        cleaned = cleaned.replace('\x00', '')  # Null bytes
        cleaned = cleaned.replace('\r', ' ')   # Carriage returns
        cleaned = cleaned.replace('\n', ' ')   # Line breaks
        cleaned = cleaned.replace('\t', ' ')   # Tabs
        
        # Remove multiple consecutive spaces
        import re
        cleaned = re.sub(r'\s+', ' ', cleaned)
        
        # Remove leading/trailing whitespace again
        cleaned = cleaned.strip()
        
        return cleaned

    def extract_text_from_exact_coordinates(self, page, rect):
        """
        Extract text from exact coordinates, filtering out any text outside the specified rectangle
        """
        try:
            # Get all text blocks from the page
            text_dict = page.get_text("dict")
            
            extracted_text = ""
            
            # Iterate through all text blocks
            for block in text_dict["blocks"]:
                if "lines" in block:  # Text block
                    for line in block["lines"]:
                        for span in line["spans"]:
                            # Get the bounding box of this text span
                            span_rect = fitz.Rect(span["bbox"])
                            
                            # Check if this span intersects with our target rectangle
                            if span_rect.intersects(rect):
                                # Get the intersection rectangle
                                intersection = span_rect & rect
                                
                                # Calculate how much of the span is within our target area
                                intersection_area = intersection.get_area()
                                span_area = span_rect.get_area()
                                
                                # Only include text if most of it is within our target area
                                if intersection_area / span_area > 0.5:  # At least 50% overlap
                                    text = span["text"]
                                    if text.strip():
                                        extracted_text += text + " "
            
            return extracted_text.strip()
            
        except Exception as e:
            # Fallback to simple get_textbox if the precise method fails
            try:
                return page.get_textbox(rect)
            except:
                return ""

    def process_picking_dockets_internal(self):
        """Internal method for picking dockets processing with barcode generation and Excel upload"""
        import re
        from barcode import Code128
        from barcode.writer import ImageWriter
        import tempfile
        
        try:
            # STEP 1: Upload to Supabase
            if SUPABASE_AVAILABLE:
                # Upload Store Orders (if selected or if we have internal data)
                if getattr(self, 'selected_excel_file', "") or getattr(self, 'internal_excel_data', []):
                    self.processing_thread.progress_signal.emit("📤 Uploading store order data to database...")
                    try:
                        # Use internal data if available, otherwise read from Excel file
                        if hasattr(self, 'internal_excel_data') and self.internal_excel_data:
                            self.processing_thread.progress_signal.emit(f"Using internal Excel data with {len(self.internal_excel_data)} rows...")
                            store_order_data = self.internal_excel_data
                            file_name = "Internal_Excel_Data"
                        else:
                            # Read Excel file maintaining row order
                            excel_file_path = getattr(self, 'selected_excel_file', "")
                            if excel_file_path:
                                self.processing_thread.progress_signal.emit(f"Reading {Path(excel_file_path).name} and preserving Excel row order...")
                                df = pd.read_excel(excel_file_path)
                                
                                # Convert DataFrame to list of dictionaries (preserves row order)
                                store_order_data = df.to_dict('records')
                                file_name = Path(excel_file_path).name
                            else:
                                self.processing_thread.progress_signal.emit("⚠️ No Excel file or internal data available - skipping database upload")
                                store_order_data = []
                                file_name = "None"
                        
                        self.processing_thread.progress_signal.emit(f"Uploading {len(store_order_data)} rows to dispatch_orders table in picking sequence order...")
                        
                        # Upload to Supabase using the function (order-preserving)
                        date_q = self.delivery_date_edit.date()
                        created_at_iso = f"{date_q.toString('yyyy-MM-dd')}T00:00:00+00:00"
                        
                        # Debug: Show what we're about to upload
                        self.processing_thread.progress_signal.emit(f"📋 About to upload {len(store_order_data)} rows to dispatch_orders table")
                        if store_order_data:
                            sample_row = store_order_data[0]
                            self.processing_thread.progress_signal.emit(f"📋 Sample row columns: {list(sample_row.keys())}")
                        
                        # Extract PDF file name from the first row if available
                        pdf_file_name = None
                        if store_order_data and 'pdf_file_name' in store_order_data[0]:
                            pdf_file_name = store_order_data[0]['pdf_file_name']
                        
                        success = upload_store_orders_from_excel(store_order_data, file_name, created_at_override=created_at_iso, pdf_file_name=pdf_file_name)
                        
                        if success:
                            self.processing_thread.progress_signal.emit(f"✅ Successfully uploaded {file_name} to database with Excel order preserved!")
                        else:
                            self.processing_thread.progress_signal.emit(f"⚠️ Failed to upload {file_name} to database - continuing with picking docket processing")
                    except Exception as e:
                        self.processing_thread.progress_signal.emit(f"⚠️ Error uploading Excel data to database: {str(e)} - continuing with picking docket processing")
            else:
                self.processing_thread.progress_signal.emit("⚠️ Supabase not available - skipping database upload")
            
            # STEP 2: Continue with picking docket processing
            # Determine output directory - use selected folder or default
            if self.selected_output_folder:
                base_output_dir = Path(self.selected_output_folder)
            else:
                base_output_dir = Path.cwd() / "picking_dockets_output"
            
            # Create date-based subfolder (YYYY-MM-DD format)
            current_date = datetime.now().strftime("%Y-%m-%d")
            output_dir = base_output_dir / current_date
            output_dir.mkdir(parents=True, exist_ok=True)
            
            # Dictionary to store pages grouped by order number
            order_pages = {}
            processed_files = 0
            total_pages_processed = 0
            
            # Track which order numbers were found in PDF files
            order_numbers_found_in_pdfs = set()
            
            # Dictionary to store generated barcodes for each order ID
            order_barcodes = {}
            
            # Tracking for barcode generation status
            barcode_generation_status = {}
            barcode_generation_errors = {}
            
            # Generate barcodes for Excel order numbers
            unique_order_numbers = self.excel_order_numbers.copy()
            
            # Check if we have order numbers to process
            if not unique_order_numbers:
                self.processing_thread.progress_signal.emit("❌ No order numbers found in Excel file!")
                self.processing_thread.progress_signal.emit("Please make sure your Excel file has order numbers in Column A")
                return {
                    "processed_files": 0,
                    "total_pages": 0,
                    "driver_files_created": 0,
                    "created_files": [],
                    "failed_files": [],
                    "driver_details": {},
                    "output_dir": str(output_dir),
                    "barcodes_generated": 0,
                    "error": "No order numbers found in Excel file"
                }
            
            self.processing_thread.progress_signal.emit("Starting picking dockets processing...")
            self.processing_thread.progress_signal.emit(f"Processing {len(self.selected_picking_pdf_files)} PDF files...")
            self.processing_thread.progress_signal.emit(f"Looking for {len(unique_order_numbers)} unique order numbers from Excel file...")
            
            # Debug: Show loaded Excel order numbers
            self.processing_thread.progress_signal.emit(f"Unique order numbers to find: {len(unique_order_numbers)}")
            for i, order_num in enumerate(unique_order_numbers[:5]):  # Show first 5
                self.processing_thread.progress_signal.emit(f"  {i+1}. '{order_num}'")
            if len(unique_order_numbers) > 5:
                self.processing_thread.progress_signal.emit(f"  ... and {len(unique_order_numbers) - 5} more order numbers")
            
            self.processing_thread.progress_signal.emit(f"Generating barcodes for {len(unique_order_numbers)} unique order numbers...")
            for order_id in unique_order_numbers:
                try:
                    # Validate order ID for barcode generation
                    if not order_id or not order_id.strip():
                        error_msg = "Empty or whitespace-only order number"
                        barcode_generation_errors[order_id] = error_msg
                        self.processing_thread.progress_signal.emit(f"❌ Skipped barcode generation for '{order_id}': {error_msg}")
                        continue
                    
                    # Check for invalid characters in Code128
                    invalid_chars = []
                    for char in order_id:
                        if ord(char) < 32 or ord(char) > 126:  # Code128 supports ASCII 32-126
                            invalid_chars.append(char)
                    
                    if invalid_chars:
                        error_msg = f"Invalid characters for Code128 barcode: {invalid_chars}"
                        barcode_generation_errors[order_id] = error_msg
                        self.processing_thread.progress_signal.emit(f"❌ Skipped barcode generation for '{order_id}': {error_msg}")
                        continue
                    
                    # Create barcode using Code128
                    code128 = Code128(order_id, writer=ImageWriter())
                    
                    # Generate barcode as bytes
                    barcode_buffer = io.BytesIO()
                    code128.write(barcode_buffer)
                    barcode_buffer.seek(0)
                    
                    # Store the barcode image data
                    order_barcodes[order_id] = barcode_buffer.getvalue()
                    barcode_generation_status[order_id] = "Generated"
                    
                    self.processing_thread.progress_signal.emit(f"✅ Generated barcode for Order ID: {order_id}")
                    
                except Exception as e:
                    error_msg = f"Barcode generation failed: {str(e)}"
                    barcode_generation_errors[order_id] = error_msg
                    self.processing_thread.progress_signal.emit(f"❌ Error generating barcode for '{order_id}': {error_msg}")
                    continue
            
            # Track files with and without matches
            files_with_matches = set()
            files_without_matches = set()
            
            # Process picking docket PDF files
            for pdf_file in self.selected_picking_pdf_files:
                self.processing_thread.progress_signal.emit(f"Processing picking docket: {Path(pdf_file).name}")
                
                # Track if this file has any matches
                file_has_matches = False
                
                try:
                    # Open PDF
                    pdf_document = fitz.open(pdf_file)
                    
                    # Process each page
                    for page_num in range(len(pdf_document)):
                        page = pdf_document[page_num]
                        
                        # Extract text from page
                        page_text = page.get_text()
                        
                        # Only use OCR if no text was found (much faster)
                        if not page_text.strip():
                            try:
                                # Render page as image for OCR with higher resolution
                                pix = page.get_pixmap(matrix=fitz.Matrix(3, 3))  # 3x resolution for better OCR
                                img_data = pix.tobytes("png")
                                img = Image.open(io.BytesIO(img_data))
                                
                                # Try multiple OCR configurations for better accuracy
                                ocr_text = ""
                                psm_modes = [6, 3, 7, 8, 13]  # Different page segmentation modes
                                
                                for psm_mode in psm_modes:
                                    try:
                                        ocr_result = pytesseract.image_to_string(img, config=f'--psm {psm_mode}')
                                        if ocr_result.strip():
                                            ocr_text = ocr_result
                                            break
                                    except Exception:
                                        continue
                                
                                # Use OCR text if found
                                if ocr_text.strip():
                                    page_text = ocr_text
                                    self.processing_thread.progress_signal.emit(
                                        f"Used OCR for page {page_num + 1} in {Path(pdf_file).name}"
                                    )
                                
                            except Exception as ocr_error:
                                self.processing_thread.progress_signal.emit(
                                    f"OCR failed for page {page_num + 1}: {str(ocr_error)}"
                                )
                        
                        # Search for exact order ID matches from Excel data (both files)
                        matched_order_id = None
                        
                        # Search for each order ID from Excel data directly in the PDF text
                        for excel_order_id in unique_order_numbers:
                            # Case-insensitive search for the exact order ID
                            if excel_order_id.upper() in page_text.upper():
                                matched_order_id = excel_order_id  # Use the exact case from Excel
                                self.processing_thread.progress_signal.emit(
                                    f"✅ Found exact match: '{excel_order_id}' on page {page_num + 1}"
                                )
                                break
                        
                        # If no exact match found, try word boundary search for more precision
                        if not matched_order_id:
                            for excel_order_id in unique_order_numbers:
                                # Use word boundaries to avoid partial matches
                                pattern = r'\b' + re.escape(excel_order_id) + r'\b'
                                if re.search(pattern, page_text, re.IGNORECASE):
                                    matched_order_id = excel_order_id
                                    self.processing_thread.progress_signal.emit(
                                        f"✅ Found word boundary match: '{excel_order_id}' on page {page_num + 1}"
                                    )
                                    break
                        
                        # If still no match, try fuzzy matching for OCR errors
                        if not matched_order_id:
                            for excel_order_id in unique_order_numbers:
                                # Try common OCR character substitutions
                                ocr_variants = self.generate_ocr_variants(excel_order_id)
                                for variant in ocr_variants:
                                    if variant.upper() in page_text.upper():
                                        matched_order_id = excel_order_id
                                        self.processing_thread.progress_signal.emit(
                                            f"✅ Found OCR variant match: '{excel_order_id}' (as '{variant}') on page {page_num + 1}"
                                        )
                                        break
                                if matched_order_id:
                                    break
                        
                        # Debug: Show what we found on this page
                        if matched_order_id:
                            self.processing_thread.progress_signal.emit(
                                f"Found Order ID '{matched_order_id}' on page {page_num + 1} of {Path(pdf_file).name}"
                            )
                            
                            # Track that this file has matches
                            file_has_matches = True
                            
                            # Track that this order number was found in PDFs
                            order_numbers_found_in_pdfs.add(matched_order_id)
                            
                            # Initialize order group if not exists
                            if matched_order_id not in order_pages:
                                order_pages[matched_order_id] = []
                            
                            # Store page info for this order
                            order_pages[matched_order_id].append({
                                'source_pdf_path': pdf_file,
                                'page_num': page_num,
                                'order_id': matched_order_id,
                                'source_file': pdf_file
                            })
                            
                            self.processing_thread.progress_signal.emit(
                                f"✓ Added page {page_num + 1} to order '{matched_order_id}' group"
                            )
                        else:
                            # Debug: Show extracted text for troubleshooting
                            if page_text.strip():
                                # Show first 200 characters of extracted text
                                debug_text = page_text.strip()[:200]
                                self.processing_thread.progress_signal.emit(
                                    f"Debug - Page {page_num + 1} text sample: '{debug_text}...'"
                                )
                        
                        total_pages_processed += 1
                    
                    processed_files += 1
                    pdf_document.close()
                    
                    # Track whether this file had matches
                    if file_has_matches:
                        files_with_matches.add(Path(pdf_file).name)
                        self.processing_thread.progress_signal.emit(f"✅ {Path(pdf_file).name} - Found matching order numbers")
                    else:
                        files_without_matches.add(Path(pdf_file).name)
                        self.processing_thread.progress_signal.emit(f"⚠️ {Path(pdf_file).name} - No matching order numbers found")
                    
                except Exception as e:
                    self.processing_thread.progress_signal.emit(f"Error processing {pdf_file}: {str(e)}")
                    files_without_matches.add(Path(pdf_file).name)
                    if 'pdf_document' in locals():
                        pdf_document.close()
                    continue
            
            # Summary of what was found
            total_matched_pages = sum(len(pages) for pages in order_pages.values())
            self.processing_thread.progress_signal.emit(f"📊 PDF Processing Summary:")
            self.processing_thread.progress_signal.emit(f"   - Processed {processed_files} PDF files")
            self.processing_thread.progress_signal.emit(f"   - Scanned {total_pages_processed} total pages")
            self.processing_thread.progress_signal.emit(f"   - Found {total_matched_pages} pages with matching order numbers")
            self.processing_thread.progress_signal.emit(f"   - Matched {len(order_pages)} different order numbers")
            
            # Report file matching status
            self.processing_thread.progress_signal.emit(f"📁 File Matching Status:")
            self.processing_thread.progress_signal.emit(f"   - Files with matches: {len(files_with_matches)}")
            self.processing_thread.progress_signal.emit(f"   - Files without matches: {len(files_without_matches)}")
            
            if files_without_matches:
                self.processing_thread.progress_signal.emit(f"   ⚠️ Files skipped (no matching order numbers):")
                for filename in sorted(files_without_matches):
                    self.processing_thread.progress_signal.emit(f"      - {filename}")
                self.processing_thread.progress_signal.emit(f"   💡 These files will not have barcoded PDFs created")
            
            # Comprehensive barcode and order number status reporting
            self.processing_thread.progress_signal.emit("📊 Barcode Generation and Order Number Status Report:")
            
            # Report barcode generation status
            successful_barcodes = len(order_barcodes)
            failed_barcodes = len(barcode_generation_errors)
            self.processing_thread.progress_signal.emit(f"   Barcode Generation: {successful_barcodes} successful, {failed_barcodes} failed")
            
            # Report order number matching status
            found_in_pdfs = len(order_numbers_found_in_pdfs)
            not_found_in_pdfs = len(unique_order_numbers) - found_in_pdfs
            self.processing_thread.progress_signal.emit(f"   Order Number Matching: {found_in_pdfs} found in PDFs, {not_found_in_pdfs} not found")
            
            # Report detailed failures
            if barcode_generation_errors:
                self.processing_thread.progress_signal.emit("   ❌ Barcode Generation Failures:")
                for order_id, error in barcode_generation_errors.items():
                    self.processing_thread.progress_signal.emit(f"      - '{order_id}': {error}")
            
            # Report order numbers not found in PDFs
            order_numbers_not_found = set(unique_order_numbers) - order_numbers_found_in_pdfs
            if order_numbers_not_found:
                self.processing_thread.progress_signal.emit("   ❌ Order Numbers Not Found in PDF Files:")
                for order_id in sorted(order_numbers_not_found):
                    self.processing_thread.progress_signal.emit(f"      - '{order_id}': No matching pages found in any PDF file")
            
            if not order_pages:
                self.processing_thread.progress_signal.emit("⚠️ No matching order numbers found in any PDF files!")
                self.processing_thread.progress_signal.emit("This could mean:")
                self.processing_thread.progress_signal.emit("   - Order numbers in PDFs don't match Excel file")
                self.processing_thread.progress_signal.emit("   - PDFs contain images that need OCR")
                self.processing_thread.progress_signal.emit("   - Text extraction failed from PDF pages")
            
            # Save generated barcodes to Supabase if available
            if SUPABASE_AVAILABLE:
                try:
                    self.processing_thread.progress_signal.emit("Saving barcodes to Supabase database...")
                    
                    # Prepare barcode data for Supabase
                    barcode_data_for_db = []
                    for order_id, pages in order_pages.items():
                        for page_info in pages:
                            if order_id in order_barcodes:
                                barcode_record = {
                                    'order_id': order_id,
                                    'driver_number': 'N/A',  # No driver assignment in this workflow
                                    'pdf_file_name': Path(page_info['source_file']).name,
                                    'page_number': page_info['page_num'] + 1,  # Convert to 1-based indexing
                                    'barcode_type': 'Code128'
                                }
                                barcode_data_for_db.append(barcode_record)
                    
                    # Save to Supabase
                    if barcode_data_for_db:
                        success = save_generated_barcodes(barcode_data_for_db)
                        if success:
                            self.processing_thread.progress_signal.emit(f"✅ Successfully saved {len(barcode_data_for_db)} barcodes to Supabase")
                        else:
                            self.processing_thread.progress_signal.emit("⚠️ Failed to save some barcodes to Supabase")
                    else:
                        self.processing_thread.progress_signal.emit("No barcodes to save to Supabase")
                        
                except Exception as e:
                    self.processing_thread.progress_signal.emit(f"⚠️ Error saving barcodes to Supabase: {str(e)}")
                    # Continue processing even if Supabase save fails
                    pass
            else:
                self.processing_thread.progress_signal.emit("⚠️ Supabase not available - barcodes not saved to database")
            
            # Modify original PDF files by adding barcodes to pages with matching order numbers
            self.processing_thread.progress_signal.emit("Adding barcodes to PDF pages with matching order numbers...")
            
            # Show summary of what was found
            total_matched_pages = sum(len(pages) for pages in order_pages.values())
            self.processing_thread.progress_signal.emit(f"Found {total_matched_pages} pages with matching Order IDs across {len(order_pages)} orders")
            self.processing_thread.progress_signal.emit("📋 Adding barcodes to pages with order IDs from Excel file")
            
            created_files = []
            failed_files = []
            
            if not order_pages:
                self.processing_thread.progress_signal.emit("No matching orders found in picking docket PDF files!")
                self.processing_thread.progress_signal.emit("Check that your picking docket PDF files contain order IDs that match those in your Excel file")
                return {
                    "processed_files": processed_files,
                    "total_pages": total_pages_processed,
                    "driver_files_created": 0,
                    "created_files": [],
                    "failed_files": [],
                    "driver_details": {},
                    "output_dir": str(output_dir),
                    "barcodes_generated": len(order_barcodes),
                    "error": "No matching orders found in picking docket PDF files"
                }
            
            # Group pages by source PDF file for processing
            pdf_files_to_modify = {}
            for order_id, pages in order_pages.items():
                for page_info in pages:
                    pdf_path = page_info['source_pdf_path']
                    if pdf_path not in pdf_files_to_modify:
                        pdf_files_to_modify[pdf_path] = []
                    pdf_files_to_modify[pdf_path].append({
                        'page_num': page_info['page_num'],
                        'order_id': order_id
                    })
            
            # Process each PDF file and add barcodes to matching pages
            for pdf_file, pages_to_modify in pdf_files_to_modify.items():
                try:
                    pdf_filename = Path(pdf_file).name
                    self.processing_thread.progress_signal.emit(f"Processing {pdf_filename} - adding barcodes to {len(pages_to_modify)} pages...")
                    
                    # Open the original PDF
                    pdf_document = fitz.open(pdf_file)
                    
                    # Sort pages by page number to process in order
                    pages_to_modify.sort(key=lambda x: x['page_num'])
                    
                    # Add barcodes to each matching page
                    for page_info in pages_to_modify:
                        page_num = page_info['page_num']
                        order_id = page_info['order_id']
                        
                        try:
                            # Get the page
                            page = pdf_document[page_num]
                            
                            # Add barcode at the top center of the page
                            if order_id in order_barcodes:
                                try:
                                    # Insert barcode image at the top center
                                    barcode_data = order_barcodes[order_id]
                                    
                                    # Calculate position for top center
                                    page_width = page.rect.width
                                    barcode_width = 700  # Long barcode
                                    barcode_height = 70  # Shorter barcode
                                    
                                    barcode_x = (page_width - barcode_width) / 2  # Center horizontally
                                    barcode_y = 20  # Top margin
                                    
                                    # Insert barcode image
                                    barcode_rect = fitz.Rect(barcode_x, barcode_y, barcode_x + barcode_width, barcode_y + barcode_height)
                                    page.insert_image(barcode_rect, stream=barcode_data)
                                    
                                    self.processing_thread.progress_signal.emit(
                                        f"Added barcode for Order {order_id} to page {page_num + 1} in {pdf_filename}"
                                    )
                                    
                                except Exception as barcode_error:
                                    self.processing_thread.progress_signal.emit(
                                        f"Error adding barcode to page {page_num + 1} for Order {order_id}: {str(barcode_error)}"
                                    )
                            
                        except Exception as e:
                            self.processing_thread.progress_signal.emit(
                                f"Error processing page {page_num + 1} for Order {order_id}: {str(e)}"
                            )
                            continue
                    
                    # Save the modified PDF to output directory
                    output_filename = f"Barcoded_{pdf_filename}"
                    output_path = output_dir / output_filename
                    
                    pdf_document.save(str(output_path))
                    pdf_document.close()
                    
                    # Verify the file was created
                    if output_path.exists():
                        created_files.append(output_filename)
                        self.processing_thread.progress_signal.emit(
                            f"✓ Successfully created {output_filename} with barcodes added"
                        )
                    else:
                        failed_files.append(output_filename)
                        self.processing_thread.progress_signal.emit(
                            f"✗ Failed to create {output_filename} - file not found after save"
                        )
                        
                except Exception as e:
                    failed_files.append(f"Barcoded_{Path(pdf_file).name}")
                    self.processing_thread.progress_signal.emit(
                        f"✗ Error processing PDF {Path(pdf_file).name}: {str(e)}"
                    )
                    if 'pdf_document' in locals():
                        pdf_document.close()
                    continue
            
            # Final summary message
            self.processing_thread.progress_signal.emit("Processing complete!")
            if SUPABASE_AVAILABLE:
                if getattr(self, 'selected_excel_file', ""):
                    self.processing_thread.progress_signal.emit(f"📤 Uploaded store orders to dispatch_orders table")
            self.processing_thread.progress_signal.emit(f"Created {len(created_files)} barcoded PDF files in {output_dir}")
            self.processing_thread.progress_signal.emit(f"📅 Files saved in date folder: {current_date}")
            self.processing_thread.progress_signal.emit(f"🏷️  Generated barcodes for {len(order_barcodes)} unique order numbers from Excel files")
            self.processing_thread.progress_signal.emit("📋 Added barcodes to pages with order IDs matching Excel file - other pages remain unchanged")
            
            # Generate summary report
            summary_path = output_dir / "picking_dockets_summary.txt"
            with open(summary_path, 'w', encoding='utf-8') as f:
                f.write("Dispatch Scanning Processing Summary\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Processing Date: {current_date}\n")
                f.write(f"Output Directory: {output_dir}\n")
                excel_file_path = getattr(self, 'selected_excel_file', "")
                if excel_file_path:
                    f.write(f"Store Orders Excel file: {Path(excel_file_path).name}\n")
                f.write(f"Database upload: {'✅ Success' if SUPABASE_AVAILABLE else '❌ Supabase not available'}\n")
                f.write(f"Unique order numbers from Excel files: {len(unique_order_numbers)}\n")
                f.write(f"Total picking docket PDF files processed: {processed_files}\n")
                f.write(f"Total pages scanned: {total_pages_processed}\n")
                f.write(f"Barcoded PDF files created: {len(created_files)}\n")
                f.write(f"Barcodes generated: {len(order_barcodes)}\n")
                f.write(f"Barcode generation failures: {len(barcode_generation_errors)}\n")
                f.write(f"Order numbers found in PDFs: {len(order_numbers_found_in_pdfs)}\n")
                f.write(f"Order numbers not found in PDFs: {len(set(unique_order_numbers) - order_numbers_found_in_pdfs)}\n")
                if failed_files:
                    f.write(f"Failed PDF files: {len(failed_files)}\n")
                f.write("\n")
                
                # Detailed barcode generation report
                f.write("Barcode Generation Details:\n")
                f.write("-" * 30 + "\n")
                if barcode_generation_errors:
                    f.write("Failed Barcode Generation:\n")
                    for order_id, error in barcode_generation_errors.items():
                        f.write(f"  - '{order_id}': {error}\n")
                    f.write("\n")
                
                # Order number matching report
                order_numbers_not_found = set(unique_order_numbers) - order_numbers_found_in_pdfs
                if order_numbers_not_found:
                    f.write("Order Numbers Not Found in PDF Files:\n")
                    for order_id in sorted(order_numbers_not_found):
                        f.write(f"  - '{order_id}': No matching pages found in any PDF file\n")
                    f.write("\n")
                
                f.write("Order Numbers Successfully Processed:\n")
                for order_id in sorted(order_numbers_found_in_pdfs):
                    page_count = len(order_pages.get(order_id, []))
                    f.write(f"  - '{order_id}': {page_count} pages found and barcoded\n")
                f.write("\n")
                f.write("WORKFLOW COMPLETED:\n")
                excel_file_path = getattr(self, 'selected_excel_file', "")
                if excel_file_path:
                    f.write(f"1. 📤 Uploaded store orders to dispatch_orders table (Excel row order preserved)\n")
                f.write(f"2. 🏷️  Generated barcodes for {len(order_barcodes)} unique order numbers from Excel files\n")
                f.write(f"3. 📄 Added barcodes to pages in original PDF files where order numbers were found\n")
                f.write(f"4. 📅 Organized all files in date folder: {current_date}\n\n")
                f.write("Each modified PDF contains the original pages with barcodes added at the top where order numbers were found.\n")
                f.write("Barcodes are generated for order numbers found in Excel Column A.\n\n")
                
                if created_files:
                    f.write("✓ Successfully Created Order PDF Files:\n")
                    for filename in created_files:
                        f.write(f"  - {filename}\n")
                    f.write("\n")
                
                if failed_files:
                    f.write("✗ Failed PDF Files:\n")
                    for filename in failed_files:
                        f.write(f"  - {filename}\n")
                    f.write("\n")
                
                f.write("Order Page Counts:\n")
                for order_id, pages in order_pages.items():
                    f.write(f"  Order {order_id}: {len(pages)} pages\n")
                
                f.write("\nExcel Order Numbers:\n")
                for order_id in sorted(unique_order_numbers):
                    found_pages = len(order_pages.get(order_id, []))
                    f.write(f"  - {order_id} ({found_pages} pages found)\n")
            
            # Collect order details for results dialog (convert to match expected format)
            order_details = {}
            for order_id, pages in order_pages.items():
                order_details[order_id] = {
                    'page_count': len(pages),
                    'orders': [order_id]  # Single order per group in this workflow
                }
            
            return {
                "processed_files": processed_files,
                "total_pages": total_pages_processed,
                "driver_files_created": len(created_files),
                "created_files": created_files,
                "failed_files": failed_files,
                "driver_details": order_details,  # Use order details instead of driver details
                "output_dir": str(output_dir),
                "barcodes_generated": len(order_barcodes),
                "barcode_generation_errors": barcode_generation_errors,
                "order_numbers_found_in_pdfs": list(order_numbers_found_in_pdfs),
                "order_numbers_not_found": list(set(unique_order_numbers) - order_numbers_found_in_pdfs),
                "database_upload": SUPABASE_AVAILABLE,
                "excel_file": Path(excel_file_path).name if excel_file_path else "None"
            }
            
        except Exception as e:
            self.processing_thread.progress_signal.emit(f"Error: {str(e)}")
            raise e
    
    def update_status(self, message):
        """Update the status bar message"""
        self.status_bar.showMessage(message)
    
    def show_progress(self, show=True):
        """Show or hide the progress bar"""
        self.progress_bar.setVisible(show)
        if show:
            self.progress_bar.setRange(0, 0)  # Indeterminate progress
        else:
            self.progress_bar.setRange(0, 1)
            self.progress_bar.setValue(1)
    
    def apply_clean_styling(self):
        """Apply professional, compact business styling"""
        self.setStyleSheet("""
            /* Main Application Styling */
            QMainWindow {
                background-color: #f5f5f5;
                font-family: 'Segoe UI', 'Arial', sans-serif;
            }
            
            /* Header Styling */
            QFrame#headerFrame {
                background-color: #3498db;
                border: none;
                border-radius: 0px;
                margin-bottom: 0px;
            }
            
            QLabel#headerTitle {
                color: white;
                font-size: 22px;
                font-weight: 600;
            }
            
            QLabel#headerSubtitle {
                color: #bdc3c7;
                font-size: 13px;
                font-weight: 400;
            }
            
            QLabel#headerIcon {
                font-size: 16px;
                color: white;
                background-color: #34495e;
                border-radius: 4px;
                padding: 4px;
                min-width: 24px;
                max-width: 24px;
                min-height: 24px;
                max-height: 24px;
            }
            
            QLabel#statusDot {
                color: #27ae60;
                font-size: 8px;
                background-color: #34495e;
                border-radius: 50%;
                padding: 2px;
                min-width: 12px;
                max-width: 12px;
                min-height: 12px;
                max-height: 12px;
            }
            
            QLabel#statusText {
                color: #bdc3c7;
                font-size: 11px;
                font-weight: 400;
            }
            
            /* Column Frames */
            QFrame#columnFrame {
                background-color: transparent;
                border: none;
            }
            
            /* Section Cards - Professional Design */
            QFrame#section {
                background-color: white;
                border: 1px solid #d5d5d5;
                border-radius: 4px;
                padding: 12px;
                margin-bottom: 8px;
            }
            
            /* Typography */
            QLabel {
                color: #2c3e50;
                font-size: 13px;
                font-weight: 400;
            }
            
            QLabel#sectionTitle {
                color: #2c3e50;
                font-size: 16px;
                font-weight: 600;
                margin-bottom: 4px;
            }
            
            QLabel#sectionSubtitle {
                color: #7f8c8d;
                font-size: 12px;
                font-style: normal;
                margin-bottom: 8px;
            }
            
            QLabel#sectionIcon {
                font-size: 12px;
                background-color: #3498db;
                color: white;
                border-radius: 3px;
                padding: 4px;
                min-width: 20px;
                max-width: 20px;
                min-height: 20px;
                max-height: 20px;
                text-align: center;
                qproperty-alignment: AlignCenter;
            }
            
            QLabel#workflowText {
                color: #2c3e50;
                font-size: 12px;
                line-height: 1.3;
                padding: 8px;
                background-color: #f8f9fa;
                border-radius: 3px;
                border-left: 3px solid #3498db;
                margin: 4px 0;
            }
            
            QLabel#requirementsText {
                color: #2c3e50;
                font-size: 12px;
                line-height: 1.3;
                padding: 8px;
                background-color: #f8f9fa;
                border-radius: 3px;
                border-left: 3px solid #3498db;
                margin: 4px 0;
            }
            
            QLabel#infoText {
                color: #7f8c8d;
                font-size: 12px;
                padding: 6px 8px;
                background-color: #f8f9fa;
                border-radius: 3px;
                border: 1px solid #e9ecef;
            }
            
            QLabel#warningText {
                color: #e67e22;
                font-size: 12px;
                padding: 6px 8px;
                background-color: #fef9e7;
                border-radius: 3px;
                font-weight: 500;
                border: 1px solid #f39c12;
            }
            
            QLabel#successText {
                color: #27ae60;
                font-size: 12px;
                padding: 6px 8px;
                background-color: #eafaf1;
                border-radius: 3px;
                font-weight: 500;
                border: 1px solid #2ecc71;
            }
            
            /* Button Styling - Professional */
            QPushButton {
                background-color: #ecf0f1;
                color: #2c3e50;
                border: 1px solid #bdc3c7;
                padding: 6px 12px;
                border-radius: 3px;
                font-weight: 500;
                font-size: 13px;
                min-height: 20px;
            }
            
            QPushButton:hover {
                background-color: #d5dbdb;
                border-color: #95a5a6;
            }
            
            QPushButton:pressed {
                background-color: #bdc3c7;
            }
            
            /* Primary Button */
            QPushButton#primaryButton {
                background-color: #3498db;
                color: white;
                border: 1px solid #2980b9;
                padding: 8px 16px;
                border-radius: 3px;
                font-weight: 600;
                font-size: 13px;
                min-height: 24px;
            }
            
            QPushButton#primaryButton:hover {
                background-color: #2980b9;
            }
            
            QPushButton#primaryButton:pressed {
                background-color: #21618c;
            }
            
            /* Secondary Button */
            QPushButton#secondaryButton {
                background-color: #95a5a6;
                color: white;
                border: 1px solid #7f8c8d;
            }
            
            QPushButton#secondaryButton:hover {
                background-color: #7f8c8d;
            }
            
            /* Input Fields */
            QLineEdit, QDateEdit {
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                padding: 6px 8px;
                background-color: white;
                color: #2c3e50;
                font-size: 13px;
            }
            
            QLineEdit:focus, QDateEdit:focus {
                border-color: #3498db;
                outline: none;
            }
            
            QLineEdit:hover, QDateEdit:hover {
                border-color: #95a5a6;
            }
            
            /* List and Table Widgets */
            QListWidget, QTableWidget {
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                color: #2c3e50;
                alternate-background-color: #f8f9fa;
                font-size: 13px;
            }
            
            QListWidget:focus, QTableWidget:focus {
                border-color: #3498db;
                outline: none;
            }
            
            /* Scrollbar Styling - Minimal */
            QScrollBar:vertical {
                background-color: #ecf0f1;
                width: 12px;
                border: none;
                border-radius: 6px;
            }
            
            QScrollBar::handle:vertical {
                background-color: #bdc3c7;
                border-radius: 6px;
                min-height: 20px;
            }
            
            QScrollBar::handle:vertical:hover {
                background-color: #95a5a6;
            }
            
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
                background: none;
            }
            
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                background: none;
            }
            
            QScrollBar:horizontal {
                background-color: #ecf0f1;
                height: 12px;
                border: none;
                border-radius: 6px;
            }
            
            QScrollBar::handle:horizontal {
                background-color: #bdc3c7;
                border-radius: 6px;
                min-width: 20px;
            }
            
            QScrollBar::handle:horizontal:hover {
                background-color: #95a5a6;
            }
            
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                width: 0px;
                background: none;
            }
            
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                background: none;
            }
            
            /* Table Styling */
            QTableWidget::item {
                padding: 6px 8px;
                border-bottom: 1px solid #ecf0f1;
                font-size: 13px;
            }
            
            QTableWidget::item:selected {
                background-color: #e3f2fd;
                color: #1976d2;
            }
            
            QTableWidget QHeaderView::section {
                background-color: #f8f9fa;
                border: none;
                border-bottom: 1px solid #bdc3c7;
                padding: 6px 8px;
                font-weight: 600;
                color: #2c3e50;
                font-size: 13px;
            }
            
            /* Status Bar */
            QStatusBar {
                background-color: #ecf0f1;
                border-top: 1px solid #bdc3c7;
                color: #2c3e50;
                font-size: 13px;
                font-weight: 400;
            }
            
            /* Progress Bar */
            QProgressBar {
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                text-align: center;
                background-color: white;
                color: #2c3e50;
                font-weight: 500;
                font-size: 13px;
            }
            
            QProgressBar::chunk {
                background-color: #3498db;
                border-radius: 2px;
            }
            
            /* Message Box Styling */
            QMessageBox {
                background-color: white;
                color: #2c3e50;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 16px;
                font-size: 12px;
            }
            
            QMessageBox QLabel {
                background-color: transparent;
                color: #2c3e50;
                font-size: 12px;
                padding: 8px;
                font-weight: 400;
            }
            
            QMessageBox QPushButton {
                background-color: #3498db;
                color: white;
                border: 1px solid #2980b9;
                padding: 8px 16px;
                border-radius: 3px;
                font-weight: 500;
                font-size: 11px;
                min-width: 60px;
            }
            
            QMessageBox QPushButton:hover {
                background-color: #2980b9;
            }
            
            QMessageBox QPushButton:pressed {
                background-color: #21618c;
            }
            
            /* Special Labels */
            QLabel#descriptionLabel {
                color: #7f8c8d;
                font-size: 12px;
                margin-bottom: 12px;
                line-height: 1.4;
            }
            
            QFrame#placeholderFrame {
                background-color: #f8f9fa;
                border: 1px dashed #bdc3c7;
                border-radius: 4px;
                padding: 24px;
                margin: 12px 0;
            }
            
            QLabel#placeholderLabel {
                color: #95a5a6;
                font-size: 12px;
                font-style: italic;
                font-weight: 400;
            }
            
            QFrame#infoFrame {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 15px;
            }
            
            QLabel#infoText {
                color: #475569;
                font-size: 13px;
                line-height: 1.5;
            }
            
            QPlainTextEdit#resultsArea {
                background-color: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                font-family: 'Consolas', monospace;
                font-size: 12px;
                color: #374151;
            }
            
            /* Order Management Tab Styling */
            QTabWidget#mainTabWidget {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
            }
            
            QTabWidget#mainTabWidget::pane {
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                background-color: white;
            }
            
            QTabWidget#mainTabWidget::tab-bar {
                alignment: left;
            }
            
            QTabWidget#mainTabWidget QTabBar::tab {
                background-color: #f8f9fa;
                color: #6b7280;
                border: 1px solid #e2e8f0;
                border-bottom: none;
                padding: 8px 16px;
                margin-right: 2px;
                border-top-left-radius: 6px;
                border-top-right-radius: 6px;
                font-weight: 500;
                font-size: 13px;
            }
            
            QTabWidget#mainTabWidget QTabBar::tab:selected {
                background-color: white;
                color: #2c3e50;
                border-bottom: 1px solid white;
                font-weight: 600;
            }
            
            QTabWidget#mainTabWidget QTabBar::tab:hover {
                background-color: #e9ecef;
                color: #495057;
            }
            
            QFrame#orderManagementHeader {
                background-color: #f8f9fa;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                margin-bottom: 12px;
            }
            
            QLabel#orderManagementTitle {
                font-size: 18px;
                font-weight: bold;
                color: #2c3e50;
            }
            
            QPushButton#refreshButton {
                background-color: #3498db;
                color: white;
                border: 1px solid #2980b9;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: 500;
                font-size: 12px;
            }
            
            QPushButton#refreshButton:hover {
                background-color: #2980b9;
            }
            
            QPushButton#refreshButton:pressed {
                background-color: #21618c;
            }
            
            QLabel#orderStatusLabel {
                color: #6b7280;
                font-size: 12px;
                font-style: italic;
            }
            
            QFrame#orderTableFrame {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
            }
            
            QTableWidget#orderTable {
                background-color: white;
                border: none;
                gridline-color: #e2e8f0;
                font-size: 12px;
            }
            
            QTableWidget#orderTable::item {
                padding: 8px 12px;
                border-bottom: 1px solid #f1f3f4;
            }
            
            QTableWidget#orderTable::item:first {
                border-left: none;
            }
            
            QTableWidget#orderTable::item:last {
                border-right: none;
            }
            
            QTableWidget#orderTable::item:selected {
                background-color: #e3f2fd;
                color: #1976d2;
            }
            
            QTableWidget#orderTable QHeaderView::section {
                background-color: #f8f9fa;
                border: none;
                border-bottom: 2px solid #e2e8f0;
                border-right: 1px solid #e2e8f0;
                padding: 8px 12px;
                font-weight: 600;
                color: #2c3e50;
                font-size: 12px;
            }
            
            QTableWidget#orderTable QHeaderView::section:first {
                border-top-left-radius: 6px;
            }
            
            QTableWidget#orderTable QHeaderView::section:last {
                border-top-right-radius: 6px;
                border-right: none;
            }
            
            /* Label Printing Tab Styling */
            QFrame#labelPrintingHeader {
                background-color: #f8f9fa;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                margin-bottom: 12px;
            }
            
            QLabel#labelPrintingTitle {
                font-size: 14px;
                font-weight: bold;
                color: #2c3e50;
            }
            
            QLabel#printerStatusLabel {
                color: #e74c3c;
                font-weight: bold;
                font-size: 12px;
            }
            
            QGroupBox#scannerGroup, QGroupBox#settingsGroup, QGroupBox#printerGroup, QGroupBox#logGroup {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                margin-top: 8px;
                padding-top: 8px;
                font-weight: 600;
                color: #2c3e50;
            }
            
            QGroupBox#scannerGroup::title, QGroupBox#settingsGroup::title, 
            QGroupBox#printerGroup::title, QGroupBox#logGroup::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                background-color: white;
            }
            
            QLineEdit#orderNumberInput {
                background-color: white;
                border: 2px solid #e2e8f0;
                border-radius: 4px;
                padding: 8px 12px;
                font-size: 14px;
                font-family: 'Consolas', monospace;
            }
            
            QLineEdit#orderNumberInput:focus {
                border-color: #3498db;
                background-color: #f8f9fa;
            }
            
            QLabel#orderInfoLabel {
                background-color: #f8f9fa;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                padding: 8px;
                min-height: 60px;
                font-family: 'Consolas', monospace;
                font-size: 11px;
            }
            
            QComboBox#crateCountSpinbox {
                background-color: white;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                padding: 6px 12px;
                font-size: 14px;
                min-width: 80px;
            }
            
            QComboBox#crateCountSpinbox:focus {
                border-color: #3498db;
            }
            
            QPushButton#printButton {
                background-color: #27ae60;
                color: white;
                border: 1px solid #229954;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: 600;
                font-size: 14px;
            }
            
            QPushButton#printButton:hover {
                background-color: #229954;
            }
            
            QPushButton#printButton:pressed {
                background-color: #1e8449;
            }
            
            QPushButton#printButton:disabled {
                background-color: #bdc3c7;
                border-color: #95a5a6;
                color: #7f8c8d;
            }
            
            QPushButton#previewButton {
                background-color: #f39c12;
                color: white;
                border: 1px solid #e67e22;
                padding: 10px 20px;
                border-radius: 4px;
                font-weight: 600;
                font-size: 14px;
            }
            
            QPushButton#previewButton:hover {
                background-color: #e67e22;
            }
            
            QPushButton#previewButton:pressed {
                background-color: #d35400;
            }
            
            QPushButton#previewButton:disabled {
                background-color: #bdc3c7;
                border-color: #95a5a6;
                color: #7f8c8d;
            }
            
            QPushButton#connectPrinterButton {
                background-color: #3498db;
                color: white;
                border: 1px solid #2980b9;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: 500;
                font-size: 12px;
            }
            
            QPushButton#connectPrinterButton:hover {
                background-color: #2980b9;
            }
            
            QPushButton#connectPrinterButton:pressed {
                background-color: #21618c;
            }
            
            QLabel#printerInfoLabel {
                background-color: #f8f9fa;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                padding: 8px;
                min-height: 40px;
                font-family: 'Consolas', monospace;
                font-size: 11px;
            }
            
            QTextEdit#printLog {
                background-color: #f8f9fa;
                border: 1px solid #e2e8f0;
                border-radius: 4px;
                font-family: 'Consolas', monospace;
                font-size: 11px;
            }
            
            /* Crate Count Dialog Styling */
            QPushButton#cancelButton {
                background-color: #95a5a6;
                color: white;
                border: 1px solid #7f8c8d;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: 500;
                font-size: 12px;
            }
            
            QPushButton#cancelButton:hover {
                background-color: #7f8c8d;
            }
            
            QPushButton#printDialogButton {
                background-color: #27ae60;
                color: white;
                border: 1px solid #229954;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: 600;
                font-size: 12px;
            }
            
            QPushButton#printDialogButton:hover {
                background-color: #229954;
            }
            
            QPushButton#printDialogButton:pressed {
                background-color: #1e8449;
            }
        """)


def main():
    app = QApplication(sys.argv)
    window = DispatchScanningApp()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main() 